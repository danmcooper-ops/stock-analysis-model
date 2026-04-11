# scripts/report_excel.py
def build_excel(rows, filename):
    """Generate a multi-sheet Excel workbook matching the reference format.

    Produces three tabs — Analysis (flat), Ownership (sector-grouped),
    Company (sector-grouped) — with full-row rating colors, sector header
    rows, and per-sheet freeze panes.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from collections import defaultdict, OrderedDict

    wb = Workbook()

    # --- Derive peers (same-industry tickers) ---
    industry_map = defaultdict(list)
    for r in rows:
        ind = r.get('industry')
        if ind:
            industry_map[ind].append(r['ticker'])
    for r in rows:
        ind = r.get('industry')
        peers = [t for t in industry_map.get(ind, []) if t != r['ticker']]
        r['_peers'] = ', '.join(peers[:8]) if peers else ''

    # --- Pre-compute derived fields ---
    for r in rows:
        r['_mcap_b'] = r['mcap'] / 1e9 if r.get('mcap') else None
        r['_description'] = r.get('description_full') or r.get('description') or ''
        r['_shares_out_m'] = r['shares_out'] / 1e6 if r.get('shares_out') else None
        r['_float_m'] = r['float_shares'] / 1e6 if r.get('float_shares') else None
        r['_short_m'] = r['shares_short'] / 1e6 if r.get('shares_short') else None
        r['_fcf_m'] = r['fcf'] / 1e6 if r.get('fcf') else None
        r['_insider_net_value_m'] = r['insider_net_value'] / 1e6 if r.get('insider_net_value') else None
        r['_founder_led'] = 'Yes' if r.get('founder_led') else 'No'
        p, fv = r.get('price'), r.get('dcf_fv')
        r['_price_fv'] = p / fv if p and fv else None
        sens = r.get('dcf_sens_range')
        fv = r.get('dcf_fv')
        if sens and len(sens) == 2:
            r['_dcf_bear'] = sens[0]
            r['_dcf_bull'] = sens[1]
        else:
            r['_dcf_bear'] = fv * 0.70 if fv is not None else None
            r['_dcf_bull'] = fv * 1.35 if fv is not None else None

    # --- Column definitions per sheet (tab order: Analysis, Ownership, Company) ---
    sheets_config = OrderedDict([
        ('Analysis', [
            ('Ticker',           'ticker',            '@'),
            ('Rating',           'rating',            '@'),
            ('Analyst Rec',      'analyst_rec',       '@'),
            ('Gates Passed',     '_gates_passed',     '@'),
            ('Sector',           'sector',            '@'),
            ('Mkt Cap ($B)',     '_mcap_b',           '#,##0.0'),
            ('Last Price',       'price',             '"$"#,##0.00'),
            ('Model Fair Value', 'dcf_fv',            '"$"#,##0.00'),
            ('Model Price/FV',   '_price_fv',         '0.00'),
            ('MS Fair Value',    'ms_fv',             '"$"#,##0.00'),
            ('MS Price/FV',      'ms_pfv',            '0.00'),
            ('vs MS',            'ms_diff',           '"+"0.0%;"--"0.0%'),
            ('Net Debt/EBITDA',  'nd_ebitda',         '0.0'),
            ('EV/EBITDA',        'ev_ebitda',         '0.0'),
            ('Sector Med EV/EBITDA', '_sector_median_ee', '0.0'),
            ('vs Sector',            '_ee_vs_sector',     '"+"0%;"--"0%'),
            ('Current Ratio',    'cr',                '0.00'),
            ('D/E',              'de',                '0.0'),
            ('Rev CAGR (5y)',    'rev_cagr',          '0.0%'),
            ('FCF Growth',       'fcf_growth',        '0.0%'),
            ('Analyst LTG',      'analyst_ltg',       '0.0%'),
            ('Margin Trend',     'margin_trend',      '+0.0%;-0.0%'),
            ('Surprise Avg',     'surprise_avg',      '+0.0%;-0.0%'),
            ('Fund. Growth',     'fundamental_growth', '0.0%'),
            ('Reinvest Rate',    'reinvestment_rate', '0.0%'),
            ('FCF ($M)',         '_fcf_m',            '#,##0.0'),
            ('Piotroski F',      'piotroski',         '0'),
            ('P/FCF',            'pfcf',              '0.0'),
            ('P/E',              'pe',                '0.0'),
            ('P/B',              'pb',                '0.0'),
            ('ROE',              'roe',               '0.0%'),
            ('ROA',              'roa',               '0.0%'),
            ('ROIC',             'roic',              '0.0%'),
            ('WACC',             'wacc',              '0.0%'),
            ('ROIC-WACC Spread', 'spread',            '0.0%'),
            ('Terminal Growth',  'terminal_growth',   '0.0%'),
            ('Exit Mult FV',     'exit_mult_fv',      '"$"#,##0.00'),
            ('TV Spread',        'tv_method_spread',  '0.0%'),
            ('MoS',              'mos',               '0.0%'),
            ('Cash Conv',        'cash_conv',         '0.0'),
            ('Accruals',         'accruals',          '0.00'),
            ('Interest Cov',     'int_cov',           '0.0'),
            ('DCF Bear',         '_dcf_bear',         '"$"#,##0.00'),
            ('DCF Bull',         '_dcf_bull',         '"$"#,##0.00'),
            ('MC Confidence',    'mc_confidence',     '@'),
            ('MC CV',            'mc_cv',             '0.0%'),
            ('DDM Eligible',     'ddm_eligible',      '@'),
            ('DDM FV',           'ddm_fv',            '"$"#,##0.00'),
            ('DDM Growth',       'ddm_growth',        '0.0%'),
            ('DDM Div CAGR',     'ddm_div_cagr',      '0.0%'),
            ('DDM Sust. Growth', 'ddm_sustainable_growth', '0.0%'),
            ('DDM Consec. Yrs',  'ddm_consecutive_years', '0'),
            ('DDM Payout Flag',  'ddm_payout_flag',   '@'),
            ('Blend Method',     '_blended_method',   '@'),
            # Reverse DCF
            ('Impl. Growth',     'implied_growth',    '0.0%'),
            ('Impl. vs Est.',    'implied_vs_estimated', '+0.0%;-0.0%'),
            # EPV
            ('EPV FV',           'epv_fv',            '"$"#,##0.00'),
            ('EPV MoS',          'epv_mos',           '0.0%'),
            ('EPV Growth FV',    'epv_growth_fv',     '"$"#,##0.00'),
            # RIM
            ('RIM FV',           'rim_fv',            '"$"#,##0.00'),
            ('RIM MoS',          'rim_mos',           '0.0%'),
            # Quality/Risk
            ('Altman Z',         'altman_z',          '0.00'),
            ('Z Zone',           'altman_z_zone',     '@'),
            ('Beneish M',        'beneish_m',         '0.00'),
            ('Manip. Flag',      'beneish_flag',      '@'),
            # DuPont
            ('DuPont Margin',    'dupont_margin',     '0.0%'),
            ('DuPont Turnover',  'dupont_turnover',   '0.00'),
            ('DuPont Leverage',  'dupont_leverage',   '0.00'),
            # 52-week range
            ('52W High',         'high_52w',          '"$"#,##0.00'),
            ('52W Low',          'low_52w',           '"$"#,##0.00'),
            ('% from 52W High',  'pct_from_52w_high', '+0.0%;-0.0%'),
            ('52W Range Pos',    'range_52w_position', '0'),
            # Portfolio
            ('Position Weight',  'position_weight',   '0.0%'),
            ('Composite Score',  '_composite_score',  '0"%"'),
            ('WS Target Low',    'target_low',        '"$"#,##0.00'),
            ('WS Target Mean',   'target_mean',       '"$"#,##0.00'),
            ('WS Target High',   'target_high',       '"$"#,##0.00'),
            ('# Analysts',       'num_analysts',      '0'),
        ]),
        ('Ownership', [
            ('Ticker',            'ticker',                  '@'),
            ('Rating',            'rating',                  '@'),
            ('Gates Passed',      '_gates_passed',           '@'),
            ('Sector',            'sector',                  '@'),
            ('Mkt Cap ($B)',      '_mcap_b',                 '#,##0.0'),
            ('Shares Out (M)',    '_shares_out_m',           '#,##0.0'),
            ('Float (M)',         '_float_m',                '#,##0.0'),
            ('Insider %',         'insider_pct',             '0.00%'),
            ('Institutional %',   'inst_pct',                '0.00%'),
            ('Shares Short (M)',  '_short_m',                '#,##0.0'),
            ('Short Ratio',       'short_ratio',             '0.00'),
            ('Short % Float',     'short_pct_float',         '0.00%'),
            ('Share Turnover',    'share_turnover_rate',     '0.00%'),
            ('Buyback Rate',      'share_buyback_rate',      '0.00%'),
            ('Shrhldr Yield',     'shareholder_yield',       '0.00%'),
            ('Div Yield',         'div_yield',               '0.00%'),
            ('Payout Ratio',      'payout_ratio',            '0.0%'),
            ('Ins Buy Ratio',     'insider_buy_ratio',       '0%'),
            ('Ins Buys 90d',      'insider_buy_count_90d',   '0'),
            ('Ins Sells 90d',     'insider_sell_count_90d',  '0'),
            ('Ins Buys 1y',       'insider_buy_count_365d',  '0'),
            ('Ins Sells 1y',      'insider_sell_count_365d', '0'),
            ('Net Ins Val ($M)',  '_insider_net_value_m',    '#,##0.0'),
        ]),
        ('Company', [
            ('Ticker',       'ticker',        None),
            ('Gates Passed', '_gates_passed', None),
            ('CEO',          'ceo_bio',       None),
            ('Founder-Led', '_founder_led',  None),
            ('Sector',      'sector',        None),
            ('Industry',    'industry',      None),
            ('Peers',       '_peers',        None),
            ('Description', '_description',  None),
        ]),
        ('Matrix', [
            # --- Fixed columns ---
            ('Ticker',          'ticker',           '@'),
            ('Raw Rating',      'rating_raw',       '@'),
            ('Final Rating',    'rating',           '@'),
            ('Composite',       '_composite_score', '0"%"'),
            ('Gates Passed',    '_gates_passed',    '@'),
            ('Sector',          'sector',           '@'),
            # --- Valuation (15%) ---
            ('MoS',             '_gate_mos',        '0.0%'),
            ('MoS Score',       '_score_mos',       '0"%"'),
            ('Price/FV',        '_gate_price_fv',   '0.00'),
            ('P/FV Score',      '_score_price_fv',  '0"%"'),
            ('EPV P/FV',        '_gate_epv_p_fv',   '0.00'),
            ('EPV P/FV Score',  '_score_epv_p_fv',  '0"%"'),
            ('Val Total',       '_score_valuation', '0"%"'),
            # --- Quality (10%) ---
            ('Piotroski',       '_gate_piotroski',  '0'),
            ('Piotroski Score', '_score_piotroski', '0"%"'),
            ('Int Cov',         '_gate_int_coverage', '0.00'),
            ('Int Cov Score',   '_score_int_coverage', '0"%"'),
            ('Accruals',        '_gate_accruals',   '0.0%'),
            ('Accruals Score',  '_score_accruals',  '0"%"'),
            ('Qual Total',      '_score_quality',   '0"%"'),
            # --- Moat (35%) ---
            ('ROIC CV',         '_gate_roic_consistency', '0.0%'),
            ('ROIC CV Score',   '_score_roic_consistency', '0"%"'),
            ('Spread > 5%',     '_gate_spread_>_5%', '0.0%'),
            ('Spread Score',    '_score_spread',    '0"%"'),
            ('Gross Margin',    '_gate_gross_margin', '0.0%'),
            ('Gross Mgn Score', '_score_gross_margin', '0"%"'),
            ('Moat Total',      '_score_moat',      '0"%"'),
            # --- Growth (30%) ---
            ('Fund Growth',     '_gate_fund_growth', '0.0%'),
            ('FG Score',        '_score_fund_growth', '0"%"'),
            ('Margins',         '_gate_margins',    '0.0%'),
            ('Margins Score',   '_score_margins',   '0"%"'),
            ('Surprise',        '_gate_surprise',   '0.0%'),
            ('Surprise Score',  '_score_surprise',  '0"%"'),
            ('Profit Pool',     '_gate_profit_pool', '0.00'),
            ('PP Score',        '_score_profit_pool', '0"%"'),
            ('Growth Total',    '_score_growth',    '0"%"'),
            # --- Ownership (10%) ---
            ('Shrhldr Yld',     '_gate_shrhldr_yield',  '0.0%'),
            ('Shrhldr Score',   '_score_shrhldr_yield', '0"%"'),
            ('Insider %',       '_gate_insider_own',    '0.0%'),
            ('Ins Own Score',   '_score_insider_own',   '0"%"'),
            ('Turnover',        '_gate_turnover',       '0.0%'),
            ('Turn Score',      '_score_turnover',      '0"%"'),
            ('Buyback',         '_gate_buyback_rate',   '0.0%'),
            ('Buyback Score',   '_score_buyback_rate',  '0"%"'),
            ('Ins Buying',      '_gate_insider_buying', '0%'),
            ('Ins Buy Score',   '_score_insider_buying', '0"%"'),
            ('Own Total',       '_score_ownership',     '0"%"'),
        ]),
    ])

    # --- Styles: white cells, gray headers ---
    gray_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    header_font = Font(bold=True, color='000000', size=11)
    header_fill = gray_fill
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    row_header_fill = gray_fill  # frozen / row-identifier columns
    row_header_font = Font(bold=True, color='000000')
    data_font = Font(color='000000')
    sector_font = Font(bold=True, size=11)
    gate_pass_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    gate_fail_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    gate_na_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    gate_yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')

    # Column widths per sheet (from reference file)
    col_widths = {
        'Analysis': {
            'ticker': 8, 'rating': 9, 'analyst_rec': 13, '_gates_passed': 14,
            'sector': 22, '_mcap_b': 15, 'price': 12, 'dcf_fv': 19,
            '_price_fv': 15, 'ms_fv': 15, 'ms_pfv': 13, 'ms_diff': 22,
            'nd_ebitda': 17, 'ev_ebitda': 11, '_sector_median_ee': 22,
            '_ee_vs_sector': 12, 'cr': 20, 'de': 8, 'rev_cagr': 22,
            'fcf_growth': 20, 'analyst_ltg': 13, 'margin_trend': 14,
            'surprise_avg': 13, 'fundamental_growth': 14, 'reinvestment_rate': 14,
            '_fcf_m': 10, 'piotroski': 13, 'pfcf': 19,
            'pe': 11, 'pb': 8, 'roe': 20, 'roa': 21, 'roic': 20,
            'wacc': 10, 'spread': 21, 'terminal_growth': 17,
            'exit_mult_fv': 13, 'tv_method_spread': 12,
            'mos': 22, 'cash_conv': 19, 'accruals': 22, 'int_cov': 19,
            '_dcf_bear': 11, '_dcf_bull': 11, 'mc_confidence': 14,
            '_composite_score': 16, 'target_low': 15,
            'target_mean': 16, 'target_high': 15, 'num_analysts': 12,
        },
        'Ownership': {
            'ticker': 8, 'rating': 9, '_gates_passed': 14, 'sector': 22,
            '_mcap_b': 15, '_shares_out_m': 16, '_float_m': 13,
            'insider_pct': 14, 'inst_pct': 17, '_short_m': 18,
            'short_ratio': 13, 'short_pct_float': 15,
            'share_turnover_rate': 15, 'share_buyback_rate': 14,
            'shareholder_yield': 14, 'div_yield': 11, 'payout_ratio': 14,
            'insider_buy_ratio': 14, 'insider_buy_count_90d': 14,
            'insider_sell_count_90d': 15, 'insider_buy_count_365d': 13,
            'insider_sell_count_365d': 14, '_insider_net_value_m': 16,
        },
        'Company': {
            'ticker': 8, '_gates_passed': 14, 'ceo_bio': 40,
            '_founder_led': 13, 'sector': 22, 'industry': 22,
            '_peers': 40, '_description': 134,
        },
        'Matrix': {
            'ticker': 8, 'rating_raw': 12, 'rating': 14,
            '_composite_score': 12, '_gates_passed': 11, 'sector': 22,
            '_gate_mos': 8, '_score_mos': 10,
            '_gate_price_fv': 10, '_score_price_fv': 10,
            '_gate_epv_p_fv': 10, '_score_epv_p_fv': 12,
            '_score_valuation': 10,
            '_gate_piotroski': 11, '_score_piotroski': 14,
            '_gate_int_coverage': 11, '_score_int_coverage': 14,
            '_gate_accruals': 10, '_score_accruals': 13,
            '_score_quality': 10,
            '_gate_roic_consistency': 13, '_score_roic_consistency': 14,
            '_gate_spread_>_5%': 12, '_score_spread': 12,
            '_gate_gross_margin': 13, '_score_gross_margin': 14,
            '_score_moat': 10,
            '_gate_fund_growth': 12, '_score_fund_growth': 10,
            '_gate_margins': 10, '_score_margins': 13,
            '_gate_surprise': 10, '_score_surprise': 13,
            '_gate_profit_pool': 12, '_score_profit_pool': 10,
            '_score_growth': 12,
            '_gate_shrhldr_yield': 12, '_score_shrhldr_yield': 14,
            '_gate_insider_own': 11, '_score_insider_own': 13,
            '_gate_turnover': 11, '_score_turnover': 12,
            '_gate_buyback_rate': 11, '_score_buyback_rate': 14,
            '_gate_insider_buying': 12, '_score_insider_buying': 13,
            '_score_ownership': 11,
        },
    }

    # Frozen columns per sheet (these get row-header gray styling)
    frozen_cols = {'Analysis': 4, 'Ownership': 4, 'Company': 2, 'Matrix': 6}

    freeze_config = {'Analysis': 'E2', 'Ownership': 'E2', 'Company': 'C2', 'Matrix': 'G4'}

    # Conditionally add Source column when validation data is present
    _has_validation = any(r.get('source_group') == 'poor' for r in rows)
    if _has_validation:
        _src_col = ('Source', 'source_group', '@')
        for _sname in ('Analysis', 'Matrix'):
            _cols = sheets_config[_sname]
            _cols.insert(1, _src_col)  # insert after Ticker
            col_widths[_sname]['source_group'] = 10
        frozen_cols['Analysis'] = 5
        frozen_cols['Matrix'] = 7
        freeze_config['Analysis'] = 'F2'
        freeze_config['Matrix'] = 'H4'

    SECTOR_ORDER = [
        'Technology', 'Communication Services', 'Consumer Cyclical',
        'Healthcare', 'Basic Materials', 'Industrials',
        'Energy', 'Consumer Defensive',
    ]

    def _group_by_sector(data_rows):
        by_sector = defaultdict(list)
        for r in data_rows:
            by_sector[r.get('sector') or 'Other'].append(r)
        grouped = OrderedDict()
        for s in SECTOR_ORDER:
            if s in by_sector:
                grouped[s] = by_sector.pop(s)
        for s in sorted(by_sector):
            grouped[s] = by_sector[s]
        return grouped

    sector_grouped = set()  # all sheets are flat tables

    first = True
    for sheet_name, cols in sheets_config.items():
        ws = wb.active if first else wb.create_sheet()
        first = False
        ws.title = sheet_name

        # Header row (Matrix has its own multi-level header logic)
        if sheet_name != 'Matrix':
            for ci, (label, _, _) in enumerate(cols, 1):
                cell = ws.cell(row=1, column=ci, value=label)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align

        n_frozen = frozen_cols.get(sheet_name, 0)
        widths = col_widths.get(sheet_name, {})

        if sheet_name in sector_grouped:
            # --- Sector-grouped sheet ---
            groups = _group_by_sector(rows)
            current_row = 2
            for sector_name, sector_rows in groups.items():
                cell = ws.cell(row=current_row, column=1, value=sector_name)
                cell.font = sector_font
                current_row += 1
                for row in sector_rows:
                    for ci, (_, key, fmt) in enumerate(cols, 1):
                        if ci == 1:
                            continue
                        val = row.get(key)
                        cell = ws.cell(row=current_row, column=ci, value=val)
                        if fmt:
                            cell.number_format = fmt
                        if key.startswith('_gate_'):
                            gp = row.get('_gp' + key[5:])
                            if gp is True:
                                cell.fill = gate_pass_fill
                            elif gp is False:
                                cell.fill = gate_fail_fill
                            else:
                                cell.fill = gate_na_fill
                            cell.alignment = Alignment(horizontal='center')
                        elif key.startswith('_score_') and isinstance(val, (int, float)):
                            if val >= 60:
                                cell.fill = gate_pass_fill
                            elif val >= 30:
                                cell.fill = gate_yellow_fill
                            else:
                                cell.fill = gate_fail_fill
                            cell.alignment = Alignment(horizontal='center')
                        elif ci <= n_frozen:
                            cell.fill = row_header_fill
                            cell.font = row_header_font
                        else:
                            cell.fill = white_fill
                            cell.font = data_font
                        if key in ('_description', 'ceo_bio', '_peers') and val:
                            cell.alignment = Alignment(wrap_text=True, vertical='top')
                    current_row += 1
        else:
            # --- Flat sheet ---
            data_start = 2

            # Matrix sheet: multi-level headers (category row + column row + thresholds)
            if sheet_name == 'Matrix':
                # -- Key → column-index mapping --
                _key_to_col = {key: ci for ci, (_, key, _) in enumerate(cols, 1)}

                # Category groups: (label, first_key, last_key, dark_hex, light_hex)
                _cat_groups = [
                    ('VALUATION (15%)', '_gate_mos',           '_score_valuation',
                     '2F5496', 'D6E4F0'),
                    ('QUALITY (10%)',   '_gate_piotroski',     '_score_quality',
                     '548235', 'E2EFDA'),
                    ('MOAT (35%)',      '_gate_roic_consistency', '_score_moat',
                     'C55A11', 'FCE4CC'),
                    ('GROWTH (30%)',    '_gate_fund_growth',   '_score_growth',
                     '7030A0', 'E4CCEF'),
                    ('OWNERSHIP (10%)', '_gate_shrhldr_yield', '_score_ownership',
                     'BF8F00', 'FFF2CC'),
                ]
                _cat_font = Font(bold=True, color='FFFFFF', size=11)
                _cat_align = Alignment(horizontal='center', vertical='center')

                # Build column → light-fill mapping for row 2 sub-headers
                _col_light_fill = {}
                for _, fk, lk, _, lt in _cat_groups:
                    lf = PatternFill(start_color=lt, end_color=lt, fill_type='solid')
                    for c in range(_key_to_col[fk], _key_to_col[lk] + 1):
                        _col_light_fill[c] = lf

                # Row 1 — frozen columns: label merged into rows 1-2
                for ci in range(1, n_frozen + 1):
                    for r in (1, 2):
                        c = ws.cell(row=r, column=ci)
                        c.fill = header_fill
                        c.font = header_font
                        c.alignment = header_align
                    ws.cell(row=1, column=ci, value=cols[ci - 1][0])
                    ws.merge_cells(start_row=1, start_column=ci,
                                   end_row=2, end_column=ci)

                # Row 1 — category merged headers (each its own colour)
                for cat_label, first_key, last_key, dk, _ in _cat_groups:
                    sc = _key_to_col[first_key]
                    ec = _key_to_col[last_key]
                    dk_fill = PatternFill(start_color=dk, end_color=dk,
                                          fill_type='solid')
                    # Pre-fill every cell so merged range has background
                    for c in range(sc, ec + 1):
                        ws.cell(row=1, column=c).fill = dk_fill
                    ws.merge_cells(start_row=1, start_column=sc,
                                   end_row=1, end_column=ec)
                    cell = ws.cell(row=1, column=sc, value=cat_label)
                    cell.font = _cat_font
                    cell.fill = dk_fill
                    cell.alignment = _cat_align

                # Row 2 — gate-level column headers (tinted per category)
                for ci in range(n_frozen + 1, len(cols) + 1):
                    cell = ws.cell(row=2, column=ci, value=cols[ci - 1][0])
                    cell.font = header_font
                    cell.fill = _col_light_fill.get(ci, header_fill)
                    cell.alignment = header_align

                # Row 3 — threshold descriptions
                gate_thresholds = {
                    '_gate_mos':            'MoS > 0%',
                    '_gate_price_fv':       'Price/FV < 1.2',
                    '_gate_epv_p_fv':       'EPV P/FV < 1.2',
                    '_gate_piotroski':      'F-Score ≥ 5',
                    '_gate_int_coverage':   'IC > 3×',
                    '_gate_accruals':       '|Accruals| < 8%',
                    '_gate_roic_consistency': 'CV < 30%',
                    '_gate_spread_>_5%':    'Spread > 5%',
                    '_gate_gross_margin':   'GM > 25%',
                    '_gate_fund_growth':    'FG > 3%',
                    '_gate_margins':        'Margin ≥ 0',
                    '_gate_surprise':       'Surprise > 0',
                    '_gate_profit_pool':    'PP ≥ 1.0×',
                    '_gate_shrhldr_yield':  'Yield > 0%',
                    '_gate_insider_own':    'Insider ≥ 5%',
                    '_gate_turnover':       'Turnover < 2%',
                    '_gate_buyback_rate':   'Buyback > 0%',
                    '_gate_insider_buying': 'Buy Ratio > 50%',
                }
                desc_font = Font(italic=True, color='666666', size=10)
                for ci, (_, key, _) in enumerate(cols, 1):
                    desc = gate_thresholds.get(key, '')
                    cell = ws.cell(row=3, column=ci, value=desc)
                    cell.font = desc_font
                    cell.fill = white_fill
                    cell.alignment = Alignment(horizontal='center')
                data_start = 4

            for ri, row in enumerate(rows, data_start):
                for ci, (_, key, fmt) in enumerate(cols, 1):
                    val = row.get(key)
                    cell = ws.cell(row=ri, column=ci, value=val)
                    if fmt:
                        cell.number_format = fmt
                    # Gate cells: colour by pass/fail boolean
                    if key.startswith('_gate_'):
                        gp = row.get('_gp' + key[5:])
                        if gp is True:
                            cell.fill = gate_pass_fill
                        elif gp is False:
                            cell.fill = gate_fail_fill
                        else:
                            cell.fill = gate_na_fill
                        cell.alignment = Alignment(horizontal='center')
                    elif key.startswith('_score_') and isinstance(val, (int, float)):
                        if val >= 60:
                            cell.fill = gate_pass_fill
                        elif val >= 30:
                            cell.fill = gate_yellow_fill
                        else:
                            cell.fill = gate_fail_fill
                        cell.alignment = Alignment(horizontal='center')
                    elif ci <= n_frozen:
                        cell.fill = row_header_fill
                        cell.font = row_header_font
                    else:
                        cell.fill = white_fill
                        cell.font = data_font
                    if key in ('_description', 'ceo_bio', '_peers') and val:
                        cell.alignment = Alignment(wrap_text=True, vertical='top')

        # Fixed column widths
        for ci, (label, key, _) in enumerate(cols, 1):
            w = widths.get(key, len(label) + 2)
            ws.column_dimensions[get_column_letter(ci)].width = w

        # Row heights
        if sheet_name == 'Matrix':
            ws.row_dimensions[1].height = 20   # category header
            ws.row_dimensions[2].height = 26   # column headers
            for ri in range(3, ws.max_row + 1):
                ws.row_dimensions[ri].height = 13
        else:
            ws.row_dimensions[1].height = 13
            for ri in range(2, ws.max_row + 1):
                ws.row_dimensions[ri].height = 13

        ws.freeze_panes = freeze_config.get(sheet_name, 'A2')

    # --- Profit Pool tab ---
    pp_ws = wb.create_sheet('Profit Pool')

    # Pre-compute display fields
    def _pp_strength(mult):
        if mult is None:    return ''
        if mult >= 2.0:     return 'Dominant'
        if mult >= 1.5:     return 'Strong'
        if mult >= 1.0:     return 'Above Avg'
        if mult >= 0.5:     return 'Below Avg'
        return 'Weak'

    def _hhi_label(hhi):
        if hhi is None:     return ''
        if hhi > 0.25:      return 'Concentrated'
        if hhi > 0.15:      return 'Moderate'
        return 'Competitive'

    # Pre-compute sector ranks and profit share gap
    from collections import defaultdict
    _sector_pp = defaultdict(list)
    for r in rows:
        if r.get('pp_multiple') is not None:
            _sector_pp[r.get('sector') or ''].append(r)
    for sector_rows_pp in _sector_pp.values():
        sector_rows_pp.sort(key=lambda r: r.get('pp_multiple') or 0, reverse=True)
        for rank_i, r in enumerate(sector_rows_pp, 1):
            r['_pp_sector_rank'] = rank_i

    for r in rows:
        r['_pp_rev_b']    = r['revenue'] / 1e9 if r.get('revenue') else None
        r['_pp_oi_m']     = r['operating_income'] / 1e6 if r.get('operating_income') else None
        r['_pp_strength'] = _pp_strength(r.get('pp_multiple'))
        r['_hhi_label']   = _hhi_label(r.get('pp_sector_hhi'))
        rs = r.get('pp_revenue_share')
        ps = r.get('pp_profit_share')
        r['_pp_gap'] = (ps - rs) if (rs is not None and ps is not None) else None
        if '_pp_sector_rank' not in r:
            r['_pp_sector_rank'] = None

    # Columns — triage order: key signals first, supporting data after, context last
    # (label, key, number_format, width)
    pp_cols = [
        # --- Frozen identifiers (4) ---
        ('Ticker',          'ticker',                   '@',            8),
        ('Rating',          'rating',                   '@',            9),
        ('Score',           '_composite_score',         '0"%"',        10),
        ('Sector',          'sector',                   '@',           22),
        # --- Primary signals: pool dominance + margin moat ---
        ('PP Multiple',     'pp_multiple',              '0.00"×"',     13),
        ('Strength',        '_pp_strength',             '@',           12),
        ('Sector Rank',     '_pp_sector_rank',          '0',           11),
        ('Margin Adv.',     'pp_margin_advantage',      '+0.0%;-0.0%', 12),
        # --- Secondary signals: how the share gap breaks down ---
        ('Profit Gap',      '_pp_gap',                  '+0.00%;-0.00%', 12),
        ('Rev Share',       'pp_revenue_share',         '0.00%',       11),
        ('Profit Share',    'pp_profit_share',          '0.00%',       12),
        ('Peer Pctile',     '_peer_pctile_pp_multiple', '0%',          11),
        # --- Margin detail ---
        ('Op. Margin',      'operating_margin',         '0.0%',        11),
        ('Sector Med. OPM', '_sector_median_opm',       '0.0%',        15),
        # --- Scale ---
        ('Revenue ($B)',    '_pp_rev_b',                '#,##0.0',     13),
        ('Op. Income ($M)', '_pp_oi_m',                 '#,##0.0',     15),
        # --- Sector structure ---
        ('HHI',             'pp_sector_hhi',            '0.000',       10),
        ('Mkt Structure',   '_hhi_label',               '@',           14),
        ('CR4',             'pp_sector_cr4',            '0.0%',        10),
        ('# in Sector',     'pp_sector_count',          '0',           11),
    ]
    n_pp_frozen = 4

    # Header row
    for ci, (label, _, _, _) in enumerate(pp_cols, 1):
        cell = pp_ws.cell(row=1, column=ci, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
    pp_ws.row_dimensions[1].height = 13

    # Tiered fills for PP Multiple
    pp_dominant_fill = PatternFill(start_color='1a9850', end_color='1a9850', fill_type='solid')
    pp_strong_fill   = PatternFill(start_color='74c476', end_color='74c476', fill_type='solid')
    pp_above_fill    = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    pp_below_fill    = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    pp_weak_fill     = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    pp_neutral_fill  = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    pp_dominant_font = Font(color='FFFFFF', bold=True)
    pp_strong_font   = Font(color='FFFFFF')
    margin_pos_fill  = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    margin_neg_fill  = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')

    def _pp_mult_fill(val):
        if val is None:   return pp_neutral_fill, data_font
        if val >= 2.0:    return pp_dominant_fill, pp_dominant_font
        if val >= 1.5:    return pp_strong_fill,   pp_strong_font
        if val >= 1.0:    return pp_above_fill,    data_font
        if val >= 0.5:    return pp_below_fill,    data_font
        return pp_weak_fill, data_font

    # Flat list: sector A-Z, then PP Multiple descending within sector
    pp_rows_all = [r for r in rows if r.get('pp_multiple') is not None]
    pp_rows_all.sort(key=lambda r: (r.get('sector') or '', -(r.get('pp_multiple') or 0)))

    current_row = 2
    for row in pp_rows_all:
        for ci, (_, key, fmt, _) in enumerate(pp_cols, 1):
            val = row.get(key)
            cell = pp_ws.cell(row=current_row, column=ci, value=val)
            if fmt:
                cell.number_format = fmt

            if key == 'pp_multiple':
                fill, font = _pp_mult_fill(val)
                cell.fill = fill
                cell.font = font
                cell.alignment = Alignment(horizontal='center')
            elif key == '_pp_strength':
                fill, font = _pp_mult_fill(row.get('pp_multiple'))
                cell.fill = fill
                cell.font = font
                cell.alignment = Alignment(horizontal='center')
            elif key == '_pp_sector_rank':
                # Rank 1 = top earner in sector — highlight gold/green
                if val == 1:
                    cell.fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
                    cell.font = Font(bold=True, color='000000')
                elif val == 2:
                    cell.fill = gate_yellow_fill
                    cell.font = data_font
                else:
                    cell.fill = white_fill
                    cell.font = data_font
                cell.alignment = Alignment(horizontal='center')
            elif key == 'pp_margin_advantage':
                # Match HTML threshold: ±5pp = meaningful
                if val is not None and val > 0.05:
                    cell.fill = margin_pos_fill
                    cell.font = Font(color='27AE60', bold=True)
                elif val is not None and val < -0.05:
                    cell.fill = margin_neg_fill
                    cell.font = Font(color='C0392B', bold=True)
                else:
                    cell.fill = white_fill
                    cell.font = data_font
            elif key == '_pp_gap':
                # Profit share excess over revenue share
                if val is not None and val > 0.05:
                    cell.fill = margin_pos_fill
                    cell.font = Font(color='27AE60', bold=True)
                elif val is not None and val < -0.05:
                    cell.fill = margin_neg_fill
                    cell.font = Font(color='C0392B', bold=True)
                else:
                    cell.fill = white_fill
                    cell.font = data_font
            elif key == '_peer_pctile_pp_multiple' and isinstance(val, (int, float)):
                if val >= 0.75:
                    cell.fill = gate_pass_fill
                elif val >= 0.25:
                    cell.fill = gate_yellow_fill
                else:
                    cell.fill = gate_fail_fill
                cell.alignment = Alignment(horizontal='center')
                cell.font = data_font
            elif ci <= n_pp_frozen:
                cell.fill = row_header_fill
                cell.font = row_header_font
            else:
                cell.fill = white_fill
                cell.font = data_font
        pp_ws.row_dimensions[current_row].height = 13
        current_row += 1

    # Column widths
    for ci, (_, _, _, w) in enumerate(pp_cols, 1):
        pp_ws.column_dimensions[get_column_letter(ci)].width = w

    pp_ws.freeze_panes = 'E2'

    # --- Add hyperlinks from Gates Passed cells → Matrix tab ---
    link_font = Font(color='4472C4', underline='single')  # blue underline
    matrix_ws = wb['Matrix']
    # Build ticker → Matrix row mapping (Matrix data starts at row 4)
    ticker_to_matrix_row = {}
    for ri in range(4, matrix_ws.max_row + 1):
        t = matrix_ws.cell(row=ri, column=1).value
        if t:
            ticker_to_matrix_row[t] = ri

    for sname in ('Analysis', 'Ownership', 'Company'):
        ws = wb[sname]
        cols_list = sheets_config[sname]
        gp_col = None
        for ci, (label, key, _) in enumerate(cols_list, 1):
            if key == '_gates_passed':
                gp_col = ci
                break
        if gp_col is None:
            continue
        # Find ticker column
        tk_col = None
        for ci, (_, key, _) in enumerate(cols_list, 1):
            if key == 'ticker':
                tk_col = ci
                break
        if tk_col is None:
            continue
        for ri in range(2, ws.max_row + 1):
            ticker = ws.cell(row=ri, column=tk_col).value
            m_row = ticker_to_matrix_row.get(ticker)
            if m_row:
                cell = ws.cell(row=ri, column=gp_col)
                cell.hyperlink = f"#Matrix!A{m_row}"
                cell.font = link_font

    wb.save(filename)

