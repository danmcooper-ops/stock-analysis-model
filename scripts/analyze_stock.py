# scripts/analyze_stock.py
import gc
import sys
import os
import io
import json
import logging
import re
import warnings
from datetime import date
from statistics import median as _median
import numpy as np
import pandas as pd
from urllib.request import urlopen, Request
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# Load .env from project root (simple key=value parser, no dependency needed)
_env_path = os.path.join(os.path.dirname(__file__), '..', '.env')
if os.path.exists(_env_path):
    try:
        with open(_env_path, encoding='utf-8') as _f:
            for _line in _f:
                _line = _line.strip()
                if _line and not _line.startswith('#') and '=' in _line:
                    _k, _v = _line.split('=', 1)
                    os.environ.setdefault(_k.strip(), _v.strip())
    except (OSError, UnicodeDecodeError) as _e:
        print(f"[WARN] Could not read .env ({_e}); continuing with process env only.",
              file=sys.stderr)

from data.yfinance_client import (YFinanceClient, EmptyYahooResponseError,
                                  MCAP_MAX_PLAUSIBLE)
from data.treasury_rate import fetch_risk_free_rate
from models.capm import (calculate_beta, r2_diagnostic, ggm_implied_re, buildup_re,
                         weekly_returns, rolling_betas, ROLLING_BETA_WINDOWS)
from models.dcf import (two_stage_ev_valuation, fair_value_per_share, dcf_sensitivity,
                        two_stage_ev_exit_multiple_valuation, monte_carlo_dcf,
                        reverse_dcf)
from models.ratios import (compute_ratios, calculate_roic, calculate_wacc,
                           calculate_fundamental_growth, compute_dupont)
from models.quality import (calculate_earnings_quality, calculate_piotroski_f,
                            calculate_revenue_cagr, calculate_interest_coverage,
                            calculate_net_debt_ebitda, get_net_debt,
                            calculate_altman_z, calculate_beneish_m)
from models.market import compute_relative_multiples, compute_analyst_consensus, extract_next_earnings
from models.ddm import (ddm_eligibility, estimate_ddm_growth, two_stage_ddm_valuation,
                         ddm_h_model_valuation, monte_carlo_ddm)
from models.epv import earnings_power_value_valuation, epv_with_growth_premium
from models.rim import residual_income_model_valuation
from models.nav import tangible_book_value_per_share
from models.portfolio import position_sizes, concentration_analysis
from models.utils import rank
from models.field_keys import (OPERATING_CF_KEYS, CAPEX_KEYS, _get,
                               DEBT_KEYS, CASH_KEYS)
from scripts.report_excel import build_excel
from scripts.report_html import build_html
from data.macro_client import MacroClient
from models.macro import (assess_macro_regime, compute_macro_adjustments,
                          print_macro_summary, generate_sector_signals,
                          compute_sector_rs_from_local)
from models.narrative import generate_stock_narrative, generate_financial_summary
from data.news_client import NewsClient
from data.tiingo_client import TiingoClient
from data.sec_legal_client import SECLegalClient
from data.finnhub_supply_client import FinnhubSupplyClient
from data.sec_supply_client import SECSupplyClient
from data.sec_xbrl_client import SECXBRLClient
from data.fx_client import get_spot_fx_rate, apply_fx_to_statement_df
from data.sec_insider_client import SECInsiderClient
from data.provenance import ProvenanceRecorder
from data.culture_client import CultureClient

from scripts.config import (ERP, TERMINAL_GROWTH_RATE,
                            GROWTH_WEIGHT_FCF, GROWTH_WEIGHT_REV,
                            GROWTH_WEIGHT_ANALYST_ST, GROWTH_WEIGHT_ANALYST_LT,
                            GROWTH_WEIGHT_EARNINGS_G, GROWTH_WEIGHT_FUNDAMENTAL,
                            SURPRISE_THRESHOLD, SURPRISE_UPLIFT,
                            MARGIN_TREND_SENSITIVITY,
                            BETA_MIN, BETA_MAX, RE_MIN, RE_MAX,
                            BETA_PRIOR_MEAN, BETA_PRIOR_SD,
                            CAPEX_DA_THRESHOLD, EXCESS_CAPEX_ADDBACK,
                            YIELD_CEILING_MULT, HYPER_GROWTH_CAP, ANALYST_HAIRCUT, FALLBACK_GROWTH,
                            DCF_YEARS, DCF_STAGE1,
                            EXIT_MULT_DIVERGENCE_THRESHOLD,
                            EXIT_MULT_DEFAULT_EV_EBITDA,
                            EXIT_MULT_MIN, EXIT_MULT_MAX,
                            MC_ITERATIONS, MC_GROWTH_SIGMA_RATIO, MC_WACC_SIGMA,
                            MC_TERMINAL_GROWTH_SIGMA, MC_EXIT_MULT_SIGMA_RATIO,
                            MC_HIGH_DIVERGENCE_SIGMA_MULT,
                            DDM_HIGH_GROWTH_YEARS, DDM_BLEND_WEIGHT,
                            DCF_BLEND_WEIGHT_WITH_DDM, DDM_DIVERGENCE_THRESHOLD,
                            BLEND_TRIGGER, BLEND_DCF_WEIGHT, BLEND_MULT_WEIGHT,
                            EV_EBITDA_OUTLIER_MAX, MIN_SECTOR_STOCKS,
                            MIN_MORNINGSTAR_SAMPLE,
                            _get_sector_config)
from scripts.scoring import (_mc_confidence_label, score_and_rate,
                             _print_validation_stats)

logger = logging.getLogger('analyze_stock')


# ---------------------------------------------------------------------------
# Local price file helpers
# ---------------------------------------------------------------------------

_FOUNDER_OVERRIDES_CACHE = None
_WIKIDATA_FOUNDERS_CACHE = None
_HONORIFICS = ('mr.', 'mrs.', 'ms.', 'miss', 'dr.', 'prof.', 'sir')

# Founder-led detection requires the founder to hold a CURRENT executive role
# (see data/founder_overrides.json for the definition). These token sets gate
# both detection layers so retired / board-only founders don't trip the flag.
#   - _FOUNDER_NONEXEC_MARKERS: if present, the role is honorary/board-only and
#     never counts (Chairman Emeritus, Founder & Director, retired, advisor).
#   - _EXEC_TITLE_TOKENS: at least one must be present for the role to count.
# Order matters: a title is executive only if it has an exec token AND no
# non-exec marker (so "Non-Executive Chairman" and "Chairman Emeritus" fail).
_FOUNDER_NONEXEC_MARKERS = ('emeritus', 'retired', 'former', 'non-executive',
                            'non executive', 'advisor', 'advisory')
# Descriptive exec phrases matched as substrings; "chief " covers every C-suite
# (Chief Executive/Operating/Technology/... Officer).
_EXEC_TITLE_PHRASES = ('chief ', 'president', 'executive chair', 'managing partner')
# Acronyms matched on word boundaries — a naive substring would mis-fire (e.g.
# "cto" lives inside "dire-cto-r", which would wrongly flag "Founder & Director").
_EXEC_TITLE_ABBR_RE = re.compile(r'\b(?:ceo|coo|cto|cfo)\b')


def _is_executive_title(title):
    """True if an officer title denotes a CURRENT executive role.

    Counts CEO / President / C-suite / COO / CTO / Executive Chair. Excludes
    honorary or board-only roles (Chairman Emeritus, Founder & Director, retired,
    advisor, non-executive) and a bare "Chairman" with no executive qualifier.
    """
    t = (title or '').lower()
    if not t:
        return False
    if any(x in t for x in _FOUNDER_NONEXEC_MARKERS):
        return False
    if any(x in t for x in _EXEC_TITLE_PHRASES):
        return True
    return bool(_EXEC_TITLE_ABBR_RE.search(t))


def _flow_to_annual(history):
    # The current EDGAR client returns one value per fiscal year keyed by
    # integer year (or its string form after JSON round-trip). Legacy
    # snapshots used date-string keys ("YYYY-MM-DD") with mixed quarterly
    # + annual entries — for those, sum 4-quarter years and pass through
    # annual-only years; mixed-period legacy years are dropped (the old
    # behavior, retained for backwards compatibility with stored data).
    if not history:
        return {}
    by_year = {}
    for k, v in history.items():
        if v is None:
            continue
        ks = str(k)
        try:
            yr = int(ks[:4])
        except (TypeError, ValueError):
            continue
        is_annual_key = isinstance(k, int) or (len(ks) == 4 and ks.isdigit())
        by_year.setdefault(yr, []).append((is_annual_key, v))
    annual = {}
    for yr, entries in by_year.items():
        annual_keyed = [v for ak, v in entries if ak]
        if annual_keyed:
            annual[yr] = annual_keyed[-1]
            continue
        vals = [v for _, v in entries]
        if len(vals) == 1:
            annual[yr] = vals[0]
        elif len(vals) == 4:
            annual[yr] = sum(vals)
    return annual


def _stock_to_annual(history):
    # Point-in-time concepts (e.g. shares outstanding): keep the latest
    # observation in each calendar year.
    if not history:
        return {}
    latest = {}
    for period_end, val in history.items():
        if val is None:
            continue
        try:
            yr = int(str(period_end)[:4])
        except (TypeError, ValueError):
            continue
        if yr not in latest or str(period_end) > latest[yr][0]:
            latest[yr] = (str(period_end), val)
    return {yr: pv[1] for yr, pv in latest.items()}


def derive_edgar_metrics(edgar_history):
    """Derive multi-year CAGRs, margin-trend and decline signals from EDGAR history.

    Returns a fixed-key dict (every key present; None when insufficient
    data): the growth/margin fields (rev_cagr_5y/10y, fcf_cagr_5y/10y,
    gross_margin_avg_5y, gross_margin_trend, dividend_cagr_5y,
    shares_cagr_5y, rev_growth_vol, op_margin_avg_10y+years, fcf_edgar,
    int_cov_edgar) plus the value-trap decline detectors (rev_down_years,
    net_debt_slope_3y, div_fcf_ratio_3y, fcf_neg_years_5y).

    Shared between the live analyze_stock pipeline and rescore_and_render
    so a snapshot whose edgar_history was populated after the fact (e.g.
    via backfill_edgar_hist) can refresh its derived growth signals
    without re-running the full 3–6 hr analysis.
    """
    out = dict(rev_cagr_5y=None, rev_cagr_10y=None,
               fcf_cagr_5y=None, fcf_cagr_10y=None,
               gross_margin_avg_5y=None, gross_margin_trend=None,
               dividend_cagr_5y=None, shares_cagr_5y=None,
               fcf_edgar=None, rev_growth_vol=None,
               op_margin_avg_10y=None, op_margin_hist_years=0,
               int_cov_edgar=None,
               # Value-trap decline detectors (consumed by
               # scoring.compute_trap_signals; None = insufficient history,
               # which the trap overlay treats as skip, never as bad):
               rev_down_years=None, net_debt_slope_3y=None,
               div_fcf_ratio_3y=None, fcf_neg_years_5y=None)
    if not edgar_history:
        return out

    rev_hist = _flow_to_annual(edgar_history.get('revenue_history', {}))
    ocf_hist = _flow_to_annual(edgar_history.get('operating_cf_history', {}))
    cap_hist = _flow_to_annual(edgar_history.get('capex_history', {}))
    gp_hist  = _flow_to_annual(edgar_history.get('gross_profit_history', {}))
    div_hist = _flow_to_annual(edgar_history.get('dividends_paid_history', {}))
    sh_hist  = _stock_to_annual(edgar_history.get('shares_history', {}))
    op_hist  = _flow_to_annual(edgar_history.get('operating_income_history', {}))
    ie_hist  = _flow_to_annual(edgar_history.get('interest_expense_history', {}))

    if rev_hist:
        sy = sorted(rev_hist.keys())
        # Revenue-growth volatility: the population std of YoY revenue growth
        # rates. A "business predictability" signal (lower = steadier top line)
        # that is growth-neutral by construction — it measures the variability
        # of growth, not its level, so a fast, steady compounder scores well
        # while a lumpy/cyclical one does not. Used by the Quality: Rev
        # Volatility gate that replaced the leverage-distortable ROE gate.
        _rv = [rev_hist[y] for y in sy if rev_hist.get(y) and rev_hist[y] > 0]
        if len(_rv) >= 4:
            _gr = [(_rv[i] - _rv[i - 1]) / _rv[i - 1] for i in range(1, len(_rv))]
            if len(_gr) >= 3:
                _gm = sum(_gr) / len(_gr)
                out['rev_growth_vol'] = (sum((x - _gm) ** 2 for x in _gr) / len(_gr)) ** 0.5
        newest_rev = rev_hist[sy[-1]] if sy else None
        if newest_rev and newest_rev > 0:
            if len(sy) >= 6:
                yr5 = rev_hist.get(sy[-6])
                if yr5 and yr5 > 0:
                    out['rev_cagr_5y'] = (newest_rev / yr5) ** (1 / 5) - 1
            if len(sy) >= 11:
                yr10 = rev_hist.get(sy[-11])
                if yr10 and yr10 > 0:
                    out['rev_cagr_10y'] = (newest_rev / yr10) ** (1 / 10) - 1
        # Consecutive YoY revenue declines ending at the latest fiscal year —
        # the "still shrinking NOW" complement to the CAGRs (a −2% 5y CAGR
        # can hide a business that already re-based and is growing again;
        # three straight down-years cannot). Adjacent fiscal years only, so
        # a gap in the filing history breaks the streak rather than bridging
        # two non-consecutive declines.
        if len(sy) >= 3:
            streak = 0
            for i in range(len(sy) - 1, 0, -1):
                prev_rev = rev_hist.get(sy[i - 1])
                cur_rev = rev_hist.get(sy[i])
                if (sy[i] - sy[i - 1] == 1 and prev_rev and prev_rev > 0
                        and cur_rev is not None and cur_rev < prev_rev):
                    streak += 1
                else:
                    break
            out['rev_down_years'] = streak

    if ocf_hist:
        common_years = sorted(set(ocf_hist) & set(cap_hist)) if cap_hist else sorted(ocf_hist)
        fcf_hist = {yr: ocf_hist[yr] - abs(cap_hist.get(yr, 0)) for yr in common_years}
        fcf_vals = [fcf_hist[yr] for yr in sorted(fcf_hist) if fcf_hist[yr] is not None]
        # Point-in-time FCF (latest fiscal year with BOTH operating cash flow
        # and capex) — used as a fallback for the `fcf` field when yfinance
        # has no cash-flow statement, which is common across the expanded
        # EDGAR universe. edgar_history is already USD-normalized to the same
        # basis as the row's revenue/mcap, so the downstream fcf_margin and
        # pfcf ratios stay FX-consistent. Requiring capex (via common_years)
        # avoids overstating FCF as bare OCF.
        if cap_hist and common_years:
            out['fcf_edgar'] = fcf_hist[common_years[-1]]
        if len(fcf_vals) >= 6 and fcf_vals[-6] > 0 and fcf_vals[-1] > 0:
            out['fcf_cagr_5y'] = (fcf_vals[-1] / fcf_vals[-6]) ** (1 / 5) - 1
        if len(fcf_vals) >= 11 and fcf_vals[-11] > 0 and fcf_vals[-1] > 0:
            out['fcf_cagr_10y'] = (fcf_vals[-1] / fcf_vals[-11]) ** (1 / 10) - 1
        # Negative-FCF years among the last ≤5 — chronic cash burn reads
        # differently from one bad year, and the CAGR fields are None for any
        # series that starts or ends negative, which is exactly the cohort a
        # trap detector cares about.
        if len(fcf_vals) >= 3:
            out['fcf_neg_years_5y'] = sum(1 for v in fcf_vals[-5:] if v < 0)
        # Dividends paid vs FCF over the trailing 3 common years — the
        # "unsustainable yield" trap bait. dividends_paid is a cash OUTFLOW
        # (typically negative in filings), hence the abs(). A payer whose 3y
        # FCF sum is non-positive gets the 9.99 cap (JSON-safe stand-in for
        # infinity); a genuine non-payer is 0.0 — zero payout-bait risk — so
        # the axis stays PRESENT for non-payers rather than skipping.
        if div_hist:
            common_dy = sorted(set(div_hist) & set(fcf_hist))[-3:]
            if len(common_dy) >= 2:
                div_sum = sum(abs(div_hist[y] or 0) for y in common_dy)
                fcf_sum = sum(fcf_hist[y] for y in common_dy)
                if div_sum <= 0:
                    out['div_fcf_ratio_3y'] = 0.0
                elif fcf_sum <= 0:
                    out['div_fcf_ratio_3y'] = 9.99
                else:
                    out['div_fcf_ratio_3y'] = min(div_sum / fcf_sum, 9.99)

    if gp_hist and rev_hist:
        common_gy = sorted(set(gp_hist) & set(rev_hist))[-5:]
        margins = [gp_hist[yr] / rev_hist[yr] for yr in common_gy
                   if rev_hist.get(yr) and rev_hist[yr] > 0]
        if len(margins) >= 3:
            out['gross_margin_avg_5y'] = sum(margins) / len(margins)
            n = len(margins)
            xs = list(range(n))
            x_mean = sum(xs) / n
            y_mean = sum(margins) / n
            denom = sum((x - x_mean) ** 2 for x in xs)
            if denom:
                out['gross_margin_trend'] = sum(
                    (xs[i] - x_mean) * (margins[i] - y_mean) for i in range(n)
                ) / denom

    if op_hist and rev_hist:
        # Normalized (through-cycle) operating margin: mean of op-income /
        # revenue over the last ≤10 common years. Feeds the Quality: Margin
        # vs Hist over-earning guard and the normalized-EBIT EPV input. The
        # margin is unitless, so EDGAR's USD normalization is irrelevant here.
        common_oy = sorted(set(op_hist) & set(rev_hist))[-10:]
        op_margins = [op_hist[yr] / rev_hist[yr] for yr in common_oy
                      if rev_hist.get(yr) and rev_hist[yr] > 0
                      and op_hist.get(yr) is not None]
        if len(op_margins) >= 2:
            out['op_margin_avg_10y'] = sum(op_margins) / len(op_margins)
            out['op_margin_hist_years'] = len(op_margins)

    if op_hist and ie_hist:
        # EBIT / interest expense straight from the filing — the fallback for
        # the ~37% of non-financial rows where yfinance surfaces no income
        # statement (AAPL and MSFT among them), leaving Quality: Int Coverage
        # N/A. Missing-data N/A still counts against the applicable-gate
        # denominator, so those rows were being scored as a failed leverage
        # test on absent data rather than bad fundamentals.
        # Same latest-fiscal-year basis and the same truthiness guards as
        # calculate_interest_coverage, so the two sources agree where both
        # resolve (HON: 6.05 either way). Negative EBIT propagates a negative
        # ratio — a real "cannot cover interest" read the gate should fail,
        # not suppress. Consumed via the fallback in prepare_scoring_fields.
        common_iy = sorted(set(op_hist) & set(ie_hist))
        if common_iy:
            _oi, _ie = op_hist.get(common_iy[-1]), ie_hist.get(common_iy[-1])
            if _oi and _ie and _ie > 0:
                out['int_cov_edgar'] = _oi / _ie

    if div_hist:
        dy = sorted(div_hist.keys())
        if len(dy) >= 6:
            d0, d1 = abs(div_hist.get(dy[-6], 0) or 0), abs(div_hist.get(dy[-1], 0) or 0)
            if d0 > 0 and d1 > 0:
                out['dividend_cagr_5y'] = (d1 / d0) ** (1 / 5) - 1

    if sh_hist:
        shy = sorted(sh_hist.keys())
        if len(shy) >= 6:
            s0, s1 = sh_hist.get(shy[-6]), sh_hist.get(shy[-1])
            if s0 and s0 > 0 and s1 and s1 > 0:
                out['shares_cagr_5y'] = (s1 / s0) ** (1 / 5) - 1

    # Net-debt trajectory over the last 4 common years, expressed as the
    # per-year build as a fraction of latest revenue — a slope, not a CAGR,
    # because net debt legitimately crosses zero and a ratio of signed values
    # is meaningless. +0.05 means the company adds net debt worth 5% of
    # revenue every year. Point-in-time concepts, so _stock_to_annual.
    # Consumed by the trap overlay only when leverage is already elevated
    # (rising debt is trap fuel, not a startup drawing a revolver once).
    td_hist = _stock_to_annual(edgar_history.get('total_debt_history', {}))
    ch_hist = _stock_to_annual(edgar_history.get('cash_history', {}))
    if td_hist and ch_hist and rev_hist:
        common_ny = sorted(set(td_hist) & set(ch_hist))[-4:]
        rev_sy = sorted(rev_hist.keys())
        rev_last = rev_hist.get(rev_sy[-1]) if rev_sy else None
        if len(common_ny) >= 4 and rev_last and rev_last > 0:
            nd_first = td_hist[common_ny[0]] - ch_hist[common_ny[0]]
            nd_last = td_hist[common_ny[-1]] - ch_hist[common_ny[-1]]
            span = common_ny[-1] - common_ny[0]
            if span > 0:
                out['net_debt_slope_3y'] = (nd_last - nd_first) / (span * rev_last)

    return out


def _select_epv_ebit(point_ebit, yf_revenue, op_margin_avg_10y,
                     op_margin_hist_years):
    """Choose the EBIT that EPV capitalizes: normalized when history allows.

    At cyclical margin peaks every FV model eats the same inflated point-in-
    time earnings, so FV Dispersion stays tight exactly when it shouldn't.
    With >=5y of EDGAR margin history, capitalize the ~10y-average operating
    margin applied to CURRENT revenue instead. The margin is unitless; it must
    multiply the yfinance income-statement revenue (quote currency), never
    EDGAR's USD-normalized revenue, so the result stays unit-consistent with
    the point-EBIT path.

    Returns (ebit_used, source) with source 'normalized' | 'point' | None.
    """
    can_normalize = (op_margin_avg_10y is not None
                     and (op_margin_hist_years or 0) >= 5
                     and yf_revenue is not None and yf_revenue > 0)
    if can_normalize:
        # Prefer the through-cycle figure even when point EBIT is missing —
        # dropping EPV entirely would starve the consensus fallback of exactly
        # the sparse-data names it exists to value.
        return op_margin_avg_10y * yf_revenue, 'normalized'
    if point_ebit is None:
        return None, None
    return point_ebit, 'point'


def _rescale_fv_band(row, ratio):
    """Scale every FV-denominated uncertainty field by `ratio` after a blend.

    When a blend rewrites dcf_fv, the sensitivity range AND the Monte Carlo
    percentiles must move with it — otherwise the reported bear/bull band and
    P10/P90 bracket a fair value that no longer exists (a blended FV could sit
    outside its own band). mc_cv is a ratio and is left untouched.
    """
    if ratio is None or ratio == 1.0:
        return
    sens = row.get('dcf_sens_range')
    if sens and len(sens) == 2 and all(isinstance(x, (int, float)) for x in sens):
        row['dcf_sens_range'] = (sens[0] * ratio, sens[1] * ratio)
    for key in ('mc_p10_fv', 'mc_p90_fv'):
        v = row.get(key)
        if isinstance(v, (int, float)):
            row[key] = v * ratio


def _load_founder_overrides():
    """Load and cache the curated founder-led overrides from
    data/founder_overrides.json. Returns a dict {ticker: bool}; underscored
    keys (e.g. _doc) are filtered out. Missing or malformed file → empty dict.
    """
    global _FOUNDER_OVERRIDES_CACHE
    if _FOUNDER_OVERRIDES_CACHE is not None:
        return _FOUNDER_OVERRIDES_CACHE
    path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..',
                        'data', 'founder_overrides.json')
    try:
        with open(path, encoding='utf-8') as _f:
            raw = json.load(_f)
        out = {k: bool(v) for k, v in raw.items()
               if not k.startswith('_') and isinstance(v, bool)}
    except Exception as e:
        logger.warning(f"founder_overrides.json unreadable ({e}) — founder overrides disabled this run")
        out = {}
    _FOUNDER_OVERRIDES_CACHE = out
    return out


def _load_wikidata_founders():
    """Load and cache Wikidata-derived founder names from
    data/wikidata_founders.json. Returns {ticker: [founder_name, ...]}.
    Missing file → empty dict. Build/refresh via build_wikidata_founders.py.
    """
    global _WIKIDATA_FOUNDERS_CACHE
    if _WIKIDATA_FOUNDERS_CACHE is not None:
        return _WIKIDATA_FOUNDERS_CACHE
    path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..',
                        'data', 'wikidata_founders.json')
    try:
        with open(path, encoding='utf-8') as _f:
            raw = json.load(_f)
        out = {k: list(v) for k, v in raw.items() if not k.startswith('_')}
    except Exception as e:
        logger.warning(f"wikidata_founders.json unreadable ({e}) — founder detection degraded this run")
        out = {}
    _WIKIDATA_FOUNDERS_CACHE = out
    return out


def _normalize_name(name):
    """Lowercase, strip honorifics, split into a set of word tokens."""
    if not name:
        return set()
    parts = name.lower().split()
    parts = [p.strip('.,') for p in parts if p.lower() not in _HONORIFICS]
    return set(p for p in parts if p)


def _wikidata_founder_match(ticker, officers):
    """Return the officer dict for a Wikidata-listed founder of *ticker* who
    currently holds an EXECUTIVE role, else None.

    Match logic: all words in the founder's normalized name must appear in some
    officer's normalized name, AND that officer's title must be executive (see
    _is_executive_title). Catches "Mark Zuckerberg" vs "Mr. Mark Elliot
    Zuckerberg"; falls through on nickname divergence like "Larry Ellison" vs
    "Lawrence J. Ellison" (manual override handles those). The executive gate
    stops board-only founders (e.g. a Chairman Emeritus listed as an officer)
    from tripping the flag.
    """
    founders = _load_wikidata_founders().get(ticker, [])
    if not founders or not officers:
        return None
    for fn in founders:
        fwords = _normalize_name(fn)
        if not fwords:
            continue
        for o in officers:
            ow = _normalize_name(o.get('name', ''))
            if ow and fwords.issubset(ow) and _is_executive_title(o.get('title')):
                return o
    return None


def _sec_reporting_officers(insider_data):
    """Officer roster derived from SEC Form 4 filings (see SECInsiderClient).

    Returns a list of {name, title} for reporting persons flagged isOfficer.
    Supplements yfinance's comp-table companyOfficers list, which omits roles
    like Executive Chairman — the source gap behind founder-led false negatives.
    """
    if not insider_data:
        return []
    return insider_data.get('reporting_officers') or []


def _load_local_prices(ticker, prices_dir):
    """Load Close price series from local Parquet file. Returns pd.Series or None."""
    if not prices_dir:
        return None
    path = os.path.join(prices_dir, f"{ticker}.parquet")
    if not os.path.exists(path):
        return None
    try:
        df = pd.read_parquet(path)[['Close']].sort_index()
        df.index = pd.to_datetime(df.index).tz_localize(None)
        return df['Close']
    except Exception as e:
        logger.warning(f"local prices: unreadable Parquet for {ticker} ({e})")
        return None


def _load_local_ohlcv(ticker, prices_dir):
    """Load Close + Volume from local Parquet. Returns DataFrame or None.

    Most files carry full OHLCV (written by download_prices.py), but ~2% hold
    only a Close column — YFinanceClient._maybe_persist_prices write-throughs
    persist just the close. A columns=[...] projection RAISES on those, so read
    the file and then check what arrived.
    """
    if not prices_dir:
        return None
    path = os.path.join(prices_dir, f"{ticker}.parquet")
    if not os.path.exists(path):
        return None
    try:
        df = pd.read_parquet(path).sort_index()
        if 'Close' not in df.columns or 'Volume' not in df.columns:
            return None
        df.index = pd.to_datetime(df.index).tz_localize(None)
        return df[['Close', 'Volume']]
    except Exception as e:
        logger.warning(f"local OHLCV: unreadable Parquet for {ticker} ({e})")
        return None


def compute_multiple_vs_history(close, edgar_history, operating_income,
                                max_years=10, min_years=5,
                                max_price_staleness_days=30):
    """Current price/EBIT multiple vs the firm's OWN ~10y median.

    The time-series cheapness axis the cross-sectional and model-based
    valuation gates don't cover: is this cheap relative to ITSELF? Replaced
    the EPV Floor gate, which correlated 0.68 with MoS (a third price-vs-
    intrinsic ratio on the same price denominator).

    Ratio-of-ratios construction: multiple(y) = adjusted close at year-end y
    ÷ operating income y (EDGAR, USD), scored as current multiple ÷ own
    median − 1. Because the SAME adjusted price series supplies every
    numerator, its constant per-share basis (split factor, ADR ratio)
    cancels between current and median — no share-count history needed.
    (The earlier price×shares construction mixed split-adjusted prices with
    as-reported share counts, inflating historical multiples of any splitter
    — GOOGL read as 3.1× its own median.) Only positive-EBIT years count.

    Residual biases, both modest: the auto-adjusted series folds dividends
    into old prices (deflating them), so dividend payers read slightly
    expensive vs history — conservative; and net issuance/buybacks shift
    the per-share claim on EBIT over time. Current EBIT comes from the
    run's USD-normalized yfinance statement, history from EDGAR USD.

    Returns (mult_vs_hist, years_used):
        mult_vs_hist = current_multiple / median(historical multiples) − 1
        (negative = trading below own history = cheap), or (None, n).
    """
    if (close is None or len(close) == 0 or not edgar_history
            or operating_income is None or operating_income <= 0):
        return None, 0
    oi_hist = edgar_history.get('operating_income_history') or {}
    if not oi_hist:
        return None, 0

    current_year = date.today().year
    years = sorted(int(y) for y in oi_hist.keys()
                   if int(str(y)) < current_year)[-max_years:]
    idx = close.index
    mults = []
    for y in years:
        oiy = oi_hist.get(y, oi_hist.get(str(y)))
        if not oiy or oiy <= 0:
            continue
        target = pd.Timestamp(f'{y}-12-31')
        # Year-end close within 30 days
        pos = idx.get_indexer([target], method='nearest')[0]
        if pos < 0 or abs((idx[pos] - target).days) > 30:
            continue
        py = float(close.iloc[pos])
        if py > 0:
            mults.append(py / oiy)

    if len(mults) < min_years:
        return None, len(mults)
    med = _median(mults)
    if not med or med <= 0:
        return None, len(mults)
    # "Current" price from the same adjusted series (basis consistency);
    # a stale tail (delisted / failed refresh) must not masquerade as today.
    last_bar = idx[-1]
    if (pd.Timestamp(date.today()) - last_bar).days > max_price_staleness_days:
        return None, len(mults)
    p_now = float(close.iloc[-1])
    if p_now <= 0:
        return None, len(mults)
    return (p_now / operating_income) / med - 1.0, len(mults)


def _rolling_betas_from_prices(stock_close, market_close):
    """Weekly rolling 1y/3y/5y betas from two close series (see
    models.capm.rolling_betas). Same method as the headline beta in
    select_cost_of_equity, so a parquet-sourced fallback is comparable.
    Returns {} on degenerate inputs."""
    try:
        s_ret, m_ret, _ = weekly_returns(_to_tznaive(stock_close),
                                         _to_tznaive(market_close))
        return rolling_betas(s_ret, m_ret,
                             prior_mean=BETA_PRIOR_MEAN, prior_sd=BETA_PRIOR_SD)
    except ValueError:
        return {}


def _realized_vol(close_series, window_days=252):
    """Annualized realized volatility over the trailing *window_days* trading days."""
    ret = close_series.pct_change().dropna()
    tail = ret.tail(window_days)
    if len(tail) < 60:
        return None
    return float(tail.std() * np.sqrt(252))


def _momentum_window(close_series, months_back, months_skip=0, as_of=None):
    """Price return from *months_back* ago up to *months_skip* months ago.

    months_skip=1 drops the most recent month — the standard guard against
    short-horizon mean reversion contaminating a momentum signal.
    """
    as_of = as_of or pd.Timestamp.today().normalize()
    start_cut = as_of - pd.DateOffset(months=months_back)
    end_cut   = as_of - pd.DateOffset(months=months_skip) if months_skip else as_of
    s = close_series.loc[close_series.index <= as_of]
    if s.empty:
        return None
    after_start = s.loc[s.index >= start_cut]
    before_end  = s.loc[s.index <= end_cut]
    if after_start.empty or before_end.empty:
        return None
    p_start = float(after_start.iloc[0])
    p_end   = float(before_end.iloc[-1])
    if p_start <= 0:
        return None
    return (p_end - p_start) / p_start


def _momentum_12_1(close_series, as_of=None):
    """12-minus-1 month price momentum (skips most recent month to avoid reversal)."""
    return _momentum_window(close_series, months_back=12, months_skip=1, as_of=as_of)


def _avg_dollar_volume(ohlcv, days=63):
    """Median daily dollar volume (Close x Volume) over the trailing *days* bars.

    Median rather than mean: one earnings-day or index-rebalance spike otherwise
    inflates apparent tradeability by an order of magnitude, which is exactly
    the wrong error to make for a liquidity floor.
    """
    tail = ohlcv.tail(days).dropna(subset=['Close', 'Volume'])
    if len(tail) < 20:
        return None
    dv = (tail['Close'] * tail['Volume']).replace(0, np.nan).dropna()
    if dv.empty:
        return None
    return float(dv.median())


# Amihud is mean(|return| / dollar volume) — a per-dollar figure around 1e-13
# for a mega-cap. Expressed in BASIS POINTS OF PRICE MOVE PER $1M TRADED it
# lands in a readable 0.01-1000 range: ~0.01 for AAPL, ~800 for a microcap.
_AMIHUD_BPS_PER_1M = 1e10


def _amihud_illiquidity(ohlcv, days=252):
    """Amihud price impact, in basis points of price move per $1M traded.

    Higher = each dollar traded moves the price more. Near zero for mega-caps
    by construction; the metric earns its keep at the illiquid end.
    """
    tail = ohlcv.tail(days + 1).dropna(subset=['Close', 'Volume'])
    if len(tail) < 60:
        return None
    ret = tail['Close'].pct_change().abs()
    dv  = (tail['Close'] * tail['Volume']).replace(0, np.nan)
    impact = (ret / dv).replace([np.inf, -np.inf], np.nan).dropna()
    if impact.empty:
        return None
    return float(impact.mean() * _AMIHUD_BPS_PER_1M)


def _volume_trend(ohlcv, short_days=21, long_days=252):
    """Recent volume against the stock's own baseline: mean(21d) / mean(252d).

    Above 1.0 means trading interest is running hotter than normal for this
    name — read alongside momentum, not on its own.
    """
    tail = ohlcv.tail(long_days).dropna(subset=['Volume'])
    if len(tail) < 60:
        return None
    long_avg = float(tail['Volume'].mean())
    if long_avg <= 0:
        return None
    short_avg = float(tail['Volume'].tail(short_days).mean())
    return short_avg / long_avg


# Trailing-window price metrics measured "to today" are only honest if the
# series actually reaches today. ~14 calendar days ~ 10 trading days.
_PRICE_STALENESS_LIMIT_DAYS = 14


def _prices_are_stale(index, as_of=None):
    """True when the last bar is too old for trailing price metrics to be honest."""
    if index is None or len(index) == 0:
        return True
    as_of = as_of or pd.Timestamp.today().normalize()
    return (as_of - index.max()).days > _PRICE_STALENESS_LIMIT_DAYS


PRICE_METRIC_FIELDS = (
    'momentum_12_1', 'momentum_6_1', 'momentum_3m', 'vol_adj_momentum',
    'avg_dollar_volume_3m', 'amihud_illiquidity', 'volume_trend',
    'price_data_stale',
)


# Per-field rounding for the JSON snapshot. Dollar volume is a large absolute
# number (whole dollars is plenty); Amihud impact is tiny and needs the digits.
_PRICE_METRIC_ROUNDING = {
    'momentum_12_1': 4, 'momentum_6_1': 4, 'momentum_3m': 4,
    'vol_adj_momentum': 3, 'avg_dollar_volume_3m': 0,
    'amihud_illiquidity': 3, 'volume_trend': 3,
}


def _round_price_metrics(metrics):
    """Round price metrics for JSON output, leaving None and bools untouched."""
    out = {}
    for k, v in metrics.items():
        digits = _PRICE_METRIC_ROUNDING.get(k)
        if v is None or digits is None:
            out[k] = v
        else:
            out[k] = round(v, digits) if digits else round(v)
    return out


def price_metrics_from_series(ohlcv, close, as_of=None):
    """Momentum + volume/liquidity metrics from already-loaded price data.

    Returns a dict of field -> value (None where unavailable). Pure — no I/O —
    so the live run can reuse the Parquet read it already does.

    Every trailing-to-today metric is suppressed when the series is stale: a
    3-month-old file yields a "12-month momentum" that is really 9 months of
    momentum and 3 months of nothing, which is worse than no number at all.
    Drawdowns, realized vol and beta are deliberately NOT computed here — they
    are fixed historical windows or feed the DCF, and stay on their own path.
    """
    out = {k: None for k in PRICE_METRIC_FIELDS}

    if close is None or len(close) < 60:
        return out

    stale = _prices_are_stale(close.index, as_of=as_of)
    out['price_data_stale'] = bool(stale)
    if stale:
        return out

    out['momentum_12_1'] = _momentum_12_1(close, as_of=as_of)
    out['momentum_6_1']  = _momentum_window(close, months_back=6, months_skip=1, as_of=as_of)
    out['momentum_3m']   = _momentum_window(close, months_back=3, as_of=as_of)

    rv = _realized_vol(close)
    if rv is not None and rv > 1e-6 and out['momentum_12_1'] is not None:
        out['vol_adj_momentum'] = out['momentum_12_1'] / rv

    # Volume metrics need the Volume column, absent from the ~2% Close-only files.
    if ohlcv is not None:
        out['avg_dollar_volume_3m'] = _avg_dollar_volume(ohlcv)
        out['amihud_illiquidity']   = _amihud_illiquidity(ohlcv)
        out['volume_trend']         = _volume_trend(ohlcv)

    return out


def compute_price_metrics(ticker, prices_dir, as_of=None):
    """Load one ticker's local Parquet and derive its price metrics.

    Thin loader around price_metrics_from_series, used by the rescore path so
    it and the live run can never drift apart.
    """
    ohlcv = _load_local_ohlcv(ticker, prices_dir)
    close = ohlcv['Close'] if ohlcv is not None else _load_local_prices(ticker, prices_dir)
    return price_metrics_from_series(ohlcv, close, as_of=as_of)


def _max_drawdown_period(close_series, start, end):
    """Max drawdown (as negative fraction) within [start, end]. Returns None if no data."""
    s = close_series.loc[(close_series.index >= pd.Timestamp(start)) &
                         (close_series.index <= pd.Timestamp(end))]
    if len(s) < 5:
        return None
    roll_max = s.cummax()
    dd = (s - roll_max) / roll_max
    return float(dd.min())


# ---------------------------------------------------------------------------
# Ticker universe helpers
# ---------------------------------------------------------------------------

def _read_wiki_tables(url):
    req = Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urlopen(req) as resp:
        html = resp.read().decode("utf-8")
    return pd.read_html(io.StringIO(html))

def get_sp500_tickers():
    tables = _read_wiki_tables("https://en.wikipedia.org/wiki/List_of_S%26P_500_companies")
    return tables[0]['Symbol'].tolist()

def get_nyse_tickers():
    try:
        return pd.read_csv("nyse_tickers.csv")['Symbol'].tolist()
    except FileNotFoundError:
        print("nyse_tickers.csv not found, skipping NYSE tickers.")
        return []

def get_dow_tickers():
    tables = _read_wiki_tables("https://en.wikipedia.org/wiki/Dow_Jones_Industrial_Average")
    for table in tables:
        if 'Symbol' in table.columns:
            return table['Symbol'].tolist()
    return []


# ---------------------------------------------------------------------------
# CAPM / cost-of-equity helpers (Worksheet Steps 4A-4B)
# ---------------------------------------------------------------------------


def _to_tznaive(series):
    """Strip tz so a tz-aware (yfinance) and tz-naive (Tiingo) price series can
    be joined — a mixed pair otherwise raises or yields zero index overlap."""
    idx = getattr(series, 'index', None)
    if idx is not None and getattr(idx, 'tz', None) is not None:
        series = series.copy()
        series.index = idx.tz_localize(None)
    return series


def select_cost_of_equity(financials, risk_free_rate, yf_client=None, ticker=None,
                          erp=None, tiingo_client=None):
    """Select cost of equity using a four-level hierarchy.

    Tries each method in order, returning the first that passes validation:
      1. CAPM with a locally computed 5-year weekly beta, shrunk toward
         BETA_PRIOR_MEAN by its estimation precision (preferred; label 'capm')
      2. CAPM with yfinance-reported beta, Blume-adjusted since it carries
         no standard error (label 'capm (yahoo beta)')
      3. GGM-implied Re (for dividend payers; label 'ggm')
      4. Build-Up (last resort; label 'buildup')

    Args:
        financials: Dict of financial data from YFinanceClient.
        risk_free_rate: Current risk-free rate (10-yr Treasury yield).
        yf_client: Optional YFinanceClient for computing beta from prices.
        ticker: Optional ticker symbol for price history lookup.
        erp: Equity risk premium override (defaults to module-level ERP).

    Returns:
        Tuple of (cost_of_equity, method_label, beta_diagnostics_or_None).
    """
    if erp is None:
        erp = ERP
    info = (financials.get('info') or {}) if financials else {}
    beta_diag = None

    # 1. CAPM with computed beta and R² quality gate
    if yf_client and ticker:
        try:
            stock_prices = yf_client.fetch_history(ticker, period="5y")
            market_prices = yf_client.fetch_history('SPY', period="5y")
            # Fall back to Tiingo if yfinance returns insufficient price data
            if tiingo_client and tiingo_client.available:
                if stock_prices is None or len(stock_prices) <= 60:
                    stock_prices = tiingo_client.fetch_history(ticker, period="5y")
                if market_prices is None or len(market_prices) <= 60:
                    market_prices = tiingo_client.fetch_history('SPY', period="5y")
            if (stock_prices is not None and market_prices is not None
                    and len(stock_prices) > 60):
                # WEEKLY returns (5y ≈ 260 obs). Daily single-stock betas are
                # attenuated by non-synchronous trading; weekly is the
                # frequency the rolling-beta diagnostic shares, so the 5y
                # window there equals this headline beta by construction.
                stock_ret, market_ret, _ = weekly_returns(
                    _to_tznaive(stock_prices), _to_tznaive(market_prices))
                if len(stock_ret) > 60:
                    n5 = ROLLING_BETA_WINDOWS['5y']
                    beta_result = calculate_beta(
                        stock_ret[-n5:], market_ret[-n5:],
                        prior_mean=BETA_PRIOR_MEAN, prior_sd=BETA_PRIOR_SD)
                    r2_class, _r2_method = r2_diagnostic(beta_result['r_squared'])
                    beta_diag = {
                        **beta_result,
                        'r2_classification': r2_class,
                        'beta_source': 'weekly_5y',
                        'rolling_betas': rolling_betas(
                            stock_ret, market_ret,
                            prior_mean=BETA_PRIOR_MEAN, prior_sd=BETA_PRIOR_SD) or None,
                    }
                    for _msg in beta_result['warnings']:
                        warnings.warn(f"{ticker}: beta: {_msg}", RuntimeWarning,
                                      stacklevel=2)

                    # The precision-weighted beta already carries the
                    # "how much to trust this regression" decision (a noisy
                    # beta is pulled toward the prior), so no R² tiering.
                    beta_used = beta_result['shrunk_beta']
                    if BETA_MIN < beta_used < BETA_MAX:
                        re = risk_free_rate + beta_used * erp
                        if RE_MIN < re < RE_MAX:
                            return re, 'capm', beta_diag
        except Exception as e:
            logger.warning(f"cost of equity: local-beta path failed for {ticker} "
                           f"(yfinance/tiingo prices: {e}); falling back to yfinance beta")

    # 2. CAPM with yfinance-reported beta. Yahoo's is a raw 5-year monthly
    #    beta with no standard error, so apply the Blume adjustment rather
    #    than leave it unshrunk next to the precision-weighted local betas.
    beta = info.get('beta')
    if beta is not None and BETA_MIN < beta < BETA_MAX:
        beta_used = (2 / 3) * beta + (1 / 3) * BETA_PRIOR_MEAN
        re = risk_free_rate + beta_used * erp
        if RE_MIN < re < RE_MAX:
            yahoo_diag = {
                'raw_beta': float(beta),
                'adjusted_beta': beta_used,
                'shrunk_beta': beta_used,
                'shrink_weight': 2 / 3,
                'r_squared': None, 'se_beta': None, 'n_observations': None,
                'r2_classification': None,
                'beta_source': 'yahoo',
                'warnings': ['Yahoo-reported beta (no local price history); '
                             'Blume-adjusted, precision unknown'],
            }
            if beta_diag:  # keep any rolling diagnostics the local path built
                yahoo_diag['rolling_betas'] = beta_diag.get('rolling_betas')
            return re, 'capm (yahoo beta)', yahoo_diag

    # 3. GGM-implied: Re = D1/P + g. yfinance's dividendRate is the forward
    #    (indicated) annual rate, so the yield is already D1/P.
    div_rate = info.get('dividendRate')
    price = info.get('currentPrice') or info.get('regularMarketPrice')
    div_yield = (div_rate / price) if (div_rate and price and price > 0) else None
    if div_yield and div_yield > 0:
        re = ggm_implied_re(div_yield, TERMINAL_GROWTH_RATE, forward=True)
        if re is not None and RE_MIN < re < RE_MAX:
            return re, 'ggm', beta_diag

    # 4. Build-Up fallback (no size/industry premiums — too imprecise)
    re = buildup_re(risk_free_rate, erp, size_premium=0, industry_premium=0)
    return re, 'buildup', beta_diag


# ---------------------------------------------------------------------------
# Growth estimation helpers
# ---------------------------------------------------------------------------

def _get_analyst_lt_growth(yf_data):
    """Extract analyst forward growth estimate from yfinance growth_estimates.

    Priority: LTG (long-term ~5yr) → +1y (next year) → 0y (current year).
    LTG is often NaN in yfinance, so +1y is the practical primary signal.
    Returns decimal (e.g. 0.12 for 12%) or None.
    """
    ge = yf_data.get('growth_estimates')
    if ge is not None and hasattr(ge, 'empty') and not ge.empty:
        try:
            # yfinance column name varies across versions: 'stockTrend', 'Stock', etc.
            col = next((c for c in ('stockTrend', 'Stock') if c in ge.columns), None)
            if col is None:
                return None
            # Try LTG first, then +1y, then 0y
            for period in ['LTG', '+1y', '0y']:
                if period in ge.index:
                    val = ge.loc[period, col]
                    if pd.notna(val) and isinstance(val, (int, float)):
                        return float(val)
        except Exception as e:
            logger.debug(f"growth estimates table unusable ({e}); trying next source")
    return None


def _get_earnings_growth(yf_data):
    """Extract analyst 1-year forward earnings growth from info dict.

    Returns decimal (e.g. 0.18 for 18%) or None.
    """
    info = yf_data.get('info') or {}
    eg = info.get('earningsGrowth')
    if eg is not None and isinstance(eg, (int, float)):
        return float(eg)
    return None


def _compute_surprise_adjustment(yf_data):
    """Compute growth adjustment from earnings surprise history.

    If the company consistently beats estimates by > SURPRISE_THRESHOLD,
    analyst estimates are systematically low → uplift growth.
    If consistently misses by > threshold → penalise growth.

    Returns adjustment in decimal (e.g. +0.015 or -0.015 or 0.0).
    Also returns the average surprise % for display.
    """
    eh = yf_data.get('earnings_history')
    if eh is None or not hasattr(eh, 'empty') or eh.empty:
        return 0.0, None

    try:
        if 'surprisePercent' not in eh.columns:
            return 0.0, None
        surprises = eh['surprisePercent'].dropna()
        if len(surprises) < 2:
            return 0.0, None
        avg_surprise = float(surprises.mean())
        if avg_surprise > SURPRISE_THRESHOLD:
            return SURPRISE_UPLIFT, avg_surprise
        elif avg_surprise < -SURPRISE_THRESHOLD:
            return -SURPRISE_UPLIFT, avg_surprise
        else:
            return 0.0, avg_surprise
    except Exception as e:
        logger.debug(f"earnings surprise history unusable ({e}); no surprise uplift applied")
        return 0.0, None


def _compute_margin_trend(yf_data):
    """Compute operating margin trend from income statement.

    Calculates operating margin (Operating Income / Total Revenue) for each
    year in the income statement and returns the average annual change.

    If margins are expanding, FCF should grow faster than revenue.
    Returns annual margin change in decimal (e.g. +0.02 = expanding 2pp/yr)
    or None if insufficient data.
    """
    inc = yf_data.get('income_statement')
    if inc is None or (hasattr(inc, 'empty') and inc.empty):
        inc = yf_data.get('income_stmt')
    if inc is None or (hasattr(inc, 'empty') and inc.empty):
        return None

    op_inc_keys = ['Operating Income', 'Total Operating Income As Reported']
    rev_keys = ['Total Revenue']

    margins = []
    # yfinance columns are fiscal years (most recent first)
    for col in inc.columns:
        year_data = inc[col]
        op_inc = None
        for k in op_inc_keys:
            if k in year_data.index and pd.notna(year_data[k]):
                op_inc = float(year_data[k])
                break
        rev = None
        for k in rev_keys:
            if k in year_data.index and pd.notna(year_data[k]):
                rev = float(year_data[k])
                break
        if op_inc is not None and rev and rev > 0:
            margins.append(op_inc / rev)

    if len(margins) < 2:
        return None

    # margins[0] is most recent, margins[-1] is oldest
    # Annual change = (newest - oldest) / number_of_gaps
    n_gaps = len(margins) - 1
    annual_change = (margins[0] - margins[-1]) / n_gaps
    return annual_change


def _compute_gross_margin(yf_data):
    """Compute median gross margin from income statement.

    Gross Margin = Gross Profit / Total Revenue for each fiscal year.
    Returns the median across available years (more robust to one-off swings
    than the mean). A high, stable gross margin (>40%) signals pricing power
    and is a classic moat indicator.

    Returns median gross margin in decimal (e.g. 0.45 = 45%) or None.
    """
    inc = yf_data.get('income_statement')
    if inc is None or (hasattr(inc, 'empty') and inc.empty):
        inc = yf_data.get('income_stmt')
    if inc is None or (hasattr(inc, 'empty') and inc.empty):
        return None

    gp_keys = ['Gross Profit']
    rev_keys = ['Total Revenue']

    margins = []
    for col in inc.columns:
        year_data = inc[col]
        gp = None
        for k in gp_keys:
            if k in year_data.index and pd.notna(year_data[k]):
                gp = float(year_data[k])
                break
        rev = None
        for k in rev_keys:
            if k in year_data.index and pd.notna(year_data[k]):
                rev = float(year_data[k])
                break
        if gp is not None and rev and rev > 0:
            margins.append(gp / rev)

    if not margins:
        return None

    margins_sorted = sorted(margins)
    n = len(margins_sorted)
    if n % 2 == 1:
        return margins_sorted[n // 2]
    return (margins_sorted[n // 2 - 1] + margins_sorted[n // 2]) / 2.0


def _extract_latest_financials(yf_data):
    """Extract most recent year revenue, operating income, net income.

    Returns dict with absolute dollar values (or None for missing fields).
    Used for profit pool analysis: sector-level revenue/profit concentration.
    """
    inc = yf_data.get('income_statement')
    if inc is None or (hasattr(inc, 'empty') and inc.empty):
        inc = yf_data.get('income_stmt')
    if inc is None or (hasattr(inc, 'empty') and inc.empty):
        return {}
    latest = inc.iloc[:, 0]  # most recent fiscal year

    rev = None
    for k in ['Total Revenue']:
        if k in latest.index and pd.notna(latest[k]):
            rev = float(latest[k])
            break

    op_inc = None
    for k in ['Operating Income', 'Total Operating Income As Reported']:
        if k in latest.index and pd.notna(latest[k]):
            op_inc = float(latest[k])
            break

    net_inc = None
    for k in ['Net Income', 'Net Income Common Stockholders']:
        if k in latest.index and pd.notna(latest[k]):
            net_inc = float(latest[k])
            break

    return {
        'revenue': rev,
        'operating_income': op_inc,
        'net_income': net_inc,
    }


# Per-share / total-dollar fields in yfinance ``info`` that need to be
# multiplied by the FX spot rate when the quote currency isn't USD. Ratios
# (trailingPE, priceToBook, enterpriseToEbitda, dividendYield, payoutRatio)
# are dimensionless and intentionally excluded.
_FX_INFO_DOLLAR_FIELDS = (
    'marketCap', 'enterpriseValue',
    'currentPrice', 'regularMarketPrice', 'previousClose',
    'fiftyTwoWeekHigh', 'fiftyTwoWeekLow',
    'dividendRate', 'bookValue',
    'targetMeanPrice', 'targetHighPrice', 'targetLowPrice',
    'lastDividendValue',
    # trailingEps IS a per-share currency amount (read by models/market.py);
    # lastDividendDate is a Unix epoch timestamp and must NOT be scaled —
    # the loop floats anything float()-able, so listing it here multiplied
    # a date by the FX rate.
    'trailingEps',
)


def _convert_financials_to_usd(yf_data, statements_are_usd=False):
    """Convert ``yf_data`` financials + dollar-denominated info fields to USD.

    ``statements_are_usd``: set when the statements came from SEC XBRL, which
    is extracted in USD units. In that case the balance-sheet / income /
    cash-flow frames are already USD even if yfinance's ``financialCurrency``
    reports a local currency (a US-listed foreign private issuer), so they must
    NOT be converted again — only the quote-currency info fields are.

    Foreign-domiciled tickers report financials in local currency but the
    valuation pipeline discounts at a USD-anchored WACC (US Treasury yield
    + global ERP). Without normalization, per-share DCF / EPV / RIM / NAV
    FVs are in the local currency while the displayed price may be in USD
    (ADRs) or the same local currency (.SW, .L exchanges) — producing the
    "wild divergence" the user observed.

    Returns ``(yf_data, fx_meta)``. ``fx_meta`` carries:
        currency_quote      — info['currency'] (price currency)
        currency_financial  — info['financialCurrency'] (statement currency)
        fx_rate_financial   — rate applied to balance_sheet / income_stmt / cash_flow (None when no conversion)
        fx_rate_quote       — rate applied to info dollar fields (None when no conversion)
        fx_converted        — True iff any conversion was applied
        fx_fetch_failed     — True iff at least one rate lookup returned None for a non-USD currency
    """
    info = yf_data.get('info') or {}
    ccy_fin = yf_data.get('currency_financial') or info.get('financialCurrency') or info.get('currency')
    ccy_quote = yf_data.get('currency_quote') or info.get('currency')
    fx_meta = {
        'currency_quote': ccy_quote,
        'currency_financial': ccy_fin,
        'fx_rate_financial': None,
        'fx_rate_quote': None,
        'fx_converted': False,
        'fx_fetch_failed': False,
    }
    needs_fin = ccy_fin and ccy_fin != 'USD' and not statements_are_usd
    needs_quote = ccy_quote and ccy_quote != 'USD'
    if not needs_fin and not needs_quote:
        return yf_data, fx_meta
    # Shallow-copy the outer dict so we don't mutate the cached payload.
    out = dict(yf_data)
    if needs_fin:
        rate_fin = get_spot_fx_rate(ccy_fin)
        if rate_fin is None:
            fx_meta['fx_fetch_failed'] = True
        else:
            fx_meta['fx_rate_financial'] = rate_fin
            fx_meta['fx_converted'] = True
            for key in ('balance_sheet', 'income_statement', 'income_stmt', 'cash_flow'):
                df = out.get(key)
                if df is not None:
                    out[key] = apply_fx_to_statement_df(df, rate_fin)
    if needs_quote:
        # Often the quote and financial currencies are the same (e.g.,
        # NESN.SW: CHF/CHF). get_spot_fx_rate hits the same cache entry.
        if needs_fin and ccy_quote == ccy_fin:
            rate_quote = fx_meta['fx_rate_financial']
        else:
            rate_quote = get_spot_fx_rate(ccy_quote)
        if rate_quote is None:
            fx_meta['fx_fetch_failed'] = True
        else:
            fx_meta['fx_rate_quote'] = rate_quote
            fx_meta['fx_converted'] = True
            new_info = dict(info)
            for f in _FX_INFO_DOLLAR_FIELDS:
                v = new_info.get(f)
                if v is None:
                    continue
                try:
                    fv = float(v)
                except (TypeError, ValueError):
                    continue
                new_info[f] = fv * rate_quote
            out['info'] = new_info
    return out, fx_meta


def _fcf_series_from_cashflow(cf):
    """Return a period-sorted pandas Series of Free Cash Flow values from a
    yfinance cash-flow DataFrame, deriving FCF = OCF + Capex when the
    pre-computed 'Free Cash Flow' row is missing.

    Background: yfinance's API stopped consistently surfacing the synthetic
    'Free Cash Flow' row for many large-cap tickers after its 2023 rewrite.
    It still returns 'Operating Cash Flow' and 'Capital Expenditure'
    separately, leaving callers to derive FCF themselves. Without this
    fallback the DCF stage silently skipped ~80% of the universe (every
    mega-cap included — AAPL, MSFT, NVDA, GOOGL, ...).

    Capex in yfinance is signed negative, so FCF = OCF + Capex.

    Args:
        cf: pandas DataFrame from yfinance — rows are line items, columns
            are reporting periods.

    Returns:
        pd.Series indexed by period (sorted ascending), or None if neither
        path can produce values.
    """
    if cf is None or cf.empty:
        return None
    # Preferred: yfinance's pre-computed synthetic row when present.
    if 'Free Cash Flow' in cf.index:
        fcf = cf.loc['Free Cash Flow'].dropna().sort_index()
        if len(fcf) > 0:
            return fcf
    # Fallback: derive FCF = OCF + Capex (Capex stored as a negative
    # number in yfinance, so addition gives FCF).
    ocf_row = None
    for key in OPERATING_CF_KEYS:
        if key in cf.index:
            candidate = cf.loc[key].dropna().sort_index()
            if len(candidate) > 0:
                ocf_row = candidate
                break
    if ocf_row is None:
        return None
    capex_row = None
    for key in CAPEX_KEYS:
        if key in cf.index:
            candidate = cf.loc[key].dropna().sort_index()
            if len(candidate) > 0:
                capex_row = candidate
                break
    if capex_row is None:
        return None
    common = ocf_row.index.intersection(capex_row.index)
    if len(common) == 0:
        return None
    fcf = (ocf_row.loc[common] + capex_row.loc[common]).sort_index().dropna()
    if len(fcf) == 0:
        return None
    return fcf


def _compute_shareholder_yield(yf_data, mcap):
    """Compute total shareholder yield = (dividends + buybacks) / market cap.

    Uses the most recent fiscal year from cash flow statement, with a
    ticker.info fallback for dividends when the cash-flow Dividend Paid row
    is missing (Yahoo leaves it blank for ~78% of tickers even when the
    company genuinely pays a dividend).

    Returns dict with 'shareholder_yield' and 'buyback_rate' (both decimal), or None.
    """
    if not mcap or mcap <= 0:
        return None
    cf = yf_data.get('cash_flow')
    info = yf_data.get('info') or {}
    cf_present = cf is not None and not (hasattr(cf, 'empty') and cf.empty)

    div_paid = 0
    buyback = 0
    issuance = 0
    if cf_present:
        latest = cf.iloc[:, 0]
        for k in ['Common Stock Dividend Paid', 'Cash Dividends Paid']:
            if k in latest.index and pd.notna(latest[k]):
                div_paid = abs(float(latest[k]))
                break
        for k in ['Repurchase Of Capital Stock', 'Common Stock Payments']:
            if k in latest.index and pd.notna(latest[k]):
                buyback = abs(float(latest[k]))
                break
        # Subtract any new issuance to get net buyback. Allow negative values:
        # a net-diluting company has buyback_rate < 0 and shareholder_yield is
        # reduced (or negated) accordingly. Flooring at zero would mask dilution.
        for k in ['Issuance Of Capital Stock', 'Common Stock Issuance']:
            if k in latest.index and pd.notna(latest[k]):
                issuance = abs(float(latest[k]))
                break

    # Dividend fallback: dividendRate (annual $/share) × sharesOutstanding gives
    # total annual dividend cash, equivalent to the cash-flow Dividend Paid row.
    # Apply whenever the cash-flow signal is missing or zero — dividendRate
    # comes from info, a different yfinance endpoint that's far more reliable.
    if div_paid == 0:
        div_rate = info.get('dividendRate')
        shares_out = info.get('sharesOutstanding')
        if div_rate and shares_out and div_rate > 0 and shares_out > 0:
            div_paid = float(div_rate) * float(shares_out)

    # Distinguish "no data" from "no shareholder returns". If cash-flow is
    # entirely absent AND dividendRate didn't fire, return None so the column
    # shows as missing rather than a misleading 0.
    if not cf_present and div_paid == 0:
        return None

    net_buyback = buyback - issuance
    total_return = div_paid + net_buyback
    shareholder_yield = total_return / mcap
    buyback_rate = net_buyback / mcap

    # Sanity cap — |yield| > 50% is almost certainly a data error (stale
    # yfinance cash-flow snapshot with mismatched period/mcap units, or a
    # one-time issuance event distorting a single fiscal year).
    if abs(shareholder_yield) > 0.50:
        shareholder_yield = None
        buyback_rate = None

    return {'shareholder_yield': shareholder_yield, 'buyback_rate': buyback_rate}


# ---------------------------------------------------------------------------
# Balance-sheet debt levels (Worksheet Step 3C)
# ---------------------------------------------------------------------------

def _debt_levels(yf_data):
    """Point-in-time debt levels for the row (USD — statement frames are
    FX-normalized upstream). Statements first; yfinance .info as fallback
    only, per field (its totalDebt/totalCash are MRQ figures — fine as
    point-in-time levels, never injected into annual frames). None stays
    None: unknown ≠ 0.

    Returns (total_debt, cash, total_liabilities, net_debt, source).
    `source` records where the *known* levels came from: None = statements
    only, 'yf_info' = every known level fell back to .info,
    'statements+yf_info' = mixed. net_debt is total_debt − cash only when
    both sides are known.
    """
    total_debt_val = cash_val = total_liabilities_val = None
    _bs = yf_data.get('balance_sheet')
    if _bs is not None and not _bs.empty:
        _latest_bs = _bs.iloc[:, 0]
        total_debt_val = _get(_latest_bs, DEBT_KEYS)
        cash_val = _get(_latest_bs, CASH_KEYS)
        total_liabilities_val = _get(
            _latest_bs,
            ['Total Liabilities Net Minority Interest', 'Total Liab'])
    _yf_info = yf_data.get('info') or {}
    _fallback = []
    if total_debt_val is None and _yf_info.get('totalDebt') is not None:
        total_debt_val = _yf_info.get('totalDebt')
        _fallback.append('debt')
    if cash_val is None and _yf_info.get('totalCash') is not None:
        cash_val = _yf_info.get('totalCash')
        _fallback.append('cash')
    if not _fallback:
        debt_source = None
    elif len(_fallback) == 2:
        debt_source = 'yf_info'
    else:
        _stmt_val = cash_val if _fallback == ['debt'] else total_debt_val
        debt_source = 'statements+yf_info' if _stmt_val is not None else 'yf_info'
    net_debt_val = (total_debt_val - cash_val
                    if total_debt_val is not None and cash_val is not None
                    else None)
    return total_debt_val, cash_val, total_liabilities_val, net_debt_val, debt_source


# ---------------------------------------------------------------------------
# Forward DCF (Worksheet Step 5A)
# ---------------------------------------------------------------------------

def run_forward_dcf(yf_data, wacc, sector=None, exit_multiple=None, roic_data=None,
                    terminal_growth_adj=0.0, wacc_sigma=None, tg_sigma=None,
                    growth_sigma_mult=1.0, growth_weight_shift=0.0):
    """Run a two-stage 10-year DCF with sector-specific parameters.

    Includes: sector growth caps, averaged FCF for cyclicals,
    owner-earnings adjustment, FCF yield mean-reversion ceiling,
    fundamental growth signal, exit multiple cross-check, and
    Monte Carlo uncertainty quantification.

    Args:
        terminal_growth_adj: Additive adjustment to terminal growth (macro overlay).
        wacc_sigma: Override for MC WACC sigma (defaults to MC_WACC_SIGMA).
        tg_sigma: Override for MC terminal growth sigma (defaults to MC_TERMINAL_GROWTH_SIGMA).
        growth_sigma_mult: Multiplicative adjustment to growth sigma (macro overlay).
        growth_weight_shift: Shift from analyst LT weight to fundamental weight (macro overlay).

    Returns:
        Tuple of (fair_value, sensitivity_range, fcf_growth, growth_diag, mc_result),
        or (None, None, None, {}, None) on insufficient data.
    """
    if wacc_sigma is None:
        wacc_sigma = MC_WACC_SIGMA
    if tg_sigma is None:
        tg_sigma = MC_TERMINAL_GROWTH_SIGMA
    cf = yf_data.get('cash_flow')
    info = yf_data.get('info') or {}

    if cf is None or cf.empty:
        return None, None, None, {}, None

    cfg = _get_sector_config(sector)
    term_g = cfg['terminal_growth'] + terminal_growth_adj

    if wacc is None or wacc <= term_g:
        return None, None, None, {}, None

    # Use the shared FCF extractor: prefers yfinance's pre-computed
    # 'Free Cash Flow' row, falls back to OCF + Capex when missing.
    fcf_series = _fcf_series_from_cashflow(cf)
    if fcf_series is None:
        return None, None, None, {}, None
    fcf_values = fcf_series.values.tolist()
    if not fcf_values:
        return None, None, None, {}, None

    # --- Fix E: For cyclical sectors, average multiple years of FCF ---
    avg_years = cfg['avg_fcf_years']
    if avg_years > 1 and len(fcf_values) >= 2:
        recent = [v for v in fcf_values[-avg_years:] if v > 0]
        base_fcf = sum(recent) / len(recent) if recent else fcf_values[-1]
    else:
        base_fcf = fcf_values[-1]

    # Latest cash-flow column (newest), used for the owner-earnings capex
    # add-back and the SBC deduction below.
    try:
        latest_cf = cf[sorted(cf.columns)[-1]]
    except Exception as e:
        logger.debug(f"cash-flow columns unsortable ({e}); using positional newest column")
        latest_cf = cf.iloc[:, -1] if cf.columns.is_monotonic_increasing else cf.iloc[:, 0]

    def _cf_line(labels):
        for lbl in labels:
            if lbl in latest_cf.index and pd.notna(latest_cf[lbl]):
                return latest_cf[lbl]
        return None

    # --- Mean-reversion: cap FCF at the firm's OWN trailing normal ---
    # Peak-cycle FCF shouldn't be extrapolated. Cap at YIELD_CEILING_MULT ×
    # the trailing average of positive FCF — a fundamental anchor. (The old
    # ceiling capped at mcap × a yield, which was price-circular.) Applied
    # HERE, before the owner-earnings adjustments, so the cap and the base
    # are on the same as-reported basis — capping after the growth-capex
    # add-back / SBC deduction silently compared adjusted FCF against a
    # raw-FCF average, undoing the add-back for exactly the capex-heavy
    # firms it exists for.
    _pos_fcf = [v for v in fcf_values if v > 0]
    if _pos_fcf and base_fcf > 0:
        trailing_avg_fcf = sum(_pos_fcf) / len(_pos_fcf)
        fcf_ceiling = YIELD_CEILING_MULT * trailing_avg_fcf
        if fcf_ceiling > 0 and base_fcf > fcf_ceiling:
            base_fcf = fcf_ceiling

    # --- Fix F: Growth capex add-back for capex-heavy companies ---
    # If Capex > 2× D&A, significant growth capex is depressing accounting FCF.
    # Add back 50% of the capex above that 2× D&A band. Measuring the excess
    # from the SAME threshold that triggers the adjustment keeps fair value
    # continuous in capex: at exactly 2× D&A the add-back is zero and it ramps
    # from there. (The excess was previously measured from 1× D&A while the
    # trigger sat at 2×, so the add-back jumped from 0 to 0.5× D&A the instant
    # the ratio crossed 2.0 — a $0.02 capex change swung fair value ~25%, and
    # spending MORE on capex could RAISE fair value. Same cliff class as the
    # old all-or-nothing SBC guard below.) Still more conservative than full
    # owner earnings (OCF - D&A), which over-inflates companies like GOOG
    # where capex is partially maintenance.
    if cfg['check_owner_earnings']:
        ocf = _cf_line(['Operating Cash Flow', 'Total Cash From Operating Activities'])
        da = _cf_line(['Depreciation And Amortization', 'Depreciation Amortization Depletion'])
        _cap = _cf_line(['Capital Expenditure', 'Capital Expenditures'])
        capex = abs(_cap) if _cap is not None else None

        if ocf and da and capex and da > 0 and capex / da > CAPEX_DA_THRESHOLD:
            # Capex above the maintenance band (2× D&A) is growth capex.
            # Add back a portion — assume half is truly discretionary growth.
            excess_capex = capex - CAPEX_DA_THRESHOLD * da
            growth_add_back = excess_capex * EXCESS_CAPEX_ADDBACK
            adjusted = base_fcf + growth_add_back
            if adjusted > 0:
                base_fcf = adjusted

    # --- Owner earnings: deduct stock-based compensation (all sectors) ---
    # OCF adds SBC back as a non-cash charge, but it is a real cost to
    # shareholders (dilution). Buffett's owner earnings subtracts it. Pairs
    # with the SBC-dilution GATE (which scores the behavior) — this corrects
    # the VALUATION. Monotone: heavier SBC always means a bigger haircut,
    # floored at 25% of pre-SBC FCF so an extreme SBC year haircuts rather
    # than aborts the valuation. (The previous all-or-nothing guard skipped
    # the deduction entirely whenever SBC ≥ FCF — the worst offenders were
    # the only firms that paid NO penalty, and a $1 change in SBC could
    # swing fair value ~75%.)
    _sbc = _cf_line(['Stock Based Compensation', 'StockBasedCompensation',
                     'Share Based Compensation'])
    if _sbc is not None and base_fcf > 0:
        _sbc = abs(float(_sbc))
        if _sbc > 0:
            base_fcf = max(base_fcf - _sbc, base_fcf * 0.25)

    if base_fcf <= 0:
        return None, None, None, {}, None

    # --- FCF growth estimation ---
    fcf_cagr = None
    if len(fcf_values) >= 2 and fcf_values[0] > 0 and fcf_values[-1] > 0:
        n = len(fcf_values) - 1
        fcf_cagr = (fcf_values[-1] / fcf_values[0]) ** (1 / n) - 1

    # Revenue CAGR as secondary signal
    rev_cagr = None
    inc = yf_data.get('income_statement')
    if inc is None or (hasattr(inc, 'empty') and inc.empty):
        inc = yf_data.get('income_stmt')
    if inc is not None and not (hasattr(inc, 'empty') and inc.empty):
        if 'Total Revenue' in inc.index:
            rev_series = inc.loc['Total Revenue'].dropna().sort_index()
            revs = [v for v in rev_series.values if v and v > 0]
            if len(revs) >= 2:
                # revs sorted ascending: [oldest, ..., newest]
                rev_cagr = (revs[-1] / revs[0]) ** (1 / (len(revs) - 1)) - 1

    analyst_st = info.get('revenueGrowth')          # short-term (1yr revenue)
    analyst_lt = _get_analyst_lt_growth(yf_data)      # long-term (~5yr)
    earnings_g = _get_earnings_growth(yf_data)        # 1yr earnings growth

    # --- Sector-specific growth cap with hyper-growth override ---
    # Raise the cap only when the firm's OWN trailing revenue growth
    # corroborates the analyst optimism — not based on how expensive the stock
    # is (the old actual_yield < HYPER_GROWTH_YIELD trigger let only
    # richly-priced names through, embedding market price in the estimate).
    best_analyst = analyst_lt or analyst_st
    growth_cap = cfg['growth_cap']
    if (best_analyst is not None and best_analyst > growth_cap
            and rev_cagr is not None and rev_cagr > growth_cap):
        growth_cap = max(growth_cap, min(HYPER_GROWTH_CAP, best_analyst * ANALYST_HAIRCUT))

    # --- 6-signal weighted average (auto-normalise when signals missing) ---
    growth_signals = []
    growth_weights = []
    if fcf_cagr is not None:
        growth_signals.append(fcf_cagr)
        growth_weights.append(GROWTH_WEIGHT_FCF)
    if rev_cagr is not None:
        growth_signals.append(rev_cagr)
        growth_weights.append(GROWTH_WEIGHT_REV)
    # Optimism haircut on analyst-sourced signals — the config comment
    # promised this ("Apply 20% haircut to analyst growth estimate") but only
    # the hyper-growth CAP applied it; the signals themselves (60% of nominal
    # blend weight) entered raw. Trim positive estimates toward realism;
    # negative forecasts pass through (no rosy bias to remove).
    def _haircut(x):
        return x * ANALYST_HAIRCUT if (x is not None and x > 0) else x

    if analyst_st is not None:
        growth_signals.append(_haircut(analyst_st))
        growth_weights.append(GROWTH_WEIGHT_ANALYST_ST)
    # Apply macro growth weight shift: move weight from analyst LT to fundamental
    _w_analyst_lt = max(0.05, GROWTH_WEIGHT_ANALYST_LT + growth_weight_shift)
    _w_fundamental = max(0.05, GROWTH_WEIGHT_FUNDAMENTAL - growth_weight_shift)
    if analyst_lt is not None:
        growth_signals.append(_haircut(analyst_lt))
        growth_weights.append(_w_analyst_lt)
    if earnings_g is not None:
        growth_signals.append(_haircut(earnings_g))
        growth_weights.append(GROWTH_WEIGHT_EARNINGS_G)

    # Signal 6: Fundamental growth (Reinvestment Rate × ROIC)
    fund_result = calculate_fundamental_growth(yf_data,
                    roic_override=roic_data.get('avg_roic') if roic_data else None)
    fundamental_g = fund_result.get('fundamental_growth')
    if fundamental_g is not None:
        growth_signals.append(fundamental_g)
        growth_weights.append(_w_fundamental)

    if growth_signals:
        total_weight = sum(growth_weights)
        weighted_avg = sum(s * w for s, w in zip(growth_signals, growth_weights, strict=False)) / total_weight
    else:
        weighted_avg = FALLBACK_GROWTH

    # --- Margin trend adjustment ---
    margin_trend = _compute_margin_trend(yf_data)
    if margin_trend is not None:
        weighted_avg += margin_trend * MARGIN_TREND_SENSITIVITY

    # --- Earnings surprise adjustment ---
    surprise_adj, surprise_avg = _compute_surprise_adjustment(yf_data)
    weighted_avg += surprise_adj

    # Floor at −15% (noise control), NOT at terminal growth: flooring at
    # term_g made decline literally unrepresentable, inflating the FV of
    # deteriorating businesses and rendering two_stage_ev's negative-growth
    # warning unreachable. A genuinely shrinking firm is now modeled shrinking.
    fcf_growth = min(growth_cap, max(-0.15, weighted_avg))

    # --- Two-stage DCF via shared model function (GGM terminal value) ---
    ggm_valuation = two_stage_ev_valuation(base_fcf, fcf_growth, wacc, term_g,
                                           total_years=DCF_YEARS, stage1_years=DCF_STAGE1)
    ev_ggm = ggm_valuation.value
    if ev_ggm is None or ev_ggm <= 0:
        return None, None, None, {}, None

    # None = balance sheet absent (leverage unknown). The DCF's EV→equity
    # bridge treats it as 0 (the historical behavior for this cohort, which
    # usually also lacks the FCF to compute a DCF at all).
    net_debt = get_net_debt(yf_data) or 0
    shares = info.get('sharesOutstanding')
    fv_ggm = fair_value_per_share(ev_ggm, net_debt, shares)

    # --- Exit multiple cross-check ---
    exit_mult_fv = None
    tv_method_spread = None
    base_ebitda = None

    # Extract base EBITDA = Operating Income + D&A
    inc_stmt = yf_data.get('income_statement')
    if inc_stmt is not None and not inc_stmt.empty:
        try:
            sorted_inc_cols = sorted(inc_stmt.columns)
            latest_inc = inc_stmt[sorted_inc_cols[-1]]
        except Exception as e:
            logger.debug(f"income-statement columns unsortable ({e}); using positional column 0")
            latest_inc = inc_stmt.iloc[:, 0]
        op_inc_val = None
        for k in ['Operating Income', 'Total Operating Income As Reported']:
            if k in latest_inc.index and pd.notna(latest_inc[k]):
                op_inc_val = float(latest_inc[k])
                break
        da_val = None
        cf_for_da = yf_data.get('cash_flow')
        if cf_for_da is not None and not cf_for_da.empty:
            try:
                sorted_cf_cols = sorted(cf_for_da.columns)
                latest_cf_da = cf_for_da[sorted_cf_cols[-1]]
            except Exception as e:
                logger.debug(f"cash-flow columns unsortable for D&A ({e}); using positional column 0")
                latest_cf_da = cf_for_da.iloc[:, 0]
            for k in ['Depreciation And Amortization', 'Depreciation Amortization Depletion']:
                if k in latest_cf_da.index and pd.notna(latest_cf_da[k]):
                    da_val = abs(float(latest_cf_da[k]))
                    break
        if op_inc_val and da_val:
            base_ebitda = op_inc_val + da_val

    if base_ebitda and base_ebitda > 0 and exit_multiple and shares and shares > 0:
        exit_valuation = two_stage_ev_exit_multiple_valuation(
            base_fcf, fcf_growth, wacc, term_g,
            base_ebitda, exit_multiple,
            total_years=DCF_YEARS, stage1_years=DCF_STAGE1)
        ev_exit = exit_valuation.value
        if ev_exit and ev_exit > 0:
            exit_mult_fv = fair_value_per_share(ev_exit, net_debt, shares)

    # Average GGM and exit multiple FVs
    fv = fv_ggm
    if fv_ggm and exit_mult_fv:
        fv = (fv_ggm + exit_mult_fv) / 2.0
        avg_fv = (fv_ggm + exit_mult_fv) / 2.0
        tv_method_spread = abs(fv_ggm - exit_mult_fv) / avg_fv if avg_fv > 0 else None
    elif exit_mult_fv:
        fv = exit_mult_fv

    # Sensitivity range (supplementary)
    sens_range = None
    if shares and shares > 0:
        sens = dcf_sensitivity(base_fcf, fcf_growth, wacc, term_g,
                               net_debt, shares, years=DCF_YEARS, stage1=DCF_STAGE1)
        vals = [v for v in sens.values() if v is not None]
        if vals:
            sens_range = (min(vals), max(vals))

    # --- Monte Carlo uncertainty quantification ---
    mc_result = None
    if base_fcf > 0 and shares and shares > 0:
        g_sigma = abs(fcf_growth) * MC_GROWTH_SIGMA_RATIO if fcf_growth != 0 else 0.02
        g_sigma *= growth_sigma_mult  # macro overlay: widen in stress
        em_sigma = exit_multiple * MC_EXIT_MULT_SIGMA_RATIO if exit_multiple else None
        # Widen sigma if TV methods diverge significantly
        if tv_method_spread and tv_method_spread > EXIT_MULT_DIVERGENCE_THRESHOLD:
            g_sigma *= MC_HIGH_DIVERGENCE_SIGMA_MULT
            if em_sigma:
                em_sigma *= MC_HIGH_DIVERGENCE_SIGMA_MULT

        mc_result = monte_carlo_dcf(
            base_fcf, fcf_growth, wacc, term_g,
            net_debt, shares,
            base_ebitda=base_ebitda, exit_multiple=exit_multiple,
            n_iterations=MC_ITERATIONS,
            growth_sigma=g_sigma, wacc_sigma=wacc_sigma,
            tg_sigma=tg_sigma, exit_mult_sigma=em_sigma,
            exit_mult_floor=EXIT_MULT_MIN,
            total_years=DCF_YEARS, stage1_years=DCF_STAGE1)

    # Use MC percentiles for bear/bull instead of sensitivity grid
    if mc_result:
        sens_range = (mc_result['p10_fv'], mc_result['p90_fv'])

    growth_diag = {
        'analyst_ltg': analyst_lt,
        'earnings_growth': earnings_g,
        'margin_trend': margin_trend,
        'surprise_avg': surprise_avg,
        'fundamental_growth': fundamental_g,
        'reinvestment_rate': fund_result.get('reinvestment_rate'),
        'exit_mult_fv': exit_mult_fv,
        'tv_method_spread': tv_method_spread,
        'mc_result': mc_result,
        # Exposed so reverse_dcf can solve on the SAME adjusted FCF base and
        # terminal growth as the forward DCF — otherwise implied_vs_estimated
        # conflates a basis mismatch with a genuine expectation gap.
        'base_fcf': base_fcf,
        'term_g': term_g,
        # Valuation envelope of the primary GGM leg: which path produced the
        # number, how much its own soft warnings degrade trust in it, and the
        # concrete caveats — so a fallback can't pose as authoritative.
        'dcf_method': ggm_valuation.method,
        'dcf_confidence': ggm_valuation.confidence,
        'dcf_warnings': list(ggm_valuation.warnings),
    }
    return fv, sens_range, fcf_growth, growth_diag, mc_result


# ---------------------------------------------------------------------------
# Dividend Discount Model helper
# ---------------------------------------------------------------------------

def _annualise_dividends(div_series, as_of_year=None):
    """Convert a per-payment dividend Series to annual DPS (oldest first).

    Groups by calendar year and sums. Two corrections vs a naive groupby:

    - The current (partial) calendar year is dropped: a mid-year run would
      otherwise read a quarterly payer's year-to-date sum as a full-year
      DPS, collapsing the dividend CAGR toward the −10% floor for most of
      the dividend universe.
    - Interior skipped years are filled with 0.0 rather than vanishing, so a
      suspension is visible to the consecutive-year eligibility screen and
      doesn't silently compress the CAGR time span.
    """
    if div_series is None or len(div_series) == 0:
        return []
    # Normalise DataFrame → Series (yfinance >=1.2 may return a single-column
    # DataFrame from stock.dividends instead of the expected Series).
    if isinstance(div_series, pd.DataFrame):
        div_series = div_series.iloc[:, 0] if not div_series.empty else pd.Series(dtype=float)
    if len(div_series) == 0:
        return []
    annual = div_series.groupby(div_series.index.year).sum().sort_index()
    if as_of_year is None:
        as_of_year = date.today().year
    # Drop the current partial year (keep it only if it's the sole year, so a
    # freshly-initiated payer still surfaces something).
    if as_of_year in annual.index and len(annual) > 1:
        annual = annual.drop(index=as_of_year)
    if annual.empty:
        return []
    # Reindex to a contiguous year range, filling suspension gaps with 0.
    full_range = range(int(annual.index.min()), int(annual.index.max()) + 1)
    annual = annual.reindex(full_range, fill_value=0.0)
    return annual.tolist()


def run_ddm_valuation(yf_data, div_series, cost_of_equity, analyst_ltg=None,
                      terminal_growth_adj=0.0):
    """Run Dividend Discount Model valuation for a single stock.

    Args:
        yf_data: Dict of financial data from YFinanceClient.
        div_series: Pandas Series of historical dividends (from fetch_dividends).
        cost_of_equity: Required return (from CAPM / select_cost_of_equity).
        analyst_ltg: Analyst long-term growth estimate (optional).
        terminal_growth_adj: Macro-overlay additive adjustment to the terminal
            growth rate, so a regime shift moves the DDM cohort in step with
            the DCF cohort (which already receives it).

    Returns:
        Dict with DDM results or dict with eligible=False for non-payers.
    """
    info = (yf_data.get('info') or {}) if yf_data else {}
    eps = info.get('trailingEps')
    dps = info.get('dividendRate')
    payout = info.get('payoutRatio')
    roe = None
    # Compute ROE from balance sheet if available
    bs = yf_data.get('balance_sheet') if yf_data else None
    inc = yf_data.get('income_statement') if yf_data else None
    if bs is not None and not bs.empty and inc is not None and not inc.empty:
        try:
            sorted_bs = sorted(bs.columns)
            equity = bs[sorted_bs[-1]].get('Stockholders Equity')
            sorted_inc = sorted(inc.columns)
            ni = inc[sorted_inc[-1]].get('Net Income')
            if equity and ni and equity > 0:
                roe = ni / equity
        except Exception as e:
            logger.debug(f"DDM ROE inputs unusable ({e}); retention inference will warn instead")

    annual_divs = _annualise_dividends(div_series)

    # 1. Eligibility check
    elig = ddm_eligibility(annual_divs, payout, eps, dps)
    result = {
        'ddm_eligible': elig['eligible'],
        'ddm_reason': elig['reason'],
        'ddm_consecutive_years': elig['consecutive_years'],
        'ddm_payout_flag': elig['payout_flag'],
        'ddm_fv': None,
        'ddm_h_fv': None,
        'ddm_growth': None,
        'ddm_div_cagr': None,
        'ddm_sustainable_growth': None,
        'ddm_mc_median': None,
        'ddm_mc_p10': None,
        'ddm_mc_p90': None,
        'ddm_mc_cv': None,
    }

    if not elig['eligible']:
        return result

    # 2. Growth estimation
    growth_est = estimate_ddm_growth(annual_divs, payout, roe, analyst_ltg)
    g = growth_est['growth']
    if g is None:
        g = TERMINAL_GROWTH_RATE  # fallback to terminal growth
    result['ddm_growth'] = g
    result['ddm_div_cagr'] = growth_est['div_cagr']
    result['ddm_sustainable_growth'] = growth_est['sustainable_growth']

    re = cost_of_equity
    tg = TERMINAL_GROWTH_RATE + terminal_growth_adj

    # 3. Two-stage DDM
    ddm_valuation = two_stage_ddm_valuation(dps, g, tg, re, years=DDM_HIGH_GROWTH_YEARS)
    ddm_fv = ddm_valuation.value
    result['ddm_fv'] = ddm_fv
    result['ddm_confidence'] = ddm_valuation.confidence if ddm_fv is not None else None
    result['ddm_warnings'] = list(ddm_valuation.warnings)

    # 4. H-Model cross-check. H is HALF the linear-decline period, so to model
    # growth fading over the same horizon the two-stage holds it (5y),
    # half_life = 5/2. Passing the full 5 modeled a 10-year fade and inflated
    # the H-model leg (and hence the averaged ddm_fv) whenever g > tg.
    h_fv = ddm_h_model_valuation(dps, g, tg, re,
                                 half_life=DDM_HIGH_GROWTH_YEARS / 2.0).value
    result['ddm_h_fv'] = h_fv

    # Average the two methods when both available
    if ddm_fv and h_fv:
        result['ddm_fv'] = (ddm_fv + h_fv) / 2.0
    elif h_fv:
        result['ddm_fv'] = h_fv

    # 5. Monte Carlo DDM
    mc = monte_carlo_ddm(dps, g, re, tg, n=MC_ITERATIONS, years=DDM_HIGH_GROWTH_YEARS)
    if mc:
        result['ddm_mc_median'] = mc['median_fv']
        result['ddm_mc_p10'] = mc['p10_fv']
        result['ddm_mc_p90'] = mc['p90_fv']
        result['ddm_mc_cv'] = mc['cv']

    return result


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

MCAP_MISSING_ALERT_THRESHOLD = 0.02


def apply_mcap_integrity_guard(results, prior_rows=None,
                               threshold=MCAP_MISSING_ALERT_THRESHOLD):
    """Recover missing market caps, then shout if too many are still gone.

    ``mcap`` / ``shares_out`` sit upstream of nearly every valuation output, so
    losing them nulls p_tbv, fcf_yield, shareholder_yield, mos, fv_dispersion,
    pfcf, net_cash_to_mcap, tangible_book_per_share, all five fair-value models
    and the Monte Carlo confidence fields. Because a missing-data N/A still
    counts against the applicable-gate denominator (see the two-kinds-of-N/A
    note in scripts/scoring.py), an affected row is scored as *failing* five
    valuation gates on absent data rather than skipping them. On 2026-07-29
    that hit 265 of 2,244 records (11.8% against a ~0.1% baseline) and left
    large caps like MA and LLY rated PASS/HOLD on data that was never fetched.

    The upstream cause is fixed in ``data.yfinance_client`` (yfinance 1.3.0's
    ``.info`` intermittently drops both fields; ``fast_info`` still has them).
    This is the second line of defence: fall back to the prior snapshot's share
    count re-priced at today's price, flag every row it touches via
    ``_mcap_source`` so the fallback can never masquerade as fresh data, and
    print a prominent warning when the residual miss rate clears *threshold*.

    Must run BEFORE ``score_and_rate`` or the recovered values never reach the
    gates. Mutates *results* in place; returns a summary dict for callers/tests.
    """
    # --- Implausibly HIGH caps: corruption, not data ---------------------
    # Yahoo hands preferred / OTC lines the parent's common share count
    # (every Fannie/Freddie preferred series carries 5.7B / 3.2B shares), so
    # FNMFO — a $50,000-par preferred quoted ~$31,500 — reported a $180T cap
    # on 2026-08-12 and rated PASS. The primary fix nulls this at fetch time
    # (data.yfinance_client._sanitize_implausible_mcap); this is the second
    # line for rows built from snapshots cached before that fix. Nulling
    # (rather than clamping) routes the row into the missing-mcap machinery
    # below — and the share count is nulled with it when shares x price
    # reproduces the same absurd figure, since price is directly observed.
    implausible = [r for r in results
                   if (r.get('mcap') or 0) > MCAP_MAX_PLAUSIBLE]
    for r in implausible:
        bad = r['mcap']
        r['mcap'] = None
        shares, price = r.get('shares_out'), r.get('price')
        if shares and price and float(shares) * float(price) > MCAP_MAX_PLAUSIBLE:
            r['shares_out'] = None
        r['_mcap_invalid'] = bad
        print(f"[mcap] {r.get('ticker')}: implausible market cap "
              f"${bad:.3g} (> ${MCAP_MAX_PLAUSIBLE:.0g}) nulled")

    missing = [r for r in results if not r.get('mcap')]
    if missing and prior_rows:
        prior_by_ticker = {r['ticker']: r for r in prior_rows if r.get('ticker')}
        for r in missing:
            prior = prior_by_ticker.get(r.get('ticker'))
            if not prior:
                continue
            # Share counts move slowly, so yesterday's count at today's price
            # is a far better estimate than nothing. Only fall back to the
            # prior mcap wholesale when today's price is unavailable too.
            # A candidate above MCAP_MAX_PLAUSIBLE means the prior snapshot
            # carries the same corruption — refuse it and leave the row
            # missing rather than resurrect the garbage just nulled above.
            shares = r.get('shares_out') or prior.get('shares_out')
            price = r.get('price')
            if (shares and price
                    and float(price) * float(shares) <= MCAP_MAX_PLAUSIBLE):
                r['shares_out'] = shares
                r['mcap'] = float(price) * float(shares)
                r['_mcap_source'] = 'prior_snapshot_shares'
            elif prior.get('mcap') and prior['mcap'] <= MCAP_MAX_PLAUSIBLE:
                r['mcap'] = prior['mcap']
                r['shares_out'] = r.get('shares_out') or prior.get('shares_out')
                r['_mcap_source'] = 'prior_snapshot_mcap'

    still_missing = [r for r in results if not r.get('mcap')]
    recovered = sum(1 for r in results if r.get('_mcap_source'))
    miss_pct = (len(still_missing) / len(results)) if results else 0.0

    if recovered:
        print(f"\n[mcap] recovered {recovered} record(s) from the prior "
              f"snapshot (flagged via _mcap_source)")
    if miss_pct > threshold:
        print("\n" + "!" * 70)
        print(f"!! DATA QUALITY WARNING: {len(still_missing)} of {len(results)} "
              f"records ({miss_pct:.1%}) have no market cap.")
        print("!! Baseline is ~0.1%. Every valuation gate for these rows "
              "(P/TBV, FCF Yield,")
        print("!! Shrhldr Yld, MoS, FV Dispersion) scores as FAILED on absent "
              "data, so")
        print("!! their ratings are NOT trustworthy. Check the info backfill "
              "in")
        print("!! data/yfinance_client.py before acting on this run.")
        print("!" * 70)
    elif still_missing:
        print(f"\n[mcap] {len(still_missing)} record(s) without market cap "
              f"({miss_pct:.1%}, within the ~0.1% baseline)")

    return {'recovered': recovered, 'still_missing': len(still_missing),
            'miss_pct': miss_pct, 'alert': miss_pct > threshold,
            'implausible_nulled': len(implausible)}


class _ModelWarningCounter(logging.Filter):
    """Counts model-layer warnings routed through logging.captureWarnings.

    Feeds the end-of-run quality summary: total warnings, and how many
    flagged a fabricated input (fallback rates, placeholder costs) that
    every downstream valuation silently inherits.
    """

    def __init__(self):
        super().__init__()
        self.total = 0
        self.fabricated = 0

    def filter(self, record):
        self.total += 1
        msg = record.getMessage()
        if 'fabricated' in msg or 'fallback' in msg:
            self.fabricated += 1
        return True


def _run_setup():
    """CLI parsing, provenance recorder, and logging/warning configuration."""
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument('--input', '-i', help='Excel file with tickers in first column')
    parser.add_argument('--validation', '-v',
                        help='Validation Excel (poor performers, same column structure as --input)')
    parser.add_argument('--macro', action='store_true',
                        help='Enable macro-economic overlay (adjusts ERP, growth, sigma based on market regime)')
    parser.add_argument('--prices-dir', default='output/prices',
                        help='Directory of per-ticker Parquet price files for rolling beta, '
                             'realized vol, momentum, and drawdown (default: output/prices)')
    parser.add_argument('--universe', choices=('sp500', 'us'), default='sp500',
                        help='Ticker universe: sp500 (S&P 500 + Dow, ~500 stocks, default) or '
                             'us (all US-listed equities via SEC EDGAR, ~7000-8000 stocks).')
    parser.add_argument('--sec-email', default=os.environ.get('SEC_EMAIL', 'stockanalysis@example.com'),
                        help='Contact email for SEC EDGAR User-Agent (used by --universe us). '
                             'Override via --sec-email or SEC_EMAIL env var.')
    parser.add_argument('--min-spread', type=float, default=None, metavar='FRAC',
                        help='Phase-1 filter: skip tickers where ROIC - WACC < FRAC. '
                             'Use 0 to keep only value-creating businesses (ROIC > WACC). '
                             'Tickers where spread cannot be computed are also skipped. '
                             'Recommended with --universe us to keep Phase-2 manageable.')
    parser.add_argument('--mcap-min', type=float, default=0, metavar='DOLLARS',
                        help='Phase-1 filter: skip tickers with market cap below this threshold '
                             '(e.g. 500e6 for $500M). Default 0 = no filter. '
                             'Useful with --universe us to drop shells and micro-caps quickly.')
    args = parser.parse_args()
    prices_dir = args.prices_dir if os.path.isdir(args.prices_dir) else None
    run_start_date = date.today()
    _prov = ProvenanceRecorder(run_start_date)

    # Observability: timestamped diagnostics on stderr (report output stays on
    # stdout), model warnings routed through logging, and — critically —
    # RuntimeWarnings forced to 'always'. Python's default once-per-location
    # filter would show a model warning for the first ticker and silently
    # suppress it for the other ~2,000.
    logging.basicConfig(
        stream=sys.stderr,
        level=logging.INFO,
        format='%(asctime)s %(levelname)s %(name)s: %(message)s',
    )
    warnings.simplefilter('always', RuntimeWarning)
    logging.captureWarnings(True)
    _model_warning_counter = _ModelWarningCounter()
    logging.getLogger('py.warnings').addFilter(_model_warning_counter)
    return {'args': args, 'prices_dir': prices_dir,
            'run_start_date': run_start_date, '_prov': _prov,
            '_model_warning_counter': _model_warning_counter}


def _run_macro_setup(args, prices_dir):
    """Risk-free rate fetch plus the opt-in macro-economic overlay."""
    # Fetch live risk-free rate (10-yr Treasury yield)
    risk_free_rate = fetch_risk_free_rate()
    from data import treasury_rate as _treasury
    risk_free_rate_source = _treasury.last_rate_source or 'live'
    print(f"Risk-free rate: {risk_free_rate:.2%} (10-yr Treasury, "
          f"source={risk_free_rate_source})")

    # --- Macro-economic overlay (opt-in via --macro) ---
    macro_regime_result = None
    macro_adj = None
    effective_erp = ERP
    effective_tg_adj = 0.0
    effective_wacc_sigma = MC_WACC_SIGMA
    effective_growth_sigma_mult = 1.0
    effective_exit_mult_adj = 0.0
    effective_growth_weight_shift = 0.0
    sector_signals = {}
    commodity_data = {}
    sector_etf_data = {}
    local_rs = None

    if args.macro:
        try:
            mc_client = MacroClient()
            macro_indicators = mc_client.fetch_macro_indicators()
            macro_regime_result = assess_macro_regime(macro_indicators)
            macro_adj = compute_macro_adjustments(macro_regime_result)
            print_macro_summary(macro_regime_result, macro_adj, base_erp=ERP)

            effective_erp = ERP + macro_adj['erp_adjustment']
            effective_tg_adj = macro_adj['terminal_growth_adjustment']
            effective_wacc_sigma = MC_WACC_SIGMA + macro_adj['wacc_sigma_adjustment']
            effective_growth_sigma_mult = macro_adj['growth_sigma_multiplier']
            effective_exit_mult_adj = macro_adj['exit_mult_adjustment']
            effective_growth_weight_shift = macro_adj['growth_weight_shift']

            # Sector headwind/tailwind analysis
            sector_etf_data = mc_client.fetch_sector_data()
            local_rs = None
            if prices_dir:
                try:
                    local_rs = compute_sector_rs_from_local(prices_dir)
                except Exception as _rs_exc:
                    print(f"  Sector RS from local prices failed ({_rs_exc}), skipping.")
            sector_signals = generate_sector_signals(sector_etf_data, macro_regime_result, local_rs=local_rs)
            # Commodity & cross-sector data for stock-level narrative
            commodity_data = mc_client.fetch_commodity_data()
        except Exception as e:
            print(f"  Macro overlay failed ({e}), proceeding with defaults.")
    return {'risk_free_rate': risk_free_rate,
            'risk_free_rate_source': risk_free_rate_source,
            'macro_regime_result': macro_regime_result,
            'macro_adj': macro_adj,
            'effective_erp': effective_erp,
            'effective_tg_adj': effective_tg_adj,
            'effective_wacc_sigma': effective_wacc_sigma,
            'effective_growth_sigma_mult': effective_growth_sigma_mult,
            'effective_exit_mult_adj': effective_exit_mult_adj,
            'effective_growth_weight_shift': effective_growth_weight_shift,
            'sector_signals': sector_signals,
            'commodity_data': commodity_data,
            'sector_etf_data': sector_etf_data,
            'local_rs': local_rs}


def _run_build_universe(args):
    """Assemble the ticker universe plus Morningstar P/FV and source-group maps."""
    ms_pfv_data = {}  # Morningstar Price/Fair Value ratios (if input file has them)
    ticker_source = {}  # ticker -> 'quality' | 'poor'

    def _load_pfv_from_xlsx(path, pfv_dict):
        """Extract tickers and P/FV ratios from a Morningstar-format xlsx."""
        import openpyxl as _ox
        wb = _ox.load_workbook(path)
        ws = wb[wb.sheetnames[0]]
        tickers = sorted(set(
            str(ws.cell(r, 1).value).strip()
            for r in range(2, ws.max_row + 1)
            if ws.cell(r, 1).value
        ))
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        pfv_col = None
        for ci, h in enumerate(headers):
            if h and 'Price/Fair Value' in str(h):
                pfv_col = ci + 1
                break
        n_pfv = 0
        if pfv_col:
            for r in range(2, ws.max_row + 1):
                tk = ws.cell(r, 1).value
                pfv = ws.cell(r, pfv_col).value
                if tk and pfv and isinstance(pfv, (int, float)) and pfv > 0:
                    pfv_dict[str(tk).strip()] = pfv
                    n_pfv += 1
        return tickers, n_pfv

    # Always start with the full SP500/NYSE/DOW universe
    sp500 = set(get_sp500_tickers())
    nyse = set(get_nyse_tickers())
    dow = set(get_dow_tickers())
    all_tickers = sorted(sp500 | nyse | dow)

    # Optional broader universe from SEC EDGAR (--universe us)
    if args.universe == 'us':
        from data.us_listings import fetch_us_listed_tickers
        print(f"Fetching US-listed universe from SEC EDGAR (User-Agent: {args.sec_email})...")
        us_tickers = fetch_us_listed_tickers(email=args.sec_email)
        before = len(all_tickers)
        all_tickers = sorted(set(all_tickers) | set(us_tickers))
        print(f"  US-listed: {len(us_tickers):,} tickers; universe expanded {before} -> {len(all_tickers):,}")

    for t in all_tickers:
        ticker_source[t] = 'quality'

    if args.input:
        # Merge input file tickers + P/FV data on top of the universe
        input_tickers, n_pfv = _load_pfv_from_xlsx(args.input, ms_pfv_data)
        existing = set(all_tickers)
        extra = [t for t in input_tickers if t not in existing]
        all_tickers = all_tickers + extra
        for t in extra:
            ticker_source[t] = 'quality'
        if n_pfv:
            print(f"Universe: {len(existing)} tickers + {len(extra)} extra from input "
                  f"+ {n_pfv} MS Price/FV ratios from {args.input}")
        else:
            print(f"Universe: {len(existing)} tickers + {len(extra)} extra from input "
                  f"(no Price/Fair Value column found in {args.input})")

    if args.validation:
        val_tickers, val_n_pfv = _load_pfv_from_xlsx(args.validation, ms_pfv_data)
        for t in val_tickers:
            ticker_source[t] = 'poor'
        existing = set(all_tickers)
        all_tickers = all_tickers + [t for t in val_tickers if t not in existing]
        print(f"Loaded {len(val_tickers)} validation tickers + {val_n_pfv} P/FV from {args.validation} "
              f"(combined universe: {len(all_tickers)} tickers, {len(ms_pfv_data)} total P/FV)")
    return {'ms_pfv_data': ms_pfv_data, 'ticker_source': ticker_source,
            'all_tickers': all_tickers}


def _run_build_clients(run_start_date):
    """Construct the Phase-1 data clients (yfinance, Tiingo, SEC EDGAR)."""
    yf_client = YFinanceClient(run_date=run_start_date)

    # Tiingo client initialized here so it's available for Phase 1 beta calculation
    tiingo_client = TiingoClient(request_delay=0.5)
    if tiingo_client.available:
        print('Tiingo API configured — using as primary news source.')
    else:
        print('Tiingo API not configured (set TIINGO_API_KEY) — falling back to yfinance/Google RSS.')

    # SEC EDGAR clients initialized here (rather than at Phase-2 setup) so the
    # SECXBRLClient is available as a Phase-1 fallback when yfinance returns
    # an empty payload (Yahoo soft-throttle). The CIK-map load is idempotent.
    sec_client = SECLegalClient(email='stockanalysis@example.com', request_delay=1.0)
    sec_client._load_cik_map()
    sec_xbrl_client = SECXBRLClient(
        cik_map=sec_client._cik_map,
        name_map=sec_client._name_map,
        email='stockanalysis@example.com',
        request_delay=1.0,
    )
    return {'yf_client': yf_client, 'tiingo_client': tiingo_client,
            'sec_client': sec_client, 'sec_xbrl_client': sec_xbrl_client}


def _run_phase1_screen(args, _prov, all_tickers, ticker_source, yf_client,
                       tiingo_client, sec_xbrl_client, risk_free_rate,
                       effective_erp):
    """Phase 1: screen the full universe, caching fundamentals for Phase 2."""
    # -----------------------------------------------------------------------
    # Phase 1: Collect data for full universe (no ROIC > WACC pre-filter)
    # -----------------------------------------------------------------------
    _skip_path = os.path.join(os.path.dirname(__file__), '..', 'data', 'skip_tickers.txt')
    _skip_set = set()
    try:
        with open(_skip_path, encoding='utf-8') as _sf:
            for _line in _sf:
                _t = _line.split('#')[0].strip().upper()
                if _t:
                    _skip_set.add(_t)
        if _skip_set:
            print(f"Skipping {len(_skip_set)} ticker(s) from skip_tickers.txt: {', '.join(sorted(_skip_set))}")
    except FileNotFoundError:
        pass

    # Tickers that appeared in the most recent prior run bypass the mcap/spread
    # Phase-1 filters so that previously-scored stocks are always re-evaluated.
    # This maintains backtest continuity even when a stock dips marginally below
    # a filter threshold due to data fluctuation.
    _carry_set = set()
    # Prior rows are retained (not just their tickers) so the market-cap
    # integrity guard further down can fall back to yesterday's share count.
    _carry_prior_rows = []
    try:
        import glob as _glob
        _prior_jsons = sorted(_glob.glob(os.path.join('output', 'results_*.json')))
        if _prior_jsons:
            _prior_path = _prior_jsons[-1]
            with open(_prior_path, encoding='utf-8') as _pf:
                _prior = json.load(_pf)
            _prior_rows = _prior.get('results', _prior) if isinstance(_prior, dict) else _prior
            _carry_prior_rows = _prior_rows
            _carry_set = {r['ticker'] for r in _prior_rows if r.get('ticker')} - _skip_set
            print(f"Carry-forward: {len(_carry_set)} ticker(s) from {os.path.basename(_prior_path)} "
                  f"will bypass Phase-1 filters")
            # Also ensure carry-forward tickers are in the universe (they may
            # have been excluded by the --universe sp500 scope or the SEC list).
            _existing = set(all_tickers)
            _extra_carried = [t for t in sorted(_carry_set) if t not in _existing]
            if _extra_carried:
                all_tickers = all_tickers + _extra_carried
                for _t in _extra_carried:
                    ticker_source[_t] = 'quality'
                print(f"  Added {len(_extra_carried)} carry-forward ticker(s) not already in universe")
    except Exception as _ce:
        print(f"[warn] carry-forward load failed: {_ce}")

    print(f"Processing {len(all_tickers)} tickers (full universe)...")
    _universe_n = len(all_tickers)
    # Fetch-failure retry: when BOTH sources return nothing, the cause is
    # usually a transient outage window, not the ticker (2026-08-18 dropped
    # DE/DECK/DELL/DEO/DG in one contiguous alphabetical block this way).
    # Such tickers are re-queued once at the end of the pass — by the time
    # the queue drains, hours have passed and the outage has cleared.
    # Neither client caches these failures (EmptyYahooResponseError is
    # raised before the yfinance cache write; the SEC client deliberately
    # skips caching request failures), so the retry re-fetches for real.
    _fetch_retry_queued = set()
    _fetch_retry_failed = set()
    qualifying = []
    screen_cache = {}
    screen_outcomes = {'quality': {'total': 0, 'passed': 0},
                       'poor': {'total': 0, 'passed': 0}}

    for i, ticker in enumerate(all_tickers, 1):
        if ticker in _skip_set:
            print(f"  [{i}/{len(all_tickers)}] {ticker} - SKIP (skip_tickers.txt)")
            sys.stdout.flush()
            continue
        _grp = ticker_source.get(ticker, 'quality')
        screen_outcomes[_grp]['total'] += 1
        try:
            # ----------------------------------------------------------------
            # Fundamentals fetch — SEC XBRL is the primary source for the
            # financial statements (authoritative, direct from filings);
            # yfinance is the primary source for market-data `info` (market
            # cap, beta, current price, dividends) and analyst series.
            # Merge them so downstream code (CAPM, market-cap-weighted WACC,
            # mcap filter) keeps working unchanged.
            #
            # Source tags written to screen_cache for audit / later gating:
            #   'sec_xbrl+yfinance'  XBRL statements + yfinance info (US filer)
            #   'yfinance'           yfinance both (foreign / OTC, no CIK)
            #   'sec_xbrl'           XBRL only (yfinance throttled for a US filer)
            # ----------------------------------------------------------------
            try:
                yf_data = yf_client.fetch_financials(ticker)
            except EmptyYahooResponseError:
                yf_data = None

            # Early mcap bail before paying the XBRL fetch cost. For ~6K
            # micro-caps the mcap filter will drop the ticker anyway —
            # no point spending 1 second on a SEC fetch we'll discard.
            # Carry-forwards bypass the mcap filter, so they still proceed.
            _carried = ticker in _carry_set
            if (yf_data is not None and args.mcap_min and not _carried):
                _early_mcap = (yf_data.get('info') or {}).get('marketCap') or 0
                if _early_mcap < args.mcap_min:
                    print(f"  [{i}/{len(all_tickers)}] {ticker} - "
                          f"SKIP mcap ${_early_mcap/1e6:.0f}M < "
                          f"${args.mcap_min/1e6:.0f}M floor")
                    sys.stdout.flush()
                    continue

            xbrl_data = None
            sec_prov = None
            _xbrl_attempted = ticker in sec_xbrl_client._cik_map
            if _xbrl_attempted:
                try:
                    # No year_limit: use all available XBRL history (10-16
                    # years). Through-cycle ROIC is the right input for a
                    # DCF / value pipeline with a long terminal-value horizon.
                    xbrl_data = sec_xbrl_client.build_yfinance_shape(ticker)
                except Exception as e:
                    logger.warning(f"SEC XBRL: build_yfinance_shape failed for {ticker}: {e}; falling back to yfinance statements")
                    xbrl_data = None  # network hiccup — fall through
                if xbrl_data is not None:
                    # Filing metadata must be read here: Phase 2 evicts the
                    # companyfacts blob from the client's memory cache.
                    sec_prov = sec_xbrl_client.get_filing_provenance(ticker)

            if xbrl_data is not None and yf_data is not None:
                yf_data = {
                    'balance_sheet':    xbrl_data['balance_sheet'],
                    'income_statement': xbrl_data['income_statement'],
                    'cash_flow':        xbrl_data['cash_flow'],
                    'info':             yf_data.get('info') or {},
                    'growth_estimates': yf_data.get('growth_estimates'),
                    'earnings_history': yf_data.get('earnings_history'),
                }
                _data_source = 'sec_xbrl+yfinance'
            elif xbrl_data is not None:
                # yfinance throttled for a US filer — XBRL-only path.
                # info is sparse → CAPM falls to buildup, WACC uses book equity.
                yf_data = xbrl_data
                _data_source = 'sec_xbrl'
                _prov.record_event('source_fallback', ticker, 'yfinance',
                                   {'from': 'yfinance', 'to': 'sec_xbrl',
                                    'reason': 'yfinance empty for US filer'})
            elif yf_data is not None:
                # No CIK (foreign / OTC) — yfinance is the only source.
                _data_source = 'yfinance'
                if _xbrl_attempted:
                    # Distinguish a failed SEC fetch from a successful fetch
                    # whose facts couldn't build usable statements (SPACs,
                    # thin filers): fetch_company_facts caches the blob iff
                    # the request succeeded.
                    _fb_reason = ('insufficient XBRL data'
                                  if sec_xbrl_client._cache.get(ticker) is not None
                                  else 'sec_xbrl fetch failed')
                    _prov.record_event('source_fallback', ticker, 'sec_xbrl',
                                       {'from': 'sec_xbrl', 'to': 'yfinance',
                                        'reason': _fb_reason})
            else:
                if ticker not in _fetch_retry_queued:
                    # First failure: re-queue at the end of the pass and undo
                    # this attempt's screen_outcomes count (the retry attempt
                    # re-increments it, so each ticker is counted once).
                    _fetch_retry_queued.add(ticker)
                    all_tickers.append(ticker)
                    screen_outcomes[_grp]['total'] -= 1
                    print(f"  [{i}/{len(all_tickers)}] {ticker} - "
                          "error: yfinance empty AND no SEC XBRL coverage "
                          "— re-queued for retry")
                else:
                    _fetch_retry_failed.add(ticker)
                    print(f"  [{i}/{len(all_tickers)}] {ticker} - "
                          "error: yfinance empty AND no SEC XBRL coverage "
                          "(retry also failed)")
                sys.stdout.flush()
                continue

            _prov.record_source(ticker, 'statements', _data_source, sec=sec_prov)
            _prov.record_source(
                ticker, 'market_data',
                'sec_xbrl' if _data_source == 'sec_xbrl' else 'yfinance')

            # FX normalization — convert local-currency financials + price/
            # mcap to USD before any model runs. Skips USD-reporting tickers
            # automatically; logs a warning flag on the row if the rate
            # lookup fails so we can audit divergent valuations later.
            # Statements from SEC XBRL are already USD — don't convert them
            # again even if yfinance reports a foreign financialCurrency.
            _stmts_usd = _data_source in ('sec_xbrl', 'sec_xbrl+yfinance')
            yf_data, fx_meta = _convert_financials_to_usd(
                yf_data, statements_are_usd=_stmts_usd)
            _prov.record_source(ticker, 'fx', 'fx_client',
                                converted=fx_meta.get('fx_converted', False))
            if fx_meta.get('fx_fetch_failed'):
                _prov.record_event('fx_fetch_failed', ticker, 'fx_client',
                                   {'currency_financial': fx_meta.get('currency_financial')})

            info = yf_data.get('info') or {}
            sector = info.get('sector', '')

            roic_data = calculate_roic(yf_data)
            cost_of_equity, re_method, beta_diag = select_cost_of_equity(
                yf_data, risk_free_rate, yf_client, ticker, erp=effective_erp,
                tiingo_client=tiingo_client)
            _prov.record_source(ticker, 'beta', re_method)
            if re_method != 'capm':
                # Anything but the locally regressed beta is a fallback:
                # Yahoo's unvetted beta, GGM, or the build-up.
                _prov.record_event('source_fallback', ticker, 'beta',
                                   {'method': re_method,
                                    'reason': 'local 5y weekly beta unavailable'})
            wacc = calculate_wacc(yf_data, cost_of_equity,
                                  risk_free_rate=risk_free_rate)
            if wacc is not None:
                s_cfg = _get_sector_config(sector)
                wacc = max(s_cfg['wacc_floor'], min(s_cfg['wacc_cap'], wacc))

            spread = (roic_data['avg_roic'] - wacc
                      if (roic_data and wacc is not None) else None)

            # --- Phase-1 filters (applied before expensive Phase-2 work) ---
            mcap = info.get('marketCap') or 0
            roic_str = f"ROIC {roic_data['avg_roic']:.1%} " if roic_data else "ROIC N/A "
            wacc_str = f"WACC {wacc:.1%} " if wacc is not None else "WACC N/A "
            spread_str = f"spread {spread:.1%}" if spread is not None else "spread N/A"

            _carried = ticker in _carry_set
            if args.mcap_min and mcap < args.mcap_min and not _carried:
                print(f"  [{i}/{len(all_tickers)}] {ticker} - SKIP mcap ${mcap/1e6:.0f}M < ${args.mcap_min/1e6:.0f}M floor")
                sys.stdout.flush()
                continue

            # Bypass the ROIC>WACC spread filter for Financial Services.
            # The standard ROIC formula (NOPAT / (Equity + Debt − Cash))
            # is economically meaningless for banks/insurers — they are
            # capital intermediaries, not capital deployers. Without the
            # bypass, names like BAC / C / KEY / RF / GS / MS get
            # silently filtered out by an unreliable computation. Bank
            # quality is evaluated downstream via NIM / Efficiency / CET1
            # / NPL from FDIC call reports (see scripts/enrich_fdic.py).
            _is_financial = (info.get('sector') == 'Financial Services')
            if args.min_spread is not None and not _carried and not _is_financial:
                if spread is None or spread < args.min_spread:
                    label = "no spread" if spread is None else f"spread {spread:.1%}"
                    print(f"  [{i}/{len(all_tickers)}] {ticker} - SKIP {label} < min {args.min_spread:.1%}")
                    sys.stdout.flush()
                    continue

            qualifying.append(ticker)
            screen_outcomes[_grp]['passed'] += 1
            screen_cache[ticker] = {
                'roic_data': roic_data, 'wacc': wacc,
                'cost_of_equity': cost_of_equity,
                're_method': re_method, 'yf_data': yf_data,
                'beta_diag': beta_diag,
                'data_source': _data_source,
                'fx_meta': fx_meta,
            }
            print(f"  [{i}/{len(all_tickers)}] {ticker} - {roic_str}{wacc_str}{spread_str} [{re_method}] <{_data_source}>")

        except Exception as e:
            print(f"  [{i}/{len(all_tickers)}] {ticker} - error: {e}")
        # Flush after every ticker so the log reflects progress if OOM-killed
        sys.stdout.flush()

    # Free memory: drop ALL cached financials and price histories.
    # Qualifying tickers' data survives via screen_cache references.
    yf_client.evict_financials()
    yf_client.clear_history_cache()
    gc.collect()

    print(f"\n{len(qualifying)} tickers collected out of {_universe_n} total.")
    if _fetch_retry_queued:
        _recovered = len(_fetch_retry_queued) - len(_fetch_retry_failed)
        print(f"  Fetch-failure retry: {len(_fetch_retry_queued)} re-queued, "
              f"{_recovered} recovered, {len(_fetch_retry_failed)} failed twice")
        if _fetch_retry_failed:
            print(f"  Failed twice: {', '.join(sorted(_fetch_retry_failed))}")
    if args.validation:
        for grp in ('quality', 'poor'):
            o = screen_outcomes[grp]
            rate = o['passed'] / o['total'] if o['total'] > 0 else 0
            print(f"  {grp:>8}: {o['passed']}/{o['total']} passed ({rate:.0%})")
    print()
    return {'qualifying': qualifying, 'screen_cache': screen_cache,
            'screen_outcomes': screen_outcomes,
            '_carry_prior_rows': _carry_prior_rows}


def _run_sector_exit_multiples(qualifying, screen_cache,
                               effective_exit_mult_adj):
    """Pre-compute sector median EV/EBITDA exit multiples (macro-adjusted)."""
    # Pre-compute sector median EV/EBITDA for exit multiple cross-check
    _pre_sector_ee = {}
    for ticker in qualifying:
        cached_pre = screen_cache[ticker]
        info_pre = (cached_pre['yf_data'].get('info') or {})
        ee_pre = info_pre.get('enterpriseToEbitda')
        sector_pre = info_pre.get('sector', '')
        if ee_pre and 0 < ee_pre < EV_EBITDA_OUTLIER_MAX:
            _pre_sector_ee.setdefault(sector_pre, []).append(ee_pre)
    sector_exit_multiples = {}
    for s, v in _pre_sector_ee.items():
        if len(v) >= MIN_SECTOR_STOCKS:
            med = sorted(v)[len(v) // 2]
            sector_exit_multiples[s] = max(EXIT_MULT_MIN, min(EXIT_MULT_MAX, med))

    # Apply macro exit multiple adjustment
    effective_exit_mult_default = max(EXIT_MULT_MIN,
        min(EXIT_MULT_MAX, EXIT_MULT_DEFAULT_EV_EBITDA + effective_exit_mult_adj))
    if effective_exit_mult_adj != 0.0:
        for s in sector_exit_multiples:
            sector_exit_multiples[s] = max(EXIT_MULT_MIN,
                min(EXIT_MULT_MAX, sector_exit_multiples[s] + effective_exit_mult_adj))
    return {'sector_exit_multiples': sector_exit_multiples,
            'effective_exit_mult_default': effective_exit_mult_default}


def _run_build_phase2_clients(sec_client, qualifying, screen_cache):
    """Construct the Phase-2 clients (news, supply chain, insider, culture)."""
    # -----------------------------------------------------------------------
    # News pipeline: Tiingo (primary) + yfinance/Google RSS (fallback)
    # -----------------------------------------------------------------------
    news_client = NewsClient(request_delay=1.0, max_age_days=30)
    _sectors_for_news = set(
        (screen_cache[t]['yf_data'].get('info') or {}).get('sector', '')
        for t in qualifying
    )
    news_client.prefetch_all_sectors(_sectors_for_news)

    # sec_client and sec_xbrl_client are initialized earlier (before Phase 1)
    # so the XBRL fallback is available during the universe screen.

    # Finnhub: supply chain relationships
    supply_client = FinnhubSupplyClient(request_delay=1.0)
    if supply_client.available:
        print('Finnhub supply chain API configured.')
    else:
        print('Finnhub supply chain API not configured (set FINNHUB_API_KEY).')

    # SEC EDGAR: supply chain extraction from 10-K filings (free fallback)
    sec_supply_client = SECSupplyClient(
        cik_map=sec_client._cik_map,
        name_map=sec_client._name_map,
        email='stockanalysis@example.com',
        request_delay=1.0,
    )

    # SEC EDGAR: insider transaction tracking from Form 4 filings
    sec_insider_client = SECInsiderClient(
        cik_map=sec_client._cik_map,
        name_map=sec_client._name_map,
        email='stockanalysis@example.com',
        request_delay=1.0,
        max_form4_files=15,
    )

    # Culture metrics client (no external API — derives signals from yfinance)
    culture_client = CultureClient()
    return {'news_client': news_client, 'supply_client': supply_client,
            'sec_supply_client': sec_supply_client,
            'sec_insider_client': sec_insider_client,
            'culture_client': culture_client}


def _run_phase2_analysis(qualifying, screen_cache, prices_dir,
                         ticker_source, ms_pfv_data, _prov, yf_client,
                         tiingo_client, news_client, sec_client,
                         supply_client, sec_supply_client, sec_xbrl_client,
                         sec_insider_client, culture_client,
                         sector_exit_multiples, effective_exit_mult_default,
                         effective_erp, effective_tg_adj,
                         effective_wacc_sigma, effective_growth_sigma_mult,
                         effective_growth_weight_shift, risk_free_rate,
                         macro_regime_result, sector_signals):
    """Phase 2: full per-ticker analysis of every qualifying ticker."""
    # -----------------------------------------------------------------------
    # Phase 2: Full analysis on qualifying tickers (Worksheet Steps 2-5)
    # -----------------------------------------------------------------------

    # Pre-load SPY local prices once for rolling-beta comparisons
    _spy_local = _load_local_prices('SPY', prices_dir)

    results = []
    for ticker in qualifying:
        print(f"Analyzing {ticker}...")
        try:
            cached = screen_cache[ticker]
            yf_data = cached['yf_data']
            wacc = cached['wacc']
            roic_data = cached['roic_data']
            cost_of_equity = cached['cost_of_equity']
            beta_diag = cached.get('beta_diag')
            fx_meta = cached.get('fx_meta') or {}

            # Phase 2 assumes roic_data and wacc are populated. Carry-forward
            # tickers bypass the Phase-1 spread filter, so we can land here
            # with either missing — skip the ticker rather than crash.
            if not roic_data or wacc is None:
                print(f"  Skipping {ticker}: ROIC or WACC unavailable today")
                continue

            # --- Price-history enrichments (local Parquet) ---
            # One read serves both the legacy signals and the market/risk
            # metrics; fall back to the Close-only loader for the ~2% of files
            # that lack a Volume column.
            _local_ohlcv = _load_local_ohlcv(ticker, prices_dir)
            _local_close = (_local_ohlcv['Close'] if _local_ohlcv is not None
                            else _load_local_prices(ticker, prices_dir))
            _price_metrics = price_metrics_from_series(_local_ohlcv, _local_close)
            _ticker_realized_vol = None
            _ticker_dd_2008      = None
            _ticker_dd_2020      = None
            _ticker_dd_2022      = None
            # Rolling betas come from phase 1 (weekly, same frame as the
            # headline beta) whether or not a local price file exists.
            _rolling_beta_diag   = (beta_diag or {}).get('rolling_betas') or {}

            if _local_close is not None and len(_local_close) > 60:
                # 1. Realized volatility (252-day) → replaces fixed MC_WACC_SIGMA
                #    Not staleness-gated: it feeds the MC WACC sigma, and a
                #    slightly stale vol estimate is better than reverting to
                #    the fixed default.
                _ticker_realized_vol = _realized_vol(_local_close)

                # 2. 12-minus-1 month momentum now comes from _price_metrics,
                #    which suppresses it when the price file is stale.

                # 3. Max drawdown in key stress periods
                _ticker_dd_2008 = _max_drawdown_period(_local_close, '2008-01-01', '2009-03-31')
                _ticker_dd_2020 = _max_drawdown_period(_local_close, '2020-01-01', '2020-09-30')
                _ticker_dd_2022 = _max_drawdown_period(_local_close, '2022-01-01', '2022-12-31')

                # 4. Rolling beta across 1y / 3y / 5y windows. Only fall back
                #    to the local parquet closes when phase 1 produced none,
                #    using the same weekly method so the windows stay
                #    comparable with the headline beta.
                if not _rolling_beta_diag and _spy_local is not None:
                    _rolling_beta_diag = _rolling_betas_from_prices(_local_close, _spy_local)
                    if _rolling_beta_diag:
                        if beta_diag is None:
                            beta_diag = {}
                        beta_diag['rolling_betas'] = _rolling_beta_diag
                        beta_diag['rolling_source'] = 'parquet'

            # Use ticker realized vol for MC WACC sigma (floor at macro-adjusted base)
            _effective_wacc_sigma_ticker = effective_wacc_sigma
            if _ticker_realized_vol is not None:
                # WACC sigma ≈ 30% of realized equity vol (equity → WACC dampening)
                _rv_wacc = _ticker_realized_vol * 0.30
                _effective_wacc_sigma_ticker = max(effective_wacc_sigma, min(_rv_wacc, 0.04))

            # Gross margin (moat gate) and ROIC consistency
            gross_margin = _compute_gross_margin(yf_data)
            # Latest financials for profit pool analysis
            latest_fins = _extract_latest_financials(yf_data)
            roic_cv = None
            roic_years = roic_data.get('roic_by_year', {})
            if len(roic_years) >= 2:
                vals = list(roic_years.values())
                mean_r = sum(vals) / len(vals)
                if mean_r > 0:
                    var_r = sum((x - mean_r) ** 2 for x in vals) / (len(vals) - 1)
                    roic_cv = (var_r ** 0.5) / mean_r

            # Company description and CEO (Worksheet Step 3)
            info = yf_data.get('info') or {}
            # mcap must be read from THIS ticker's info, not the stale Phase 1 variable
            mcap = info.get('marketCap') or 0

            # Shareholder yield (dividends + buybacks) / market cap
            sy_result = _compute_shareholder_yield(yf_data, mcap)
            shareholder_yield = sy_result['shareholder_yield'] if sy_result else None
            share_buyback_rate = sy_result['buyback_rate'] if sy_result else None
            description = info.get('longBusinessSummary') or ''
            company_name = info.get('shortName') or info.get('longName') or ''
            sector = info.get('sector') or ''
            industry = info.get('industry') or ''
            country = info.get('country') or ''
            # Discount-rate discipline: floor/cap the cost of equity with the
            # same sector bounds that bound WACC, so the equity-flow models
            # (DDM, RIM, EPV growth premium) can't discount at a raw GGM/
            # build-up Re below the blended-capital floor while the DCF is
            # floored. Diagnostic 'er' keeps the raw estimate.
            re_for_models = cost_of_equity
            if cost_of_equity is not None:
                _re_cfg = _get_sector_config(sector)
                re_for_models = max(_re_cfg['wacc_floor'],
                                    min(_re_cfg['wacc_cap'], cost_of_equity))
            # Tiingo is primary news source; fall back to yfinance/Google RSS
            if tiingo_client.available:
                tiingo_news = tiingo_client.fetch_ticker_news(ticker, max_age_days=30, max_items=12)
            else:
                tiingo_news = []
            if tiingo_news:
                ticker_news = tiingo_news
                news_sentiment = tiingo_client.fetch_ticker_sentiment(ticker, max_age_days=30, max_items=12)
            else:
                ticker_news = news_client.get_combined_news(ticker, sector, max_total=12)
                news_sentiment = None
            legal_data = sec_client.fetch_legal_filings(ticker, days_back=730)
            supply_data = supply_client.fetch_supply_chain(ticker)
            if not supply_data.get('available'):
                supply_data = sec_supply_client.fetch_supply_chain(ticker)
            finnhub_peers = supply_client.fetch_peers(ticker)

            # SEC EDGAR: XBRL cross-validation. Skip when statements already
            # came from XBRL — validating XBRL against itself is meaningless.
            _stmt_source = (screen_cache.get(ticker) or {}).get('data_source', 'yfinance')
            if _stmt_source in ('sec_xbrl+yfinance', 'sec_xbrl'):
                xbrl_validation = None
            else:
                xbrl_validation = sec_xbrl_client.validate_against_yfinance(ticker, yf_data)
                if xbrl_validation and xbrl_validation.get('fields_flagged', 0) > 0:
                    _prov.record_event(
                        'cross_source_conflict', ticker, 'sec_xbrl',
                        {'fields_flagged': xbrl_validation.get('fields_flagged'),
                         'edgar_quality_score': xbrl_validation.get('edgar_quality_score')})
            # SEC EDGAR: long-duration revenue/earnings history
            edgar_history = sec_xbrl_client.fetch_historical_financials(ticker, min_years=10)
            # Evict the raw XBRL JSON blob (~1-10 MB) now that both
            # validate_against_yfinance and fetch_historical_financials are done.
            sec_xbrl_client._cache.pop(ticker, None)
            # SEC EDGAR: insider transactions from Form 4
            insider_data = sec_insider_client.fetch_insider_activity(ticker, days_back=365)

            officers = info.get('companyOfficers') or []
            ceo_officer = next(
                (o for o in officers
                 if 'ceo' in (o.get('title') or '').lower() or
                    'chief executive' in (o.get('title') or '').lower()),
                officers[0] if officers else None
            )
            ceo = ceo_officer.get('name') if ceo_officer else None
            # Build CEO biography from available yfinance officer data
            ceo_bio = None
            if ceo_officer:
                bio_parts = [ceo_officer.get('name', 'N/A'), ceo_officer.get('title', '')]
                age = ceo_officer.get('age')
                if age:
                    bio_parts.append(f"Age {age}")
                year_born = ceo_officer.get('yearBorn')
                if year_born and not age:
                    bio_parts.append(f"Born {year_born}")
                total_pay = ceo_officer.get('totalPay')
                if total_pay:
                    if total_pay >= 1e6:
                        pay_str = f"${total_pay/1e6:.1f}M"
                    else:
                        pay_str = f"${total_pay:,.0f}"
                    fy = ceo_officer.get('fiscalYear', '')
                    bio_parts.append(f"Compensation: {pay_str}" + (f" (FY{fy})" if fy else ""))
                ceo_bio = " | ".join(p for p in bio_parts if p)

            # Culture raw metrics: employees, CEO pay, comp risk, SBC
            _culture_raw = culture_client.extract(info, yf_data)
            _culture_gd = culture_client.fetch_glassdoor(company_name, ticker)

            # Step 2: Relative multiples
            multiples = compute_relative_multiples(yf_data)
            current_price = multiples.get('price')
            shares = multiples.get('shares')

            # Time-series cheapness: current price/EBIT vs own ~10y median
            # (Valuation: Mult vs Hist gate — replaced EPV Floor, r=0.68 vs MoS)
            mult_vs_hist, mult_hist_years = compute_multiple_vs_history(
                _local_close, edgar_history, latest_fins.get('operating_income'))

            # Analyst consensus (Worksheet Step 8)
            analyst = compute_analyst_consensus(yf_data)
            # Next scheduled earnings date (from the same cached info fetch)
            next_earn = extract_next_earnings(yf_data)

            # Step 5A: Forward DCF (sector-aware: Fixes C/D/E/F)
            dcf_fv, dcf_sens_range, fcf_growth, growth_diag, mc_result = run_forward_dcf(
                yf_data, wacc, sector=sector,
                exit_multiple=sector_exit_multiples.get(sector, effective_exit_mult_default),
                roic_data=roic_data,
                terminal_growth_adj=effective_tg_adj,
                wacc_sigma=_effective_wacc_sigma_ticker,
                tg_sigma=MC_TERMINAL_GROWTH_SIGMA,
                growth_sigma_mult=effective_growth_sigma_mult,
                growth_weight_shift=effective_growth_weight_shift)
            mos = (dcf_fv - current_price) / dcf_fv if (dcf_fv and current_price and dcf_fv > 0) else None

            # Step 5B: Dividend Discount Model (for dividend payers)
            try:
                div_series = yf_client.fetch_dividends(ticker)
            except Exception as e:
                logger.warning(f"yfinance: dividend history fetch failed for {ticker}: {e}; DDM skipped")
                div_series = pd.Series(dtype=float)
            ddm_result = run_ddm_valuation(
                yf_data, div_series, re_for_models,
                analyst_ltg=growth_diag.get('analyst_ltg'),
                terminal_growth_adj=effective_tg_adj)

            # Step 3A/3B: Earnings quality
            eq = calculate_earnings_quality(yf_data)

            # Step 3B: Piotroski F
            piotroski = calculate_piotroski_f(yf_data)

            # Revenue CAGR (3Y from yfinance)
            rev_cagr = calculate_revenue_cagr(yf_data)

            # Extended CAGRs and derived metrics from EDGAR history.
            # Same logic also runs from rescore_and_render so a snapshot
            # whose edgar_history was backfilled after the live run gets
            # refreshed signals without re-running analyze_stock end-to-end.
            _edgar_metrics = derive_edgar_metrics(edgar_history)
            rev_cagr_5y         = _edgar_metrics['rev_cagr_5y']
            rev_cagr_10y        = _edgar_metrics['rev_cagr_10y']
            fcf_cagr_5y         = _edgar_metrics['fcf_cagr_5y']
            fcf_cagr_10y        = _edgar_metrics['fcf_cagr_10y']
            gross_margin_avg_5y = _edgar_metrics['gross_margin_avg_5y']
            gross_margin_trend  = _edgar_metrics['gross_margin_trend']
            dividend_cagr_5y    = _edgar_metrics['dividend_cagr_5y']
            shares_cagr_5y      = _edgar_metrics['shares_cagr_5y']

            # Step 3C: Balance sheet health
            int_cov = calculate_interest_coverage(yf_data)
            nd_ebitda = calculate_net_debt_ebitda(yf_data)

            # Debt levels — see _debt_levels for the statements-first /
            # per-field .info fallback and provenance rules.
            (total_debt_val, cash_val, total_liabilities_val,
             net_debt_val, debt_source) = _debt_levels(yf_data)

            # Traditional ratios
            ratios = compute_ratios(yf_data)

            # Free cash flow — use annual cash flow statement (same source as DCF).
            # Falls back to OCF + Capex when yfinance doesn't surface the
            # synthetic 'Free Cash Flow' row (common for large-caps post-2023).
            cf = yf_data.get('cash_flow')
            fcf = None
            fcf_series = _fcf_series_from_cashflow(cf)
            if fcf_series is not None and len(fcf_series) > 0:
                fcf = fcf_series.iloc[-1]  # most recent annual

            # --- NEW MODELS ---

            # Altman Z-Score (exists in comparisons, now wired)
            altman_z = calculate_altman_z(yf_data)
            altman_z_zone = None
            if altman_z is not None:
                if altman_z > 2.99:
                    altman_z_zone = 'safe'
                elif altman_z >= 1.81:
                    altman_z_zone = 'grey'
                else:
                    altman_z_zone = 'distress'

            # Beneish M-Score
            beneish = calculate_beneish_m(yf_data)

            # DuPont Decomposition
            dupont = compute_dupont(yf_data)

            # EPV (Earnings Power Value)
            inc_stmt = yf_data.get('income_statement')
            _epv_ebit = None
            _epv_eff_tax = 0.21
            _epv_yf_revenue = None
            if inc_stmt is not None and not inc_stmt.empty:
                _latest_inc = inc_stmt.iloc[:, 0]
                _epv_ebit = _latest_inc.get('Operating Income')
                _rev_val = _latest_inc.get('Total Revenue')
                if pd.notna(_rev_val) and _rev_val is not None:
                    _epv_yf_revenue = float(_rev_val)
                if pd.notna(_epv_ebit) and _epv_ebit is not None:
                    _tax_prov = _latest_inc.get('Tax Provision')
                    _pretax = _latest_inc.get('Pretax Income')
                    if (pd.notna(_tax_prov) and pd.notna(_pretax) and
                            _pretax and _pretax != 0):
                        _epv_eff_tax = max(0, min(float(_tax_prov) / float(_pretax), 0.50))
                else:
                    _epv_ebit = None

            _epv_ebit_used, _epv_ebit_source = _select_epv_ebit(
                float(_epv_ebit) if _epv_ebit is not None else None,
                _epv_yf_revenue,
                _edgar_metrics.get('op_margin_avg_10y'),
                _edgar_metrics.get('op_margin_hist_years', 0))

            bs = yf_data.get('balance_sheet')
            # Equity bridge: get_net_debt = total debt - cash, so passing it
            # as total_debt (with excess_cash=0) nets cash against debt in
            # one term. NOPAT/WACC is enterprise value; without this bridge
            # levered firms get an EPV floor overstated by debt-per-share.
            # None = leverage unknown (no balance sheet) → disqualify EPV
            # rather than fabricate an unlevered per-share floor, since EPV
            # is the widest-resolving leg of the consensus fallback.
            _epv_net_debt = get_net_debt(yf_data)
            if _epv_net_debt is None:
                epv_fv = None
                epv_valuation = None
            else:
                epv_valuation = earnings_power_value_valuation(
                    _epv_ebit_used,
                    _epv_eff_tax, wacc, shares,
                    excess_cash=0, total_debt=_epv_net_debt)
                epv_fv = epv_valuation.value
            epv_growth_fv = epv_with_growth_premium(
                epv_fv, ratios.get('ROE'), re_for_models)

            # RIM (Residual Income Model)
            _book_value = info.get('bookValue')
            if _book_value is None and shares and shares > 0:
                if bs is not None and not bs.empty:
                    _eq_val = bs.iloc[:, 0].get('Stockholders Equity')
                    if pd.notna(_eq_val) and _eq_val:
                        _book_value = float(_eq_val) / shares
            # Retention = 1 − payout, from the same info payload DDM uses.
            # Passing it makes clean-surplus book growth match reality (and
            # silences the g/ROE inference warning). None → the model infers
            # a Gordon-consistent retention as before.
            _rim_payout = info.get('payoutRatio')
            _rim_retention = (max(0.0, min(1.0, 1.0 - _rim_payout))
                              if isinstance(_rim_payout, (int, float)) else None)
            rim_valuation = residual_income_model_valuation(
                _book_value, ratios.get('ROE'), re_for_models,
                g=TERMINAL_GROWTH_RATE + effective_tg_adj,
                retention_ratio=_rim_retention)
            rim_fv = rim_valuation.value

            # NAV (Tangible Book Value per share) — universal asset-floor
            # sanity check that strips goodwill and intangibles out of equity.
            tangible_book_per_share = tangible_book_value_per_share(yf_data)
            nav_fv = tangible_book_per_share if (
                tangible_book_per_share and tangible_book_per_share > 0) else None
            nav_mos = ((nav_fv - current_price) / nav_fv
                if (nav_fv and current_price and nav_fv > 0) else None)
            p_tbv = (current_price / nav_fv
                if (nav_fv and current_price and nav_fv > 0) else None)

            # Reverse DCF (solve for implied growth)
            rev_dcf = None
            if dcf_fv and current_price and current_price > 0 and fcf and shares:
                # get_net_debt (or 0) matches the forward DCF's own EV→equity
                # bridge basis. Kept LOCAL to this branch: it must not clobber
                # the row's statement-derived net_debt_val, whose None means
                # "leverage unknown", not zero.
                _rev_net_debt = get_net_debt(yf_data) or 0
                # Solve on the same adjusted FCF base + terminal growth the
                # forward DCF used, so implied_vs_estimated measures the
                # expectation gap, not a basis/terminal-rate mismatch.
                _rev_fcf = growth_diag.get('base_fcf') or fcf
                _rev_tg = growth_diag.get('term_g')
                rev_dcf = reverse_dcf(
                    current_price, _rev_fcf, wacc, shares, _rev_net_debt,
                    terminal_g=_rev_tg if _rev_tg is not None else 0.03)

            # 52-Week Range
            high_52w = info.get('fiftyTwoWeekHigh')
            low_52w = info.get('fiftyTwoWeekLow')
            pct_from_52w_high = ((current_price - high_52w) / high_52w
                                 if (current_price and high_52w and high_52w > 0) else None)
            pct_from_52w_low = ((current_price - low_52w) / low_52w
                                if (current_price and low_52w and low_52w > 0) else None)
            range_52w_position = ((current_price - low_52w) / (high_52w - low_52w) * 100
                                  if (current_price and high_52w and low_52w
                                      and high_52w > low_52w) else None)

            # Founder-led detection (three layers, all gated on the founder
            # holding a CURRENT executive role — see data/founder_overrides.json
            # for the definition; board-only / retired founders do NOT count).
            # The officer pool merges two sources: yfinance companyOfficers (the
            # comp-table subset) PLUS SEC Form 4 reporting persons flagged as
            # officers. The SEC roster recovers founder-execs the comp table
            # omits — e.g. an Executive Chairman like Reed Hastings (NFLX).
            #   1) Title scan — a pooled officer whose title contains "founder"
            #      AND denotes an executive role (CEO/President/C-suite/COO/CTO/
            #      Exec Chair). Skips "Founder & Director", "Chairman Emeritus".
            #   2) Wikidata cross-reference — a Wikidata-listed founder of this
            #      CIK who is a current executive officer in the pool.
            #   3) Manual overrides — final say, beats both above.
            # founder_role records what triggered the flag (incl. source).
            officer_pool = [{'name': _o.get('name'), 'title': _o.get('title'),
                             'src': 'yfinance'} for _o in (officers or [])]
            for _so in _sec_reporting_officers(insider_data):
                officer_pool.append({'name': _so.get('name'),
                                     'title': _so.get('title'), 'src': 'SEC Form 4'})
            founder_led = False
            founder_role = None
            for _o in officer_pool:
                _title = _o.get('title') or ''
                if 'founder' in _title.lower() and _is_executive_title(_title):
                    founder_led = True
                    founder_role = '%s — %s [%s]' % (
                        _o.get('name') or '?', _title, _o.get('src'))
                    break
            if not founder_led:
                _m = _wikidata_founder_match(ticker, officer_pool)
                if _m:
                    founder_led = True
                    founder_role = '%s — %s [%s] (Wikidata founder)' % (
                        _m.get('name') or '?', _m.get('title') or '?',
                        _m.get('src') or '?')
            _foverrides = _load_founder_overrides()
            if ticker in _foverrides:
                founder_led = bool(_foverrides[ticker])
                founder_role = 'manual override' if founder_led else None

            # Ownership data from yfinance info
            shares_out = info.get('sharesOutstanding')
            float_shares = info.get('floatShares')
            insider_pct = info.get('heldPercentInsiders')
            inst_pct = info.get('heldPercentInstitutions')
            shares_short = info.get('sharesShort')
            short_ratio = info.get('shortRatio')
            short_pct_float = info.get('shortPercentOfFloat')

            # Share turnover rate = avg daily volume / shares outstanding
            avg_vol = info.get('averageVolume')
            share_turnover_rate = None
            if avg_vol and shares_out and shares_out > 0:
                share_turnover_rate = avg_vol / shares_out

            # Dividend yield and payout ratio (for narrative & template)
            _div_rate = info.get('dividendRate')
            _div_price = info.get('currentPrice') or info.get('regularMarketPrice')
            div_yield = (_div_rate / _div_price
                         if (_div_rate and _div_price and _div_price > 0) else None)
            payout_ratio = info.get('payoutRatio')

            # Balance sheet risk flags (goodwill, R&D, SGA)
            goodwill_pct = None
            rd_intensity = None
            sga_pct_rev = None
            sga_yoy_change = None
            inc_stmt = yf_data.get('income_statement')
            if bs is not None and not bs.empty:
                total_assets = bs.iloc[:, 0].get('Total Assets')
                gw = bs.iloc[:, 0].get('Goodwill')
                if (pd.notna(gw) and gw and pd.notna(total_assets)
                        and total_assets and total_assets > 0):
                    goodwill_pct = float(gw) / float(total_assets)
            if inc_stmt is not None and not inc_stmt.empty:
                _rev_latest = inc_stmt.iloc[:, 0].get('Total Revenue')
                _rd_latest = inc_stmt.iloc[:, 0].get('Research And Development')
                _sga_latest = inc_stmt.iloc[:, 0].get('Selling General And Administration')
                if (pd.notna(_rd_latest) and _rd_latest
                        and pd.notna(_rev_latest) and _rev_latest and _rev_latest > 0):
                    rd_intensity = float(_rd_latest) / float(_rev_latest)
                if (pd.notna(_sga_latest) and _sga_latest
                        and pd.notna(_rev_latest) and _rev_latest and _rev_latest > 0):
                    sga_pct_rev = float(_sga_latest) / float(_rev_latest)
                # SGA YoY change (compare most recent two years)
                if inc_stmt.shape[1] >= 2:
                    _sga_prior = inc_stmt.iloc[:, 1].get('Selling General And Administration')
                    if (pd.notna(_sga_latest) and _sga_latest
                            and pd.notna(_sga_prior) and _sga_prior and _sga_prior > 0):
                        sga_yoy_change = (float(_sga_latest) / float(_sga_prior)) - 1

            # Morningstar: fair value and difference vs model
            ms_diff = None
            ms_fv = None
            ms_pfv = ms_pfv_data.get(ticker)
            # ms_fv = price / (Morningstar P/FV) needs no model of ours — do
            # NOT gate it on dcf_fv, so the consensus cohort (DCF-less names)
            # can also be benchmarked against Morningstar. ms_diff, which
            # compares OUR fair value, stays gated on dcf_fv.
            if ms_pfv and current_price:
                ms_fv = current_price / ms_pfv
                if ms_fv <= 0:
                    ms_fv = None
                elif dcf_fv:
                    ms_diff = (dcf_fv / ms_fv) - 1

            row = {
                'ticker': ticker,
                'source_group': ticker_source.get(ticker, 'quality'),
                # Statement-source tag (sec_xbrl+yfinance | sec_xbrl | yfinance);
                # previously only printed to the console from screen_cache.
                'data_source': _stmt_source,
                # Company info (Step 3)
                'description': description,
                'company_name': company_name,
                'sector': sector,
                'industry': industry,
                'country': country,
                # FX normalization audit trail. ``fx_converted=True`` means
                # the financials and/or info dollar fields were multiplied
                # by the recorded rate(s) to land in USD. Tickers that
                # already reported in USD (most US-listed and some ADRs)
                # have fx_converted=False and the rates are None.
                'currency_quote': fx_meta.get('currency_quote'),
                'currency_financial': fx_meta.get('currency_financial'),
                'fx_rate_financial': fx_meta.get('fx_rate_financial'),
                'fx_rate_quote': fx_meta.get('fx_rate_quote'),
                'fx_converted': fx_meta.get('fx_converted', False),
                'fx_fetch_failed': fx_meta.get('fx_fetch_failed', False),
                'ceo': ceo,
                'ceo_bio': ceo_bio,
                'founder_led': founder_led,
                'founder_role': founder_role,
                # Culture raw inputs (narrative built in post-processing)
                'employees': _culture_raw.get('employees'),
                'ceo_total_pay': _culture_raw.get('ceo_total_pay'),
                'compensation_risk': _culture_raw.get('compensation_risk'),
                'sbc': _culture_raw.get('sbc'),
                'glassdoor_rating': _culture_gd.get('glassdoor_rating'),
                'glassdoor_ceo_pct': _culture_gd.get('glassdoor_ceo_pct'),
                'glassdoor_rec_pct': _culture_gd.get('glassdoor_rec_pct'),
                'fcf': fcf,
                # Ownership
                'shares_out': shares_out,
                'float_shares': float_shares,
                'insider_pct': insider_pct,
                'inst_pct': inst_pct,
                'shares_short': shares_short,
                'short_ratio': short_ratio,
                'short_pct_float': short_pct_float,
                'share_turnover_rate': share_turnover_rate,
                'share_buyback_rate': share_buyback_rate,
                # Insider activity (Form 4)
                'insider_buy_ratio': insider_data.get('insider_buy_ratio') if insider_data and insider_data.get('available') else None,
                'insider_buy_count_90d': insider_data.get('buy_count_90d') if insider_data and insider_data.get('available') else None,
                'insider_sell_count_90d': insider_data.get('sell_count_90d') if insider_data and insider_data.get('available') else None,
                'insider_buy_count_365d': insider_data.get('buy_count_365d') if insider_data and insider_data.get('available') else None,
                'insider_sell_count_365d': insider_data.get('sell_count_365d') if insider_data and insider_data.get('available') else None,
                'insider_net_shares': insider_data.get('net_shares_365d') if insider_data and insider_data.get('available') else None,
                'insider_net_value': insider_data.get('net_value_365d') if insider_data and insider_data.get('available') else None,
                'insider_transactions': (insider_data.get('transactions', [])[:10] if insider_data and insider_data.get('available') else []),
                'roic_by_year': roic_data.get('roic_by_year'),
                # Per-year NOPAT / invested capital (Moat: Incr ROIC gate —
                # scoring derives incremental ROIC = ΔNOPAT/ΔIC from these)
                '_nopat_by_year': roic_data.get('nopat_by_year'),
                '_ic_by_year': roic_data.get('invested_capital_by_year'),
                'roic_cv': roic_cv,
                'gross_margin': gross_margin,
                'shareholder_yield': shareholder_yield,
                'div_yield': div_yield,
                'payout_ratio': payout_ratio,
                # Core screen
                'roic': roic_data['avg_roic'],
                'wacc': wacc,
                'spread': roic_data['avg_roic'] - wacc,
                'mcap': multiples.get('market_cap'),
                # Time-series cheapness (Valuation: Mult vs Hist gate)
                'mult_vs_hist': mult_vs_hist,
                'mult_hist_years': mult_hist_years,
                # Latest financials (for profit pool analysis)
                'revenue': latest_fins.get('revenue'),
                'operating_income': latest_fins.get('operating_income'),
                'net_income': latest_fins.get('net_income'),
                'operating_margin': (latest_fins['operating_income'] / latest_fins['revenue']
                    if latest_fins.get('operating_income') is not None and latest_fins.get('revenue') and latest_fins['revenue'] > 0
                    else None),
                'er': cost_of_equity,
                're_method': cached['re_method'],
                # Beta diagnostics (Step 4A)
                'beta_raw': beta_diag.get('raw_beta') if beta_diag else None,
                # beta_adjusted is the beta that entered CAPM: precision-
                # weighted (local) or Blume (Yahoo fallback).
                'beta_adjusted': beta_diag.get('shrunk_beta') if beta_diag else None,
                'beta_shrink_weight': beta_diag.get('shrink_weight') if beta_diag else None,
                'beta_r2': beta_diag.get('r_squared') if beta_diag else None,
                'beta_se': beta_diag.get('se_beta') if beta_diag else None,
                'beta_n_obs': beta_diag.get('n_observations') if beta_diag else None,
                'beta_r2_class': beta_diag.get('r2_classification') if beta_diag else None,
                'beta_source': beta_diag.get('beta_source') if beta_diag else None,
                'beta_warnings': (beta_diag.get('warnings') or None) if beta_diag else None,
                # Valuation (Step 5)
                'dcf_fv': dcf_fv,
                'price': current_price,
                'mos': mos,
                'dcf_sens_range': dcf_sens_range,
                'fcf_growth': fcf_growth,
                'analyst_ltg': growth_diag.get('analyst_ltg'),
                'margin_trend': growth_diag.get('margin_trend'),
                'surprise_avg': growth_diag.get('surprise_avg'),
                'fundamental_growth': growth_diag.get('fundamental_growth'),
                'reinvestment_rate': growth_diag.get('reinvestment_rate'),
                'terminal_growth': _get_sector_config(sector)['terminal_growth'],
                'exit_mult_fv': growth_diag.get('exit_mult_fv'),
                'tv_method_spread': growth_diag.get('tv_method_spread'),
                # Valuation envelope (DCF primary leg)
                'dcf_method': growth_diag.get('dcf_method'),
                'dcf_confidence': growth_diag.get('dcf_confidence'),
                'dcf_warnings': growth_diag.get('dcf_warnings'),
                'mc_p10_fv': mc_result['p10_fv'] if mc_result else None,
                'mc_p90_fv': mc_result['p90_fv'] if mc_result else None,
                'mc_cv': mc_result['cv'] if mc_result else None,
                'mc_confidence': _mc_confidence_label(mc_result['cv']) if mc_result and mc_result.get('cv') is not None else None,
                'ms_diff': ms_diff,
                'ms_fv': ms_fv,
                'ms_pfv': ms_pfv,
                # Multiples (Step 2)
                'pe': multiples.get('pe'),
                'ev_ebitda': multiples.get('ev_ebitda'),
                'enterprise_value': multiples.get('enterprise_value'),
                'pfcf': multiples.get('pfcf'),
                'pb': multiples.get('pb'),
                # Analyst consensus (Step 8)
                'analyst_rec': analyst.get('rec_key', '').upper() if analyst.get('rec_key') else None,
                'num_analysts': analyst.get('num_analysts'),
                'target_mean': analyst.get('target_mean'),
                'target_high': analyst.get('target_high'),
                'target_low': analyst.get('target_low'),
                'earnings_next_date': next_earn.get('earnings_next_date'),
                'earnings_date_est': next_earn.get('earnings_date_est'),
                # Quality (Step 3B)
                'piotroski': piotroski,
                'cash_conv': eq.get('cash_conversion'),
                'accruals': eq.get('accruals_ratio'),
                'rev_cagr': rev_cagr,
                'rev_cagr_5y': rev_cagr_5y,
                'rev_cagr_10y': rev_cagr_10y,
                'fcf_cagr_5y': fcf_cagr_5y,
                'fcf_cagr_10y': fcf_cagr_10y,
                'gross_margin_avg_5y': gross_margin_avg_5y,
                'gross_margin_trend': gross_margin_trend,
                'dividend_cagr_5y': dividend_cagr_5y,
                'shares_cagr_5y': shares_cagr_5y,
                # EDGAR-derived point-in-time FCF (OCF − capex); scoring uses
                # it as a fallback for `fcf` when yfinance has no cash flow.
                'fcf_edgar': _edgar_metrics.get('fcf_edgar'),
                # EDGAR-derived EBIT/interest; scoring uses it as a fallback
                # for `int_cov` when yfinance has no income statement.
                'int_cov_edgar': _edgar_metrics.get('int_cov_edgar'),
                # Revenue-growth volatility (Quality: Rev Volatility gate) and
                # the run's risk-free rate (Valuation: FCF Yield gate hurdle).
                'rev_growth_vol': _edgar_metrics.get('rev_growth_vol'),
                # Through-cycle operating margin (Quality: Margin vs Hist gate
                # + normalized-EBIT EPV input)
                'op_margin_avg_10y': _edgar_metrics.get('op_margin_avg_10y'),
                'op_margin_hist_years': _edgar_metrics.get('op_margin_hist_years', 0),
                '_risk_free_rate': risk_free_rate,
                # EDGAR XBRL validation
                'edgar_quality_score': xbrl_validation.get('edgar_quality_score') if xbrl_validation else None,
                'edgar_fields_flagged': xbrl_validation.get('fields_flagged', 0) if xbrl_validation else 0,
                'edgar_discrepancies': xbrl_validation.get('discrepancies', []) if xbrl_validation else [],
                'edgar_history': edgar_history,
                # Balance sheet (Step 3C)
                'int_cov': int_cov,
                'nd_ebitda': nd_ebitda,
                'total_debt': total_debt_val,
                'net_debt': net_debt_val,
                'cash': cash_val,
                'total_liabilities': total_liabilities_val,
                'debt_source': debt_source,
                # Traditional ratios
                'roe': ratios.get('ROE'),
                'de': ratios.get('Debt-to-Equity'),
                'cr': ratios.get('Current Ratio'),
                'roa': ratios.get('ROA'),
                # Macro overlay
                'macro_regime': macro_regime_result['regime'] if macro_regime_result else None,
                'macro_composite': macro_regime_result['composite_score'] if macro_regime_result else None,
                'macro_erp': effective_erp,
                'sector_headwinds': sector_signals.get(sector, {}).get('headwinds', []),
                'sector_tailwinds': sector_signals.get(sector, {}).get('tailwinds', []),
                'news_headlines': ticker_news,
                'news_sentiment': news_sentiment,
                'legal_filings': legal_data.get('filings', []),
                'legal_count': legal_data.get('count', 0),
                'legal_latest': legal_data.get('latest_date'),
                'suppliers': supply_data.get('suppliers', []),
                'customers': supply_data.get('customers', []),
                'supply_chain_available': supply_data.get('available', False),
                'finnhub_peers': finnhub_peers,
                # DDM (Dividend Discount Model)
                'ddm_eligible': ddm_result.get('ddm_eligible', False),
                'ddm_reason': ddm_result.get('ddm_reason'),
                'ddm_fv': ddm_result.get('ddm_fv'),
                'ddm_h_fv': ddm_result.get('ddm_h_fv'),
                'ddm_growth': ddm_result.get('ddm_growth'),
                'ddm_div_cagr': ddm_result.get('ddm_div_cagr'),
                'ddm_sustainable_growth': ddm_result.get('ddm_sustainable_growth'),
                'ddm_payout_flag': ddm_result.get('ddm_payout_flag', False),
                'ddm_consecutive_years': ddm_result.get('ddm_consecutive_years'),
                'ddm_mc_median': ddm_result.get('ddm_mc_median'),
                'ddm_mc_p10': ddm_result.get('ddm_mc_p10'),
                'ddm_mc_p90': ddm_result.get('ddm_mc_p90'),
                'ddm_mc_cv': ddm_result.get('ddm_mc_cv'),
                'ddm_confidence': ddm_result.get('ddm_confidence'),
                'ddm_warnings': ddm_result.get('ddm_warnings'),
                # Reverse DCF
                'implied_growth': rev_dcf['implied_growth'] if rev_dcf and rev_dcf.get('converged') else None,
                'implied_vs_estimated': ((rev_dcf['implied_growth'] - fcf_growth)
                    if rev_dcf and rev_dcf.get('converged')
                    and fcf_growth is not None else None),
                # EPV (Earnings Power Value)
                'epv_fv': epv_fv,
                'epv_pfv': (current_price / epv_fv
                    if (epv_fv and current_price and epv_fv > 0) else None),
                'epv_mos': ((epv_fv - current_price) / epv_fv
                    if (epv_fv and current_price and epv_fv > 0) else None),
                'epv_growth_fv': epv_growth_fv,
                # 'normalized' (10y-avg margin × current revenue) or 'point'
                'epv_ebit_source': _epv_ebit_source,
                'epv_confidence': (epv_valuation.confidence
                    if epv_valuation is not None and epv_fv is not None else None),
                'epv_warnings': (list(epv_valuation.warnings)
                    if epv_valuation is not None else []),
                # RIM (Residual Income Model)
                'rim_fv': rim_fv,
                'rim_mos': ((rim_fv - current_price) / rim_fv
                    if (rim_fv and current_price and rim_fv > 0) else None),
                'rim_confidence': rim_valuation.confidence if rim_fv is not None else None,
                'rim_warnings': list(rim_valuation.warnings),
                # NAV (Tangible Book Value)
                'tangible_book_per_share': tangible_book_per_share,
                'nav_fv': nav_fv,
                'nav_mos': nav_mos,
                'p_tbv': p_tbv,
                # Raw sign-preserving TBV/share: lets scoring distinguish
                # "negative tangible book" (P/TBV structurally inapplicable)
                # from "balance sheet missing" (N/A scores worst)
                'tangible_book_ps': tangible_book_per_share,
                # Altman Z-Score
                'altman_z': altman_z,
                'altman_z_zone': altman_z_zone,
                # Beneish M-Score
                'beneish_m': beneish['m_score'] if beneish else None,
                'beneish_flag': beneish['manipulation_flag'] if beneish else None,
                # DuPont Decomposition
                'dupont_margin': dupont['margin'] if dupont else None,
                'dupont_turnover': dupont['turnover'] if dupont else None,
                'dupont_leverage': dupont['leverage'] if dupont else None,
                # 52-Week Range
                'high_52w': high_52w,
                'low_52w': low_52w,
                'pct_from_52w_high': pct_from_52w_high,
                'pct_from_52w_low': pct_from_52w_low,
                'range_52w_position': range_52w_position,
                # Balance sheet risk flags
                'goodwill_pct': goodwill_pct,
                'rd_intensity': rd_intensity,
                'sga_pct_rev': sga_pct_rev,
                'sga_yoy_change': sga_yoy_change,
                # --- Price-history signals ---
                'realized_vol':    round(_ticker_realized_vol, 4) if _ticker_realized_vol is not None else None,
                'drawdown_2008':   round(_ticker_dd_2008, 4) if _ticker_dd_2008 is not None else None,
                'drawdown_2020':   round(_ticker_dd_2020, 4) if _ticker_dd_2020 is not None else None,
                'drawdown_2022':   round(_ticker_dd_2022, 4) if _ticker_dd_2022 is not None else None,
                'rolling_betas':   _rolling_beta_diag or None,
                # --- Market & risk metrics (momentum / liquidity) ---
                **_round_price_metrics(_price_metrics),
            }
            # Rating set later by score_and_rate from composite score plus critical caps
            row['rating'] = None
            row['_provenance'] = _prov.ticker_block(ticker)
            results.append(row)
        except Exception as e:
            print(f"  Error analyzing {ticker}: {e}")
        finally:
            # Drop the heavy yf_data reference now that this ticker is done
            if ticker in screen_cache:
                screen_cache[ticker].pop('yf_data', None)
            sys.stdout.flush()

    # All per-ticker analysis complete — release remaining caches
    screen_cache.clear()
    yf_client.evict_financials()
    gc.collect()
    return results


def _run_postprocess(results, ms_pfv_data, _carry_prior_rows):
    """Sector comparisons, valuation blends, profit pools, scoring, sizing."""
    # -----------------------------------------------------------------------
    # Post-processing: sector-median EV/EBITDA comparison + DCF cross-check
    # -----------------------------------------------------------------------

    # 1. Compute sector median EV/EBITDA
    _sector_ee = {}
    for r in results:
        s, ee = r.get('sector'), r.get('ev_ebitda')
        if s and ee and 0 < ee < EV_EBITDA_OUTLIER_MAX:  # filter outliers
            _sector_ee.setdefault(s, []).append(ee)
    sector_median_ee = {s: _median(v) for s, v in _sector_ee.items() if len(v) >= MIN_SECTOR_STOCKS}

    for r in results:
        s = r.get('sector')
        ee = r.get('ev_ebitda')
        med = sector_median_ee.get(s)
        r['_sector_median_ee'] = med
        r['_ee_vs_sector'] = (ee / med - 1) if ee and med and med > 0 else None

        # Stash the pristine DCF before ANY blend overwrites r['dcf_fv'].
        # fv_dispersion keys off this so the DDM leg (blended into dcf_fv
        # below) isn't double-counted, which would mute the disagreement gate
        # for exactly the dividend payers where DCF and DDM diverge.
        r['_dcf_fv_preblend'] = r.get('dcf_fv')

        # 2. DCF cross-check: compute multiples-implied fair value from sector median
        # If DCF FV > 1.5× multiples FV, blend toward multiples (40% weight)
        # This reins in over-estimates where DCF extrapolates peak FCF
        ev_raw = r.get('enterprise_value')
        ev_eb = r.get('ev_ebitda')
        dcf_fv = r.get('dcf_fv')
        price = r.get('price')
        shares = r.get('shares_out')

        if (ev_raw and ev_eb and ev_eb > 0 and med and shares and shares > 0
                and dcf_fv and price):
            ebitda = ev_raw / ev_eb
            multiples_ev = med * ebitda
            # Net debt = EV − equity market cap; price is guaranteed here, so
            # no falsy-price fallback that would set net_debt = full EV.
            net_debt = ev_raw - (price * shares)
            multiples_fv = (multiples_ev - net_debt) / shares
            if multiples_fv > 0:
                r['_multiples_fv'] = multiples_fv
                # Blend when DCF is significantly above multiples-implied value
                # BUT only when DCF > price (model sees upside → possible over-estimate).
                # If DCF < price the model already under-values; pulling down worsens it.
                if dcf_fv > multiples_fv * BLEND_TRIGGER and price and dcf_fv > price:
                    blended = BLEND_DCF_WEIGHT * dcf_fv + BLEND_MULT_WEIGHT * multiples_fv
                    blend_ratio = blended / dcf_fv if dcf_fv > 0 else 1.0
                    r['dcf_fv'] = blended
                    # Recompute dependent fields
                    if blended > 0:
                        r['mos'] = (blended - price) / blended
                    # Scale the whole uncertainty band (sens range + MC
                    # percentiles) so bear/bull and P10/P90 stay consistent.
                    _rescale_fv_band(r, blend_ratio)
                    r['_blended'] = True
                else:
                    r['_blended'] = False
            else:
                r['_multiples_fv'] = None
                r['_blended'] = False
        else:
            r['_multiples_fv'] = None
            r['_blended'] = False

        # Recompute Morningstar fields after potential blending
        ms_pfv_val = ms_pfv_data.get(r['ticker'])
        if ms_pfv_val and price:
            ms_fv = price / ms_pfv_val
            if ms_fv > 0:
                r['ms_fv'] = ms_fv
                r['ms_pfv'] = ms_pfv_val
                if r.get('dcf_fv'):
                    r['ms_diff'] = (r['dcf_fv'] / ms_fv) - 1

    # DDM blending: for eligible stocks, blend 70% DCF + 30% DDM — but only
    # when the DDM is trustworthy. An unfunded dividend (payout > 100%) or a
    # DDM that disagrees with the DCF by more than the divergence threshold is
    # NOT blended in; the flags stop being cosmetic. When blended, the whole
    # uncertainty band is rescaled like the multiples blend does.
    for r in results:
        r['_blended_method'] = 'DCF'
        r['_ddm_low_confidence'] = False
        if r.get('ddm_eligible') and r.get('ddm_fv') and r.get('dcf_fv'):
            ddm_fv = r['ddm_fv']
            dcf_fv = r['dcf_fv']
            if ddm_fv > 0 and dcf_fv > 0:
                avg_fv = (dcf_fv + ddm_fv) / 2.0
                divergence = abs(dcf_fv - ddm_fv) / avg_fv if avg_fv > 0 else 0
                low_confidence = (divergence > DDM_DIVERGENCE_THRESHOLD
                                  or bool(r.get('ddm_payout_flag')))
                r['_ddm_low_confidence'] = low_confidence
                if not low_confidence:
                    blended = DCF_BLEND_WEIGHT_WITH_DDM * dcf_fv + DDM_BLEND_WEIGHT * ddm_fv
                    blend_ratio = blended / dcf_fv if dcf_fv > 0 else 1.0
                    r['dcf_fv'] = blended
                    r['_blended_method'] = 'DCF+DDM'
                    price = r.get('price')
                    if price and blended > 0:
                        r['mos'] = (blended - price) / blended
                    _rescale_fv_band(r, blend_ratio)

    # Rating is set downstream by score_and_rate from composite score plus critical caps

    # -----------------------------------------------------------------------
    # Profit pool analysis (sector-level revenue/profit concentration)
    # Must run BEFORE screening matrix so pp_multiple is available for gates
    # -----------------------------------------------------------------------
    # 1. Aggregate sector totals
    _sector_rev = {}     # sector → total revenue
    _sector_opinc = {}   # sector → total operating income (clamped ≥0)
    _sector_tickers = {} # sector → [(ticker, revenue, operating_income)]
    for r in results:
        s = r.get('sector')
        rev = r.get('revenue')
        opinc = r.get('operating_income')
        if s and rev and rev > 0:
            _sector_rev[s] = _sector_rev.get(s, 0) + rev
            if opinc is not None:
                _sector_opinc[s] = _sector_opinc.get(s, 0) + max(opinc, 0)
            _sector_tickers.setdefault(s, []).append((r['ticker'], rev, opinc or 0))

    # 2. Sector-level operating margin median
    _sector_opm = {}
    for r in results:
        s = r.get('sector')
        opm = r.get('operating_margin')
        if s and opm is not None:
            _sector_opm.setdefault(s, []).append(opm)
    sector_median_opm = {s: _median(v) for s, v in _sector_opm.items()
                         if len(v) >= MIN_SECTOR_STOCKS}

    # 3. Per-ticker profit pool metrics
    for r in results:
        s = r.get('sector')
        rev = r.get('revenue')
        opinc = r.get('operating_income')

        # Revenue share (fraction of sector total revenue in analysis universe)
        sec_rev = _sector_rev.get(s, 0)
        r['pp_revenue_share'] = (rev / sec_rev) if (rev and sec_rev > 0) else None

        # Profit share (fraction of sector total operating income)
        sec_opinc = _sector_opinc.get(s, 0)
        r['pp_profit_share'] = (max(opinc, 0) / sec_opinc
                                if (opinc is not None and sec_opinc > 0) else None)

        # Profit pool multiple = profit_share / revenue_share
        # > 1 means disproportionate profit capture; < 1 means under-earning
        # ps == 0.0 (zero/negative operating income) is a real value — 0.00x,
        # maximal under-earning — not missing data, so test `is not None`
        # rather than truthiness (which rendered 125+ loss-makers as N/A).
        rs = r.get('pp_revenue_share')
        ps = r.get('pp_profit_share')
        r['pp_multiple'] = (ps / rs) if (ps is not None and rs and rs > 0) else None

        # Margin advantage vs sector median operating margin
        opm = r.get('operating_margin')
        med_opm = sector_median_opm.get(s)
        r['pp_margin_advantage'] = ((opm - med_opm)
                                    if (opm is not None and med_opm is not None) else None)
        r['_sector_median_opm'] = med_opm

        # Sector-level concentration metrics (same for all tickers in sector)
        tickers_in_sector = _sector_tickers.get(s, [])
        if len(tickers_in_sector) >= 3 and sec_rev > 0:
            shares = [(t_rev / sec_rev) for _, t_rev, _ in tickers_in_sector]
            r['pp_sector_hhi'] = round(sum(sh ** 2 for sh in shares), 4)
            top4 = sorted(shares, reverse=True)[:4]
            r['pp_sector_cr4'] = round(sum(top4), 4)
            r['pp_sector_count'] = len(tickers_in_sector)
        else:
            r['pp_sector_hhi'] = None
            r['pp_sector_cr4'] = None
            r['pp_sector_count'] = len(tickers_in_sector) if tickers_in_sector else 0

    apply_mcap_integrity_guard(results, _carry_prior_rows)

    score_and_rate(results)

    # Position sizing and concentration analysis
    weights = position_sizes(results)
    for r in results:
        r['position_weight'] = weights.get(r['ticker'])
    concentration = concentration_analysis(
        [r for r in results if r.get('rating') in ('BUY', 'LEAN BUY')])
    if concentration.get('concentration_flag'):
        print(f"\n  Portfolio concentration warning: {concentration['top_sector']} "
              f"= {concentration['top_sector_weight']:.0%} "
              f"(HHI={concentration['hhi']:.2f})")
    return {'sector_median_ee': sector_median_ee,
            'sector_median_opm': sector_median_opm}


def _run_narratives(results, args, sector_etf_data, macro_regime_result,
                    commodity_data, sector_median_ee, sector_median_opm):
    """Peer percentiles, culture and stock narratives, and the final sort."""
    # -----------------------------------------------------------------------
    # Peer percentile ranking (sector-relative position for key metrics)
    # -----------------------------------------------------------------------
    _peer_metrics = ['roic', 'gross_margin', 'rev_cagr', 'nd_ebitda', 'piotroski',
                     'rd_intensity', 'goodwill_pct', 'operating_margin', 'pp_multiple']
    _peer_buckets = {}  # metric → sector → sorted list of values
    for metric in _peer_metrics:
        _peer_buckets[metric] = {}
        for r in results:
            s = r.get('sector')
            v = r.get(metric)
            if s and isinstance(v, (int, float)) and not isinstance(v, bool):
                _peer_buckets[metric].setdefault(s, []).append(v)
        for s in _peer_buckets[metric]:
            _peer_buckets[metric][s].sort()

    for r in results:
        s = r.get('sector')
        for metric in _peer_metrics:
            vals = _peer_buckets[metric].get(s, [])
            v = r.get(metric)
            if v is not None and len(vals) >= 3:
                # Percentile: fraction of peers this value exceeds
                below = sum(1 for x in vals if x < v)
                pctile = below / len(vals)
                r[f'_peer_pctile_{metric}'] = round(pctile, 2)
            else:
                r[f'_peer_pctile_{metric}'] = None

    # -----------------------------------------------------------------------
    # Culture narrative: workforce productivity, pay, ownership culture
    # -----------------------------------------------------------------------
    # Step 1 — derive per-employee metrics
    for r in results:
        emp     = r.get('employees')
        rev     = r.get('revenue')
        fcf_val = r.get('fcf')
        ceo_pay = r.get('ceo_total_pay')
        sbc     = r.get('sbc')

        rpe = (rev / emp) if (emp and rev and rev > 0) else None
        r['revenue_per_emp'] = rpe
        r['fcf_per_emp']  = (fcf_val / emp) if (emp and fcf_val is not None) else None
        r['ceo_pay_ratio'] = (ceo_pay / rpe) if (ceo_pay and rpe and rpe > 0) else None
        r['sbc_per_emp']   = (sbc / emp) if (sbc and emp) else None

    # Step 2 — sector-percentile buckets for revenue per employee
    _cult_sector_rpe: dict = {}
    for r in results:
        s, rpe = r.get('sector'), r.get('revenue_per_emp')
        if s and rpe and rpe > 0:
            _cult_sector_rpe.setdefault(s, []).append(rpe)
    for s in _cult_sector_rpe:
        _cult_sector_rpe[s].sort()

    # Step 3 — multi-year RPE trend from EDGAR revenue history
    for r in results:
        emp = r.get('employees')
        if not emp:
            r['rpe_cagr'] = None
            continue
        rev_hist = _flow_to_annual((r.get('edgar_history') or {}).get('revenue_history') or {})
        years = sorted(rev_hist.keys())
        if len(years) >= 3:
            earliest = rev_hist[years[0]]
            latest   = rev_hist[years[-1]]
            n_years  = years[-1] - years[0]
            if earliest and latest and earliest > 0 and n_years > 0:
                rpe_earliest = earliest / emp
                rpe_latest   = latest   / emp
                r['rpe_cagr'] = (rpe_latest / rpe_earliest) ** (1 / n_years) - 1
            else:
                r['rpe_cagr'] = None
        else:
            r['rpe_cagr'] = None

    # Step 4 — employment-related legal flag
    _EMPLOYMENT_KEYWORDS = {
        'labor', 'labour', 'employee', 'employment', 'wage', 'salary',
        'discrimination', 'wrongful termination', 'class action', 'nlrb',
        'union', 'strike', 'layoff', 'hostile work',
    }
    for r in results:
        filings = r.get('legal_filings') or []
        flag = False
        for f in filings:
            text = ' '.join([
                (f.get('description') or ''),
                (f.get('summary') or ''),
            ]).lower()
            if any(kw in text for kw in _EMPLOYMENT_KEYWORDS):
                flag = True
                break
        r['employment_legal_flag'] = flag

    # Step 5 — layoff / culture news signal
    _LAYOFF_KEYWORDS = {
        'layoff', 'lay off', 'laid off', 'job cut', 'workforce reduction',
        'redundan', 'downsiz', 'restructur', 'reorg',
    }
    _CULTURE_POS_KEYWORDS = {
        'best place', 'top employer', 'great place to work',
        'best company', 'culture award',
    }
    for r in results:
        headlines = r.get('news_headlines') or []
        layoff_signal = False
        culture_award = False
        for h in headlines:
            text = (h.get('title') or '').lower()
            if any(kw in text for kw in _LAYOFF_KEYWORDS):
                layoff_signal = True
            if any(kw in text for kw in _CULTURE_POS_KEYWORDS):
                culture_award = True
        r['layoff_news_signal'] = layoff_signal
        r['culture_award_signal'] = culture_award

    # Step 6 — plain-English narrative
    def _fmt_emp(n):
        if n >= 1_000_000: return f"{n / 1_000_000:.1f}M"
        if n >= 1_000:     return f"{n // 1_000:,}K"
        return str(n)

    def _fmt_money(v):
        if abs(v) >= 1_000_000: return f"${v / 1_000_000:.1f}M"
        if abs(v) >= 1_000:     return f"${v / 1_000:.0f}K"
        return f"${v:.0f}"

    for r in results:
        s         = r.get('sector') or 'its sector'
        emp       = r.get('employees')
        rpe       = r.get('revenue_per_emp')
        rpe_cagr  = r.get('rpe_cagr')
        fcf_pe    = r.get('fcf_per_emp')
        ceo_ratio = r.get('ceo_pay_ratio')
        sbc_pe    = r.get('sbc_per_emp')
        rd        = r.get('rd_intensity')
        crisk     = r.get('compensation_risk')
        gd_rating = r.get('glassdoor_rating')
        gd_ceo    = r.get('glassdoor_ceo_pct')
        gd_rec    = r.get('glassdoor_rec_pct')
        emp_legal = r.get('employment_legal_flag', False)
        layoff    = r.get('layoff_news_signal', False)
        cult_award = r.get('culture_award_signal', False)

        sector_rpes = _cult_sector_rpe.get(r.get('sector'), [])
        rpe_pct = None
        if rpe and len(sector_rpes) >= 3:
            rpe_pct = sum(1 for x in sector_rpes if x < rpe) / len(sector_rpes)

        sentences = []

        # --- Glassdoor (highest credibility — leads if available) ---------
        if gd_rating is not None:
            stars = f"{gd_rating:.1f}/5"
            rec_str = f", with {gd_rec}% of employees recommending it to a friend" if gd_rec else ""
            ceo_str = f" and {gd_ceo}% CEO approval" if gd_ceo else ""
            sentences.append(
                f"Glassdoor-rated {stars}{rec_str}{ceo_str}."
            )

        # --- Workforce size -----------------------------------------------
        if emp:
            sentences.append(f"Employs approximately {_fmt_emp(emp)} people.")

        # --- Revenue per employee with trend -----------------------------
        if rpe and rpe_pct is not None:
            if rpe_pct >= 0.75:   pct_desc = "top quartile"
            elif rpe_pct >= 0.50: pct_desc = "above the sector median"
            elif rpe_pct >= 0.25: pct_desc = "below the sector median"
            else:                 pct_desc = "bottom quartile"
            trend_str = ""
            if rpe_cagr is not None:
                if rpe_cagr > 0.05:
                    trend_str = f", improving at {rpe_cagr:.0%}/yr — growing workforce leverage"
                elif rpe_cagr < -0.05:
                    trend_str = f", declining at {abs(rpe_cagr):.0%}/yr — weakening productivity"
            sentences.append(
                f"Revenue per employee of {_fmt_money(rpe)} ranks in the "
                f"{pct_desc} among {s} peers{trend_str}."
            )
        elif rpe:
            sentences.append(f"Revenue per employee is {_fmt_money(rpe)}.")

        # --- FCF per employee --------------------------------------------
        if fcf_pe and fcf_pe > 0 and emp:
            sentences.append(
                f"Each employee generates {_fmt_money(fcf_pe)} of free cash flow annually."
            )

        # --- SBC per employee (ownership culture) ------------------------
        if sbc_pe and sbc_pe > 0:
            sentences.append(
                f"Stock-based compensation of {_fmt_money(sbc_pe)} per employee "
                f"reflects an ownership culture."
            )

        # --- CEO pay alignment -------------------------------------------
        if ceo_ratio is not None:
            if ceo_ratio <= 10:   ratio_desc = "modest"
            elif ceo_ratio <= 30: ratio_desc = "reasonable"
            elif ceo_ratio <= 75: ratio_desc = "elevated"
            elif ceo_ratio <= 150: ratio_desc = "high"
            else:                  ratio_desc = "very high"
            sentences.append(
                f"CEO compensation is {ceo_ratio:.0f}\u00d7 revenue per employee "
                f"({ratio_desc} relative to workforce productivity)."
            )

        # --- yfinance compensation risk ----------------------------------
        if crisk is not None:
            if crisk <= 3:
                crisk_desc = "low compensation governance risk"
            elif crisk <= 6:
                crisk_desc = "moderate compensation governance risk"
            else:
                crisk_desc = "elevated compensation governance risk"
            sentences.append(
                f"Governance score flags {crisk_desc} ({crisk}/10)."
            )

        # --- R&D intensity -----------------------------------------------
        if rd and rd > 0.01:
            if rd >= 0.20:   rd_desc = "heavy"
            elif rd >= 0.10: rd_desc = "significant"
            elif rd >= 0.05: rd_desc = "moderate"
            else:            rd_desc = "limited"
            sentences.append(
                f"R&D investment of {rd:.0%} of revenue ({rd_desc}) signals "
                f"commitment to product development and talent."
            )

        # --- Contradiction detection -------------------------------------
        if rpe_pct is not None and rd and rd >= 0.10 and rpe_pct < 0.25:
            sentences.append(
                "Note: heavy R&D investment has not yet translated to "
                "above-average workforce productivity — watch for commercialisation lag."
            )

        # --- External signals --------------------------------------------
        if cult_award:
            sentences.append(
                "Recent news includes recognition as a top employer or culture award."
            )
        if layoff:
            sentences.append(
                "Recent headlines include layoff or workforce-reduction announcements."
            )
        if emp_legal:
            sentences.append(
                "Active legal proceedings include employment or labour-related filings."
            )

        r['culture_narrative'] = " ".join(sentences) if sentences else None

    # -----------------------------------------------------------------------
    # Stock-level narrative (replaces sector-only headwinds/tailwinds)
    # -----------------------------------------------------------------------
    for r in results:
        hw, tw = generate_stock_narrative(
            r,
            sector_data=sector_etf_data if args.macro else None,
            macro_regime_result=macro_regime_result,
            commodity_data=commodity_data,
            sector_medians={'sector_median_ee': sector_median_ee,
                            'sector_median_opm': sector_median_opm},
        )
        r['sector_headwinds'] = hw
        r['sector_tailwinds'] = tw
        r['financial_summary'] = generate_financial_summary(r)

    results.sort(key=lambda r: (r.get('_composite_score') or 0, r.get('spread') or 0), reverse=True)


def _run_validation_stats(results, ms_pfv_data, args, screen_outcomes):
    """Morningstar comparison statistics and validation-cohort stats."""
    # -----------------------------------------------------------------------
    # Morningstar comparison statistics
    # -----------------------------------------------------------------------
    if ms_pfv_data:
        ms_pairs = []
        for r in results:
            pfv_val = ms_pfv_data.get(r['ticker'])
            if pfv_val and r.get('price') and r.get('dcf_fv') and r['dcf_fv'] > 0:
                ms_fv = r['price'] / pfv_val
                if ms_fv > 0:
                    ms_pairs.append((r['dcf_fv'], ms_fv))
        if len(ms_pairs) >= MIN_MORNINGSTAR_SAMPLE:
            model_fvs = [p[0] for p in ms_pairs]
            ms_fvs = [p[1] for p in ms_pairs]
            rel_errors = [(m - ms) / ms for m, ms in ms_pairs]
            mae = sum(abs(e) for e in rel_errors) / len(rel_errors)
            mse = sum(rel_errors) / len(rel_errors)
            within_20 = sum(1 for e in rel_errors if abs(e) <= 0.20) / len(rel_errors)
            ratios_sorted = sorted(m / ms for m, ms in ms_pairs)
            median_ratio = ratios_sorted[len(ratios_sorted) // 2]
            # Spearman rank correlation
            n = len(model_fvs)
            rank_m = rank(model_fvs)
            rank_ms = rank(ms_fvs)
            d_sq = sum((rm - rms) ** 2 for rm, rms in zip(rank_m, rank_ms, strict=False))
            spearman_rho = 1 - (6 * d_sq) / (n * (n ** 2 - 1)) if n > 1 else 0.0

            print(f"\nMorningstar comparison ({len(ms_pairs)} stocks):")
            print(f"  Mean Absolute Error: {mae:.1%}")
            print(f"  Mean Signed Error:   {mse:+.1%} "
                  f"({'overestimates' if mse > 0 else 'underestimates'})")
            print(f"  Within ±20%:         {within_20:.0%}")
            print(f"  Median FV Ratio:     {median_ratio:.2f}")
            print(f"  Spearman ρ:          {spearman_rho:.3f}")

            # Per-group MS comparison (when validation data present)
            if args.validation:
                for grp_name in ('quality', 'poor'):
                    grp_pairs = []
                    for r in results:
                        if r.get('source_group') != grp_name:
                            continue
                        pfv_val = ms_pfv_data.get(r['ticker'])
                        if pfv_val and r.get('price') and r.get('dcf_fv') and r['dcf_fv'] > 0:
                            gms_fv = r['price'] / pfv_val
                            if gms_fv > 0:
                                grp_pairs.append((r['dcf_fv'], gms_fv))
                    if len(grp_pairs) >= 3:
                        grp_rel_errors = [(m - ms) / ms for m, ms in grp_pairs]
                        grp_mae = sum(abs(e) for e in grp_rel_errors) / len(grp_rel_errors)
                        grp_mse = sum(grp_rel_errors) / len(grp_rel_errors)
                        grp_w20 = sum(1 for e in grp_rel_errors if abs(e) <= 0.20) / len(grp_rel_errors)
                        print(f"  {grp_name:>8} ({len(grp_pairs)} stocks): "
                              f"MAE={grp_mae:.1%}  MSE={grp_mse:+.1%}  within20={grp_w20:.0%}")

    if args.validation:
        _print_validation_stats(results, screen_outcomes)


def _write_outputs(results, run_start_date, _prov, risk_free_rate,
                   risk_free_rate_source, macro_regime_result, macro_adj,
                   local_rs, prices_dir, sector_etf_data=None):
    """Write the JSON snapshot, provenance events, HTML and Excel reports."""
    os.makedirs("output", exist_ok=True)
    today_str = run_start_date.isoformat()  # pin to run-start so a midnight-spanning run stays single-dated
    _run_prov = _prov.run_block(results)

    # Save results as JSON for backtesting pipeline. Written BEFORE the
    # HTML/Excel renders so the Phase-2 snapshot survives a render crash.
    json_filename = os.path.join("output", f"results_{run_start_date.isoformat()}.json")
    def _make_json_safe(val, _depth=0):
        """Recursively convert a value to a JSON-safe structure (max depth 8)."""
        if _depth > 8:
            return None
        # None and bool must come before int (bool is a subclass of int in Python)
        if val is None or isinstance(val, bool):
            return val
        if isinstance(val, int):
            return val
        if isinstance(val, float):
            # inf/nan are not valid JSON; replace with None (→ JSON null)
            import math
            return None if (math.isnan(val) or math.isinf(val)) else val
        if isinstance(val, str):
            return val
        # numpy scalars — np.int64/int32 are NOT subclasses of int;
        # np.float32 is NOT a subclass of float; handle explicitly
        try:
            import numpy as _np
            if isinstance(val, _np.integer):
                return int(val)
            if isinstance(val, _np.floating):
                v = float(val)
                import math
                return None if (math.isnan(v) or math.isinf(v)) else v
            if isinstance(val, _np.bool_):
                return bool(val)
        except ImportError:
            pass
        if isinstance(val, dict):
            return {str(k): _make_json_safe(v, _depth + 1) for k, v in val.items()}
        if isinstance(val, (list, tuple)):
            return [_make_json_safe(x, _depth + 1) for x in val]
        # pandas Timestamp, Decimal, and other stringifiable types
        try:
            return str(val)
        except Exception:
            # Deliberately silent: per-value guard in the JSON writer;
            # logging here could emit thousands of no-signal lines.
            return None

    json_rows = []
    for r in results:
        jr = {k: _make_json_safe(v) for k, v in r.items()}
        json_rows.append(jr)
    json_meta = {
        'date': run_start_date.isoformat(),
        'risk_free_rate': risk_free_rate,
        'risk_free_rate_source': risk_free_rate_source,
        'count': len(results),
        'provenance': _run_prov,
    }
    if macro_regime_result:
        json_meta['macro_regime'] = macro_regime_result
        json_meta['macro_adjustments'] = macro_adj
        if local_rs:
            json_meta['sector_local_rs'] = local_rs
    json_meta['results'] = json_rows
    with open(json_filename, 'w', encoding='utf-8') as f:
        json.dump(json_meta, f, indent=2, default=str)
    _prov.write_events('output')

    # Macro Outlook dashboard payload (FRED). Not gated on --macro: the tab
    # is read-only context and the FRED client caches daily. Fails soft —
    # offline or FRED-down runs simply omit the tab.
    macro_dash = None
    try:
        from data.fred_client import FREDClient
        from scripts.macro_dashboard import build_macro_payload, make_narrative_client
        macro_dash = build_macro_payload(FREDClient(), macro_regime_result,
                                         macro_adj,
                                         sector_data=sector_etf_data,
                                         local_rs=local_rs,
                                         narrative_client=make_narrative_client())
        if macro_dash:
            _n = len(macro_dash['sidecar']['series'])
            _nar = ' + Claude narrative' if macro_dash['sidecar'].get('narrative') else ''
            print(f"Macro dashboard: {_n} FRED series{_nar}")
        else:
            print("Macro dashboard skipped (no FRED data).")
    except Exception as e:
        print(f"Macro dashboard skipped ({e}).")

    html_filename = os.path.join("output", f"stock_analysis_results_{today_str}.html")
    build_html(results, html_filename, prices_dir=prices_dir, run_date=run_start_date,
               run_provenance=_run_prov, macro_payload=macro_dash)
    xlsx_filename = os.path.join("output", f"stock_analysis_results_{today_str}.xlsx")
    build_excel(results, xlsx_filename)

    print(f"\nAnalysis complete. {len(results)} stocks.")
    print(f"  HTML: {html_filename}")
    print(f"  Excel: {xlsx_filename}")
    print(f"  JSON: {json_filename}")


def _run_quality_summary(risk_free_rate, risk_free_rate_source,
                         _model_warning_counter, _prov=None):
    """End-of-run quality gate: surface substituted/fabricated inputs."""
    # Run-quality gate: surface, in one place, every way this run's numbers
    # rest on substituted rather than observed inputs.
    _log = logging.getLogger('analyze_stock')
    if risk_free_rate_source == 'fallback':
        _log.warning(
            'RUN QUALITY: risk-free rate was a hardcoded fallback — every '
            'CAPM/WACC/DCF number this run inherits a fabricated %.2f%% rate',
            risk_free_rate * 100)
    if _prov is not None:
        _beta_fallbacks = {}
        for _ev in getattr(_prov, 'events', []):
            if _ev.get('type') == 'source_fallback' and _ev.get('source') == 'beta':
                _m = (_ev.get('detail') or {}).get('method', 'unknown')
                _beta_fallbacks[_m] = _beta_fallbacks.get(_m, 0) + 1
        if _beta_fallbacks:
            _log.warning(
                'RUN QUALITY: %d tickers priced without a local regression beta '
                '(cost of equity from %s)',
                sum(_beta_fallbacks.values()),
                ', '.join(f'{k}: {v}' for k, v in sorted(_beta_fallbacks.items())))
    if _model_warning_counter.fabricated:
        _log.warning(
            'RUN QUALITY: %d model warnings flagged fabricated/fallback inputs '
            '(see WARNING lines above for tickers affected)',
            _model_warning_counter.fabricated)
    _log.info('Model warnings this run: %d total, %d about fabricated inputs',
              _model_warning_counter.total, _model_warning_counter.fabricated)


def _main():
    """Entry point: screen tickers, run DCF analysis, generate reports."""
    setup = _run_setup()
    args = setup['args']
    prices_dir = setup['prices_dir']
    run_start_date = setup['run_start_date']
    _prov = setup['_prov']
    _model_warning_counter = setup['_model_warning_counter']

    macro = _run_macro_setup(args, prices_dir)
    risk_free_rate = macro['risk_free_rate']
    risk_free_rate_source = macro['risk_free_rate_source']
    macro_regime_result = macro['macro_regime_result']
    macro_adj = macro['macro_adj']
    effective_erp = macro['effective_erp']
    effective_tg_adj = macro['effective_tg_adj']
    effective_wacc_sigma = macro['effective_wacc_sigma']
    effective_growth_sigma_mult = macro['effective_growth_sigma_mult']
    effective_exit_mult_adj = macro['effective_exit_mult_adj']
    effective_growth_weight_shift = macro['effective_growth_weight_shift']
    sector_signals = macro['sector_signals']
    commodity_data = macro['commodity_data']
    sector_etf_data = macro['sector_etf_data']
    local_rs = macro['local_rs']

    universe = _run_build_universe(args)
    ms_pfv_data = universe['ms_pfv_data']
    ticker_source = universe['ticker_source']
    all_tickers = universe['all_tickers']

    clients = _run_build_clients(run_start_date)
    yf_client = clients['yf_client']
    tiingo_client = clients['tiingo_client']
    sec_client = clients['sec_client']
    sec_xbrl_client = clients['sec_xbrl_client']

    phase1 = _run_phase1_screen(args, _prov, all_tickers, ticker_source,
                                yf_client, tiingo_client, sec_xbrl_client,
                                risk_free_rate, effective_erp)
    qualifying = phase1['qualifying']
    screen_cache = phase1['screen_cache']
    screen_outcomes = phase1['screen_outcomes']
    _carry_prior_rows = phase1['_carry_prior_rows']

    exit_mults = _run_sector_exit_multiples(qualifying, screen_cache,
                                            effective_exit_mult_adj)
    sector_exit_multiples = exit_mults['sector_exit_multiples']
    effective_exit_mult_default = exit_mults['effective_exit_mult_default']

    phase2_clients = _run_build_phase2_clients(sec_client, qualifying,
                                               screen_cache)
    news_client = phase2_clients['news_client']
    supply_client = phase2_clients['supply_client']
    sec_supply_client = phase2_clients['sec_supply_client']
    sec_insider_client = phase2_clients['sec_insider_client']
    culture_client = phase2_clients['culture_client']

    results = _run_phase2_analysis(
        qualifying, screen_cache, prices_dir, ticker_source, ms_pfv_data,
        _prov, yf_client, tiingo_client, news_client, sec_client,
        supply_client, sec_supply_client, sec_xbrl_client, sec_insider_client,
        culture_client, sector_exit_multiples, effective_exit_mult_default,
        effective_erp, effective_tg_adj, effective_wacc_sigma,
        effective_growth_sigma_mult, effective_growth_weight_shift,
        risk_free_rate, macro_regime_result, sector_signals)

    post = _run_postprocess(results, ms_pfv_data, _carry_prior_rows)
    sector_median_ee = post['sector_median_ee']
    sector_median_opm = post['sector_median_opm']

    _run_narratives(results, args, sector_etf_data, macro_regime_result,
                    commodity_data, sector_median_ee, sector_median_opm)

    _run_validation_stats(results, ms_pfv_data, args, screen_outcomes)

    _write_outputs(results, run_start_date, _prov, risk_free_rate,
                   risk_free_rate_source, macro_regime_result, macro_adj,
                   local_rs, prices_dir, sector_etf_data=sector_etf_data)

    _run_quality_summary(risk_free_rate, risk_free_rate_source,
                         _model_warning_counter, _prov)


if __name__ == "__main__":
    _main()
