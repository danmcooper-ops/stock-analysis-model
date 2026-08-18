# scripts/backtest.py
"""
Backtest + calibrate pipeline for the stock analysis model.

Subcommands:
  measure            Run backtest analysis on accumulated snapshots — measures
                     whether the model's ratings predict forward returns.
  calibrate          Walk-forward parameter calibration. Splits snapshots into
                     rolling train/test windows; for each window, sweeps the
                     parameter grid and records out-of-sample performance.
  optimize-weights   Cohen's d weight optimization on a single snapshot.
                     Searches the 4-weight valuation/quality/moat/growth grid
                     for the combo that best separates quality vs poor stocks.

Examples:
  python scripts/backtest.py measure --horizons 30,90,180
  python scripts/backtest.py calibrate --objective hit_rate --max-evals 200
  python scripts/backtest.py optimize-weights output/results_2026-04-26.json

All three modes consume the same results_YYYY-MM-DD.json snapshots produced
by analyze_stock.py and share the objective functions and snapshot loaders.
"""
import sys
import os
import json
import glob
import math
import itertools
from datetime import date, datetime, timedelta
from collections import defaultdict

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

import numpy as np

from models.utils import rank
from scripts.param_set import default_params, merge_params, validate_params


# Rating order (best → worst) for display
RATING_ORDER = ['STRONG_BUY', 'BUY', 'LEAN BUY', 'HOLD', 'PASS']
BENCHMARK = 'SPY'

# Continuous signals evaluated for return predictivity via quartile spread
# (see signal_quartile_accuracy). Each maps a field carried on every stock
# row → a human label. The whole point of the exercise: decide whether a
# signal earns a place in the composite score BEFORE trusting it with real
# position sizing. A signal only belongs in the score if its top quartile
# reliably out-earns its bottom quartile on forward EXCESS return.
SIGNAL_SPECS = [
    ('pp_multiple',     'PP Multiple (profit capture)'),
    ('pool_share_cagr', 'Pool-Share CAGR (5y trajectory)'),
    ('trap_score',      'Value-Trap Score (higher = worse expected)'),
]

# A snapshot needs at least this many stocks carrying the signal to form
# stable cross-sectional quartiles; thinner snapshots are skipped for that
# signal rather than split into noisy 3-name buckets.
MIN_SIGNAL_PER_SNAPSHOT = 20



# ===========================================================================
# BACKTEST: Forward return measurement
# ===========================================================================

# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------

def load_results(results_dir='output'):
    """Load all results JSON files, sorted by date."""
    pattern = os.path.join(results_dir, 'results_*.json')
    files = sorted(glob.glob(pattern))
    all_results = []
    for f in files:
        with open(f) as fh:
            data = json.load(fh)
            all_results.append(data)
    return all_results


# ---------------------------------------------------------------------------
# Forward return computation
# ---------------------------------------------------------------------------

# Maximum calendar-day distance allowed when snapping a target date to the
# nearest available price bar.  Beyond this, the data doesn't actually cover
# the requested window (stale parquet, not-yet-elapsed horizon) and the
# observation must be dropped rather than silently measured over the wrong
# span (or recorded as a fake 0.0 when both ends snap to the same last bar).
MAX_SNAP_GAP_DAYS = 7


def _nearest_bar(index, target, max_gap_days=MAX_SNAP_GAP_DAYS):
    """Index position of the bar nearest *target*, or None if too far away."""
    pos = index.get_indexer([target], method='nearest')[0]
    if pos < 0:
        return None
    if abs((index[pos] - target).days) > max_gap_days:
        return None
    return pos


def _fallback_period(run_dt):
    """Smallest yfinance period string that reaches back to run_dt."""
    import pandas as pd
    days_back = (pd.Timestamp.now() - run_dt).days
    for days, period in ((300, '1y'), (650, '2y'), (1750, '5y')):
        if days_back <= days:
            return period
    return 'max'


def fetch_forward_returns(tickers, run_date_str, horizon_days, yf_client,
                          prices_dir=None):
    """
    Compute forward returns for each ticker + SPY from run_date to
    run_date + horizon_days.

    When *prices_dir* is supplied and a per-ticker Parquet file exists there,
    it is used directly (fast, no network).  Falls back to yf_client for any
    ticker not found locally.

    Returns:
        dict: {ticker: {'ret': float, 'start': float, 'end': float}}
        None for tickers where data is unavailable.
    """
    import pandas as pd

    run_dt  = pd.Timestamp(run_date_str)
    eval_dt = run_dt + pd.Timedelta(days=horizon_days)
    returns = {}

    all_tickers = list(set(tickers + [BENCHMARK]))

    for ticker in all_tickers:
        try:
            hist = None

            # --- Local parquet path (preferred) ---
            if prices_dir:
                parquet = os.path.join(prices_dir, f"{ticker}.parquet")
                if os.path.exists(parquet):
                    df = pd.read_parquet(parquet)[['Close']].sort_index()
                    df.index = pd.to_datetime(df.index).tz_localize(None)
                    hist = df['Close']

            # --- Live yfinance fallback ---
            if hist is None:
                hist = yf_client.fetch_history(
                    ticker, period=_fallback_period(run_dt))
                if hist is None or len(hist) < 10:
                    continue

            # Normalise index timezone
            if hasattr(hist.index, 'tz') and hist.index.tz is not None:
                hist.index = hist.index.tz_localize(None)

            # Coverage guard: drop the observation when no bar lies within
            # MAX_SNAP_GAP_DAYS of either end (stale parquet / immature
            # horizon would otherwise be silently measured over the wrong span).
            start_idx = _nearest_bar(hist.index, run_dt)
            end_idx   = _nearest_bar(hist.index, eval_dt)
            if start_idx is None or end_idx is None:
                continue

            start_price = float(hist.iloc[start_idx])
            end_price   = float(hist.iloc[end_idx])

            # Reject non-finite prices (NaN/inf bars exist in some parquets);
            # an unguarded NaN return poisons the rank-IC objective.
            if start_price > 0 and math.isfinite(start_price) and math.isfinite(end_price):
                ret = (end_price - start_price) / start_price
                if math.isfinite(ret):
                    returns[ticker] = {
                        'ret':   ret,
                        'start': start_price,
                        'end':   end_price,
                    }
        except Exception:
            continue

    return returns


# ---------------------------------------------------------------------------
# Forward-return annotation (shared by measure + calibrate)
# ---------------------------------------------------------------------------

def is_matured(run_date_str, horizon_days, today=None):
    """True if a snapshot's forward-return window has fully elapsed.

    A (snapshot, horizon) pair is matured when ``run_date + horizon <= today``.
    Immature pairs MUST NOT be fetched: fetch_forward_returns uses a
    nearest-date lookup that would silently return the last available price
    (a truncated, wrong "forward return") for a future eval date.
    """
    today = today or date.today()
    run_dt = date.fromisoformat(str(run_date_str)[:10])
    return run_dt + timedelta(days=horizon_days) <= today


def _returns_sidecar_path(cache_dir, run_date_str, horizon_days):
    return os.path.join(cache_dir, f"{run_date_str}_h{horizon_days}.json")


def annotate_snapshot_returns(snapshot, horizons, yf_client, prices_dir=None,
                              cache_dir='output/returns', today=None):
    """Annotate one snapshot's rows in place with forward returns per horizon.

    For each MATURED horizon ``h`` (run_date + h <= today), every row that has
    price data gets::

        row['_fwd'][h] = {'excess_return', 'ret', 'end_price', 'spy_return'}

    Immature horizons are skipped entirely (no key written) so callers can
    distinguish "not matured yet" from "no price data".  Forward returns depend
    only on (date, horizon, ticker) — never on scoring params — so this is
    computed once per snapshot, outside any parameter sweep.

    Cache-aware: reads/writes a per-(date, horizon) sidecar JSON at
    ``cache_dir/{date}_h{horizon}.json``.  Matured historical returns are
    immutable, so a present sidecar is reused without refetching.  Only matured
    pairs are ever written, and any sidecar is re-validated against is_matured
    on read (a stale immature file is ignored).

    Returns:
        dict {horizon: n_rows_annotated}, with ``None`` for immature horizons.
    """
    today = today or date.today()
    run_date = snapshot.get('date')
    rows = snapshot.get('results', [])
    out = {}
    if not run_date or not rows:
        return {h: None for h in horizons}

    for h in horizons:
        if not is_matured(run_date, h, today):
            out[h] = None
            continue

        ticker_returns = None  # {ticker: {excess_return, ret, end_price, spy_return}}

        # 1. Reuse sidecar cache when present and genuinely matured.
        if cache_dir:
            path = _returns_sidecar_path(cache_dir, run_date, h)
            if os.path.exists(path):
                try:
                    with open(path) as f:
                        cached = json.load(f)
                    if is_matured(cached.get('run_date', run_date),
                                  cached.get('horizon_days', h), today):
                        ticker_returns = cached.get('tickers', {})
                except Exception:
                    ticker_returns = None

        # 2. Otherwise compute fresh via the shared fetch helper and cache it.
        if ticker_returns is None:
            tickers = [r['ticker'] for r in rows if r.get('ticker')]
            raw = fetch_forward_returns(tickers, run_date, h, yf_client,
                                        prices_dir=prices_dir)
            spy = raw.get(BENCHMARK)
            spy_ret = spy['ret'] if spy else 0.0
            ticker_returns = {}
            for t, v in raw.items():
                if t == BENCHMARK:
                    continue
                ticker_returns[t] = {
                    'excess_return': v['ret'] - spy_ret,
                    'ret': v['ret'],
                    'start_price': v['start'],
                    'end_price': v['end'],
                    'spy_return': spy_ret,
                }
            # Only persist a non-empty result. An empty set means the price
            # data didn't reach the eval date (stale parquet) — caching it would
            # poison future runs after prices are refreshed, since sidecars are
            # re-validated only against is_matured, not against data coverage.
            if cache_dir and ticker_returns:
                try:
                    os.makedirs(cache_dir, exist_ok=True)
                    with open(_returns_sidecar_path(cache_dir, run_date, h), 'w') as f:
                        json.dump({
                            'run_date': run_date,
                            'horizon_days': h,
                            'computed_at': today.isoformat(),
                            'spy_return': spy_ret,
                            'tickers': ticker_returns,
                        }, f)
                except Exception:
                    pass

        # 3. Annotate rows in place (horizon-keyed).
        n = 0
        for r in rows:
            t = r.get('ticker')
            fr = ticker_returns.get(t) if t else None
            if fr is None:
                continue
            r.setdefault('_fwd', {})[h] = fr
            n += 1
        out[h] = n

    return out


# ---------------------------------------------------------------------------
# Single-run analysis
# ---------------------------------------------------------------------------

def analyze_run(run, horizon_days, yf_client, prices_dir=None,
                cache_dir='output/returns', today=None):
    """
    Analyze one snapshot at one horizon.

    Forward returns come from the shared ``annotate_snapshot_returns`` helper
    (same code path + sidecar cache as calibrate), which also enforces the
    maturity guard — an immature (snapshot, horizon) yields no annotated rows
    rather than silently-truncated returns.

    Returns dict with:
        run_date, horizon, spy_return,
        buckets: {rating: {mean, median, count, alpha, hit_rate}},
        gates_corr: Spearman rho between gates_passed and return,
        fv_metrics: {mae, signed_error, within_20, n},
        details: [{ticker, rating, gates, fv, price, end_price, ret, excess}]
    """
    run_date = run.get('date')
    stocks = run.get('results', [])
    if not run_date or not stocks:
        return None

    annotate_snapshot_returns(run, [horizon_days], yf_client,
                              prices_dir=prices_dir, cache_dir=cache_dir,
                              today=today)

    # SPY return is identical across rows for a given (date, horizon).
    spy_ret = 0.0
    for s in stocks:
        fwd = (s.get('_fwd') or {}).get(horizon_days)
        if fwd:
            spy_ret = fwd.get('spy_return', 0.0)
            break

    # Build detail rows
    details = []
    bucket_returns = defaultdict(list)  # rating -> list of returns
    gates_vals = []
    return_vals = []
    fv_preds = []
    fv_actuals = []

    for s in stocks:
        fwd = (s.get('_fwd') or {}).get(horizon_days)
        if not fwd:
            continue

        ticker = s.get('ticker')
        rating = s.get('rating', 'UNKNOWN')
        gates = s.get('_gates_passed_num', 0)
        fv = s.get('dcf_fv')
        ret = fwd['ret']
        end_price = fwd['end_price']
        excess = fwd['excess_return']

        detail = {
            'ticker': ticker,
            'rating': rating,
            'gates_passed': s.get('_gates_passed', 'N/A'),
            'gates_num': gates,
            'dcf_fv': fv,
            'start_price': fwd.get('start_price'),
            'end_price': end_price,
            'return': ret,
            'spy_return': spy_ret,
            'excess_return': excess,
        }
        details.append(detail)

        bucket_returns[rating].append(ret)

        if isinstance(gates, (int, float)) and gates >= 0:
            gates_vals.append(gates)
            return_vals.append(ret)

        if fv and fv > 0 and end_price > 0:
            fv_preds.append(fv)
            fv_actuals.append(end_price)

    if not details:
        return None

    # --- Rating bucket stats ---
    buckets = {}
    for rating in RATING_ORDER:
        rets = bucket_returns.get(rating, [])
        if not rets:
            continue
        arr = np.array(rets)
        buckets[rating] = {
            'mean': float(np.mean(arr)),
            'median': float(np.median(arr)),
            'count': len(rets),
            'alpha': float(np.mean(arr)) - spy_ret,
            'hit_rate': float(np.sum(arr > spy_ret) / len(arr)),
        }

    # --- Gates-passed correlation ---
    gates_corr = None
    if len(gates_vals) >= 5:
        n = len(gates_vals)
        rank_g = rank(gates_vals)
        rank_r = rank(return_vals)
        d_sq = sum((rg - rr) ** 2 for rg, rr in zip(rank_g, rank_r))
        gates_corr = 1 - (6 * d_sq) / (n * (n ** 2 - 1)) if n > 1 else 0.0

    # --- FV accuracy ---
    fv_metrics = None
    if len(fv_preds) >= 5:
        preds = np.array(fv_preds)
        acts = np.array(fv_actuals)
        pct_err = (preds - acts) / acts
        fv_metrics = {
            'mae': float(np.mean(np.abs(pct_err))),
            'signed_error': float(np.mean(pct_err)),
            'within_20': float(np.sum(np.abs(pct_err) <= 0.20) / len(pct_err)),
            'n': len(pct_err),
        }

    # Store source stock data for enhanced analytics (sector, MoS, targets)
    source_stocks = {}
    for s in stocks:
        t = s.get('ticker')
        if t:
            source_stocks[t] = {
                'sector': s.get('sector'),
                'dcf_fv': s.get('dcf_fv'),
                'target_mean': s.get('target_mean'),
                'rating': s.get('rating'),
                # Profit-pool signals under evaluation (see signal_quartile_
                # accuracy). pp_multiple = disproportionate profit capture;
                # pool_share_cagr = 5y trajectory of sector-profit-pool share.
                'pp_multiple': s.get('pp_multiple'),
                'pool_share_cagr': s.get('pool_share_cagr'),
                # Value-trap overlay under evaluation, plus the cohort /
                # incrementality inputs filter_metrics_to_cohort reads:
                # mos defines the cheap cohort, beneish/altman identify rows
                # the existing rating caps already veto.
                'trap_score': s.get('trap_score'),
                'trap_score_ex_momentum': s.get('trap_score_ex_momentum'),
                'mos': s.get('mos'),
                'beneish_flag': s.get('beneish_flag'),
                'altman_z_zone': s.get('altman_z_zone'),
            }

    return {
        'run_date': run_date,
        'horizon': horizon_days,
        'n_stocks': len(details),
        'spy_return': spy_ret,
        'buckets': buckets,
        'gates_corr': gates_corr,
        'fv_metrics': fv_metrics,
        'details': details,
        '_source_stocks': source_stocks,
    }


# ---------------------------------------------------------------------------
# Full backtest across all snapshots and horizons
# ---------------------------------------------------------------------------

def run_backtest(results_dir, horizons, yf_client, prices_dir=None):
    """Run backtest for all snapshots × horizons. Returns list of result dicts."""
    all_results = load_results(results_dir)
    if not all_results:
        print("No results files found in", results_dir)
        return []

    print(f"Loaded {len(all_results)} snapshot(s).")
    for r in all_results:
        print(f"  {r.get('date', '?')}: {r.get('count', 0)} stocks")

    metrics = []
    skipped = 0

    for run in all_results:
        run_date_str = run.get('date')
        if not run_date_str:
            continue
        run_dt = datetime.strptime(run_date_str, '%Y-%m-%d')

        for h in horizons:
            eval_dt = run_dt + timedelta(days=h)
            if eval_dt > datetime.now():
                skipped += 1
                print(f"  {run_date_str} + {h}d → {eval_dt.date()}: skipped (future)")
                continue

            print(f"\n  Analyzing {run_date_str} + {h}d → {eval_dt.date()} ...")
            result = analyze_run(run, h, yf_client, prices_dir=prices_dir)
            if result:
                metrics.append(result)

    if skipped and not metrics:
        print(f"\nAll {skipped} snapshot-horizon pairs have evaluation dates in the future.")
        print("Continue running analyze_stock.py over time — the backtest will activate")
        print("once enough time has passed for forward returns to be measured.")

    return metrics


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

def build_backtest_excel(all_metrics, filename):
    """Write backtest results to a styled Excel workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    gray = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    white = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    hdr_font = Font(bold=True, color='000000', size=11)
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    data_font = Font(color='000000')
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

    def write_header(ws, headers):
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = hdr_font
            cell.fill = gray
            cell.alignment = hdr_align

    def style_row(ws, row, n_cols, frozen=1):
        for ci in range(1, n_cols + 1):
            cell = ws.cell(row=row, column=ci)
            if ci <= frozen:
                cell.fill = gray
                cell.font = Font(bold=True, color='000000')
            else:
                cell.fill = white
                cell.font = data_font

    def auto_width(ws, headers):
        for ci, h in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(ci)].width = max(len(h) + 3, 12)

    # ---- Summary tab ----
    ws = wb.active
    ws.title = 'Summary'
    summary_headers = [
        'Snapshot', 'Horizon (d)', '# Stocks', 'SPY Return',
        'BUY Mean', 'BUY Alpha', 'BUY Hit Rate', 'BUY #',
        'LEAN BUY Mean', 'LEAN BUY Alpha', 'LEAN BUY #',
        'HOLD Mean', 'HOLD Alpha', 'HOLD #',
        'PASS Mean', 'PASS Alpha', 'PASS #',
        'Gates ρ', 'FV MAE', 'FV Bias', 'FV ±20%',
    ]
    write_header(ws, summary_headers)

    for ri, m in enumerate(all_metrics, 2):
        b = m['buckets']
        fv = m.get('fv_metrics') or {}
        row_data = [
            m['run_date'], m['horizon'], m['n_stocks'],
            m['spy_return'],
            # BUY
            b.get('BUY', {}).get('mean'),
            b.get('BUY', {}).get('alpha'),
            b.get('BUY', {}).get('hit_rate'),
            b.get('BUY', {}).get('count'),
            # LEAN BUY
            b.get('LEAN BUY', {}).get('mean'),
            b.get('LEAN BUY', {}).get('alpha'),
            b.get('LEAN BUY', {}).get('count'),
            # HOLD
            b.get('HOLD', {}).get('mean'),
            b.get('HOLD', {}).get('alpha'),
            b.get('HOLD', {}).get('count'),
            # PASS
            b.get('PASS', {}).get('mean'),
            b.get('PASS', {}).get('alpha'),
            b.get('PASS', {}).get('count'),
            # Correlation & FV
            m.get('gates_corr'),
            fv.get('mae'), fv.get('signed_error'), fv.get('within_20'),
        ]
        for ci, val in enumerate(row_data, 1):
            ws.cell(row=ri, column=ci, value=val)
        style_row(ws, ri, len(summary_headers), frozen=2)

        # Format percentages
        for ci in [4, 5, 6, 7, 9, 10, 12, 13, 15, 16, 19, 20, 21]:
            cell = ws.cell(row=ri, column=ci)
            if cell.value is not None:
                cell.number_format = '0.0%'
        ws.cell(row=ri, column=18).number_format = '0.000'

    auto_width(ws, summary_headers)
    ws.freeze_panes = 'C2'
    ws.row_dimensions[1].height = 13

    # ---- Detail tab ----
    ws2 = wb.create_sheet('Detail')
    detail_headers = [
        'Snapshot', 'Horizon (d)', 'Ticker', 'Rating', 'Gates Passed',
        'DCF Fair Value', 'Start Price', 'End Price',
        'Return', 'SPY Return', 'Excess Return',
    ]
    write_header(ws2, detail_headers)

    row_num = 2
    for m in all_metrics:
        for d in m['details']:
            row_data = [
                m['run_date'], m['horizon'], d['ticker'], d['rating'],
                d['gates_passed'], d.get('dcf_fv'), d['start_price'],
                d['end_price'], d['return'], d['spy_return'],
                d['excess_return'],
            ]
            for ci, val in enumerate(row_data, 1):
                ws2.cell(row=row_num, column=ci, value=val)
            style_row(ws2, row_num, len(detail_headers), frozen=3)

            # Format
            for ci in [6, 7, 8]:
                ws2.cell(row=row_num, column=ci).number_format = '"$"#,##0.00'
            for ci in [9, 10, 11]:
                ws2.cell(row=row_num, column=ci).number_format = '0.0%'

            # Color excess return
            cell = ws2.cell(row=row_num, column=11)
            if cell.value is not None:
                cell.fill = green_fill if cell.value > 0 else red_fill

            row_num += 1

    auto_width(ws2, detail_headers)
    ws2.freeze_panes = 'D2'
    for r in range(1, row_num):
        ws2.row_dimensions[r].height = 13

    # ---- Stats tab (aggregated across all snapshots) ----
    ws3 = wb.create_sheet('Stats')
    ws3.cell(row=1, column=1, value='Aggregated Rating Performance')
    ws3.cell(row=1, column=1).font = Font(bold=True, size=13)

    stats_headers = ['Rating', 'Total Obs', 'Mean Return', 'Median Return',
                     'Mean Alpha', 'Hit Rate', 'Avg Gates']
    for ci, h in enumerate(stats_headers, 1):
        cell = ws3.cell(row=3, column=ci, value=h)
        cell.font = hdr_font
        cell.fill = gray
        cell.alignment = hdr_align

    # Aggregate across all runs
    agg = defaultdict(lambda: {'returns': [], 'alphas': [], 'hits': 0, 'n': 0})
    all_gates = []
    all_returns = []

    for m in all_metrics:
        spy_ret = m['spy_return']
        for d in m['details']:
            rating = d['rating']
            agg[rating]['returns'].append(d['return'])
            agg[rating]['alphas'].append(d['excess_return'])
            if d['return'] > spy_ret:
                agg[rating]['hits'] += 1
            agg[rating]['n'] += 1
            if d.get('gates_num', -1) >= 0:
                all_gates.append(d['gates_num'])
                all_returns.append(d['return'])

    ri = 4
    for rating in RATING_ORDER:
        a = agg.get(rating)
        if not a or not a['returns']:
            continue
        rets = np.array(a['returns'])
        alps = np.array(a['alphas'])
        ws3.cell(row=ri, column=1, value=rating)
        ws3.cell(row=ri, column=2, value=a['n'])
        ws3.cell(row=ri, column=3, value=float(np.mean(rets)))
        ws3.cell(row=ri, column=4, value=float(np.median(rets)))
        ws3.cell(row=ri, column=5, value=float(np.mean(alps)))
        ws3.cell(row=ri, column=6, value=a['hits'] / a['n'] if a['n'] else 0)
        ws3.cell(row=ri, column=7, value='')
        style_row(ws3, ri, 7, frozen=1)
        for ci in [3, 4, 5, 6]:
            ws3.cell(row=ri, column=ci).number_format = '0.0%'
        ri += 1

    # Gates-passed correlation (aggregated)
    ri += 1
    ws3.cell(row=ri, column=1, value='Gates-Passed Correlation')
    ws3.cell(row=ri, column=1).font = Font(bold=True, size=13)
    ri += 1
    if len(all_gates) >= 5:
        n = len(all_gates)
        rank_g = rank(all_gates)
        rank_r = rank(all_returns)
        d_sq = sum((rg - rr) ** 2 for rg, rr in zip(rank_g, rank_r))
        agg_rho = 1 - (6 * d_sq) / (n * (n ** 2 - 1)) if n > 1 else 0.0
        ws3.cell(row=ri, column=1, value='Spearman ρ (gates vs return)')
        ws3.cell(row=ri, column=2, value=round(agg_rho, 4))
        ws3.cell(row=ri, column=2).number_format = '0.0000'
        ri += 1
        ws3.cell(row=ri, column=1, value='Observations')
        ws3.cell(row=ri, column=2, value=n)
    else:
        ws3.cell(row=ri, column=1, value='Not enough data yet')

    # FV accuracy (aggregated)
    ri += 2
    ws3.cell(row=ri, column=1, value='Fair Value Accuracy')
    ws3.cell(row=ri, column=1).font = Font(bold=True, size=13)
    ri += 1
    all_fv_preds = []
    all_fv_acts = []
    for m in all_metrics:
        for d in m['details']:
            fv = d.get('dcf_fv')
            ep = d.get('end_price')
            if fv and fv > 0 and ep and ep > 0:
                all_fv_preds.append(fv)
                all_fv_acts.append(ep)

    if len(all_fv_preds) >= 5:
        preds = np.array(all_fv_preds)
        acts = np.array(all_fv_acts)
        pct_err = (preds - acts) / acts
        metrics_labels = [
            ('Mean Absolute Error', float(np.mean(np.abs(pct_err)))),
            ('Mean Signed Error (bias)', float(np.mean(pct_err))),
            ('Within ±20%', float(np.sum(np.abs(pct_err) <= 0.20) / len(pct_err))),
            ('Observations', len(pct_err)),
        ]
        for label, val in metrics_labels:
            ws3.cell(row=ri, column=1, value=label)
            ws3.cell(row=ri, column=2, value=val)
            if isinstance(val, float):
                ws3.cell(row=ri, column=2).number_format = '0.0%'
            ri += 1
    else:
        ws3.cell(row=ri, column=1, value='Not enough data yet')

    auto_width(ws3, stats_headers)
    ws3.column_dimensions['A'].width = 30
    ws3.freeze_panes = 'A4'
    for r in range(1, ri + 1):
        ws3.row_dimensions[r].height = 13

    # ---- Sector Accuracy tab ----
    ws4 = wb.create_sheet('Sector Accuracy')
    sec_acc = sector_accuracy(all_metrics)
    sec_headers = ['Sector', 'Count', 'Mean Return', 'Mean Alpha', 'Hit Rate']
    write_header(ws4, sec_headers)

    ri = 2
    for sector in sorted(sec_acc, key=lambda s: sec_acc[s]['mean_alpha'], reverse=True):
        sa = sec_acc[sector]
        ws4.cell(row=ri, column=1, value=sector)
        ws4.cell(row=ri, column=2, value=sa['count'])
        ws4.cell(row=ri, column=3, value=sa['mean_return'])
        ws4.cell(row=ri, column=4, value=sa['mean_alpha'])
        ws4.cell(row=ri, column=5, value=sa['hit_rate'])
        style_row(ws4, ri, 5, frozen=1)
        for ci in [3, 4, 5]:
            ws4.cell(row=ri, column=ci).number_format = '0.0%'
        ri += 1

    auto_width(ws4, sec_headers)
    ws4.column_dimensions['A'].width = 25
    ws4.freeze_panes = 'B2'
    for r in range(1, ri):
        ws4.row_dimensions[r].height = 13

    # ---- MoS Buckets tab ----
    ws5 = wb.create_sheet('MoS Buckets')
    mos_acc = mos_bucket_accuracy(all_metrics)
    mos_headers = ['MoS Bucket', 'Count', 'Mean Return', 'Hit Rate']
    write_header(ws5, mos_headers)

    ri = 2
    for bucket in ['>30%', '10-30%', '0-10%', 'Negative']:
        if bucket in mos_acc:
            ma = mos_acc[bucket]
            ws5.cell(row=ri, column=1, value=bucket)
            ws5.cell(row=ri, column=2, value=ma['count'])
            ws5.cell(row=ri, column=3, value=ma['mean_return'])
            ws5.cell(row=ri, column=4, value=ma['hit_rate'])
            style_row(ws5, ri, 4, frozen=1)
            for ci in [3, 4]:
                ws5.cell(row=ri, column=ci).number_format = '0.0%'
            ri += 1

    auto_width(ws5, mos_headers)
    ws5.freeze_panes = 'B2'
    for r in range(1, ri):
        ws5.row_dimensions[r].height = 13

    # ---- Consensus tab ----
    ws6 = wb.create_sheet('Consensus')
    ws6.cell(row=1, column=1, value='Model vs Analyst Targets')
    ws6.cell(row=1, column=1).font = Font(bold=True, size=13)

    consensus = consensus_comparison(all_metrics)
    buy_hit = strong_buy_hit_rate(all_metrics)

    ri = 3
    if buy_hit is not None:
        ws6.cell(row=ri, column=1, value='BUY Hit Rate (beat SPY)')
        ws6.cell(row=ri, column=1).font = hdr_font
        ws6.cell(row=ri, column=2, value=buy_hit)
        ws6.cell(row=ri, column=2).number_format = '0.0%'
        ri += 2

    if consensus:
        ws6.cell(row=ri, column=1, value='Metric')
        ws6.cell(row=ri, column=1).font = hdr_font
        ws6.cell(row=ri, column=1).fill = gray
        ws6.cell(row=ri, column=2, value='Value')
        ws6.cell(row=ri, column=2).font = hdr_font
        ws6.cell(row=ri, column=2).fill = gray
        ri += 1
        metrics_data = [
            ('Mean Bias (Model/Target − 1)', consensus['mean_bias']),
            ('Median Bias', consensus['median_bias']),
            ('# Stocks Compared', consensus['n_stocks']),
        ]
        for label, val in metrics_data:
            ws6.cell(row=ri, column=1, value=label)
            ws6.cell(row=ri, column=2, value=val)
            if isinstance(val, float):
                ws6.cell(row=ri, column=2).number_format = '0.0%'
            style_row(ws6, ri, 2, frozen=1)
            ri += 1
    else:
        ws6.cell(row=ri, column=1, value='Not enough data for consensus comparison')

    ws6.column_dimensions['A'].width = 35
    ws6.column_dimensions['B'].width = 15
    for r in range(1, ri + 1):
        ws6.row_dimensions[r].height = 13

    # ---- Signal Quartiles tab ----
    ws7 = wb.create_sheet('Signal Quartiles')
    ws7.cell(row=1, column=1,
             value='Forward Alpha by Signal Quartile '
                   '(Q1 = lowest signal, Q4 = highest)')
    ws7.cell(row=1, column=1).font = Font(bold=True, size=13)
    ws7.cell(row=2, column=1,
             value='A signal earns a place in the composite score only if '
                   'Q4 reliably out-earns Q1 on forward excess return. '
                   'Obs are NOT independent (overlapping daily snapshots) — '
                   'weigh the per-snapshot ρ sign-consistency, not the raw n.')
    ws7.cell(row=2, column=1).font = Font(italic=True, size=10, color='555555')

    sig_headers = ['Signal', 'Quartile', 'Mean Alpha', 'Median Alpha',
                   'Hit Rate', 'Count', 'Q4−Q1 Spread', 'Mean Snap ρ',
                   'ρ>0 Frac', 'Snapshots']
    for ci, h in enumerate(sig_headers, 1):
        cell = ws7.cell(row=4, column=ci, value=h)
        cell.font = hdr_font
        cell.fill = gray
        cell.alignment = hdr_align

    ri = 5
    for key, label in SIGNAL_SPECS:
        sq = signal_quartile_accuracy(all_metrics, key)
        if sq is None:
            ws7.cell(row=ri, column=1, value=label)
            ws7.cell(row=ri, column=2, value='not enough data yet')
            style_row(ws7, ri, len(sig_headers), frozen=1)
            ri += 1
            continue
        first = True
        for q in (1, 2, 3, 4):
            qd = sq['quartiles'].get(q)
            if not qd:
                continue
            ws7.cell(row=ri, column=1, value=label if first else '')
            ws7.cell(row=ri, column=2, value='Q%d' % q)
            ws7.cell(row=ri, column=3, value=qd['mean_alpha'])
            ws7.cell(row=ri, column=4, value=qd['median_alpha'])
            ws7.cell(row=ri, column=5, value=qd['hit_rate'])
            ws7.cell(row=ri, column=6, value=qd['count'])
            if first:
                ws7.cell(row=ri, column=7, value=sq['spread'])
                ws7.cell(row=ri, column=8, value=sq['mean_rho'])
                ws7.cell(row=ri, column=9, value=sq['rho_pos_frac'])
                ws7.cell(row=ri, column=10, value=sq['n_snapshots'])
            style_row(ws7, ri, len(sig_headers), frozen=1)
            for ci in (3, 4, 5, 7, 9):
                c = ws7.cell(row=ri, column=ci)
                if c.value is not None:
                    c.number_format = '0.0%'
            ws7.cell(row=ri, column=8).number_format = '0.000'
            # Color the spread green/red on the signal's first row.
            if first and sq['spread'] is not None:
                sc = ws7.cell(row=ri, column=7)
                sc.fill = green_fill if sq['spread'] > 0 else red_fill
            first = False
            ri += 1

    auto_width(ws7, sig_headers)
    ws7.column_dimensions['A'].width = 32
    ws7.freeze_panes = 'C5'
    for r in range(1, ri + 1):
        ws7.row_dimensions[r].height = 13


    wb.save(filename)
    return filename


# ---------------------------------------------------------------------------
# Enhanced analytics
# ---------------------------------------------------------------------------

def sector_accuracy(all_metrics):
    """Compute per-sector return and hit rate stats across all runs.

    Returns dict: {sector: {'mean_return': float, 'hit_rate': float,
                             'count': int, 'mean_alpha': float}}
    """
    sector_data = defaultdict(lambda: {'returns': [], 'alphas': []})

    for m in all_metrics:
        spy_ret = m['spy_return']
        stocks = m.get('_source_stocks', {})
        for d in m['details']:
            ticker = d['ticker']
            stock_info = stocks.get(ticker, {})
            sector = stock_info.get('sector', 'Unknown')
            sector_data[sector]['returns'].append(d['return'])
            sector_data[sector]['alphas'].append(d['excess_return'])

    result = {}
    for sector, data in sector_data.items():
        rets = np.array(data['returns'])
        alps = np.array(data['alphas'])
        result[sector] = {
            'mean_return': float(np.mean(rets)),
            'hit_rate': float(np.sum(alps > 0) / len(alps)) if len(alps) > 0 else 0,
            'count': len(rets),
            'mean_alpha': float(np.mean(alps)),
        }
    return result


def mos_bucket_accuracy(all_metrics):
    """Compute accuracy by margin-of-safety buckets.

    Buckets: >30%, 10-30%, 0-10%, negative.
    Returns dict: {bucket_label: {'mean_return': float, 'hit_rate': float, 'count': int}}
    """
    buckets = {
        '>30%': {'returns': [], 'alphas': []},
        '10-30%': {'returns': [], 'alphas': []},
        '0-10%': {'returns': [], 'alphas': []},
        'Negative': {'returns': [], 'alphas': []},
    }

    for m in all_metrics:
        stocks = m.get('_source_stocks', {})
        for d in m['details']:
            ticker = d['ticker']
            stock_info = stocks.get(ticker, {})
            fv = stock_info.get('dcf_fv') or d.get('dcf_fv')
            price = d.get('start_price')
            if fv and fv > 0 and price and price > 0:
                mos = (fv - price) / fv
                if mos > 0.30:
                    key = '>30%'
                elif mos > 0.10:
                    key = '10-30%'
                elif mos >= 0:
                    key = '0-10%'
                else:
                    key = 'Negative'
                buckets[key]['returns'].append(d['return'])
                buckets[key]['alphas'].append(d['excess_return'])

    result = {}
    for label, data in buckets.items():
        if data['returns']:
            rets = np.array(data['returns'])
            alps = np.array(data['alphas'])
            result[label] = {
                'mean_return': float(np.mean(rets)),
                'hit_rate': float(np.sum(alps > 0) / len(alps)) if len(alps) > 0 else 0,
                'count': len(rets),
            }
    return result


def strong_buy_hit_rate(all_metrics):
    """What percentage of BUY-rated stocks actually outperformed SPY?

    Returns float (0-1) or None if no BUY stocks.
    """
    buy_alphas = []
    for m in all_metrics:
        for d in m['details']:
            if d['rating'] == 'BUY':
                buy_alphas.append(d['excess_return'])
    if not buy_alphas:
        return None
    return sum(1 for a in buy_alphas if a > 0) / len(buy_alphas)


def consensus_comparison(all_metrics):
    """Compare model fair values vs analyst target prices.

    Returns dict with: mean_bias, median_bias, n_stocks.
    """
    biases = []
    for m in all_metrics:
        stocks = m.get('_source_stocks', {})
        for d in m['details']:
            ticker = d['ticker']
            stock_info = stocks.get(ticker, {})
            model_fv = stock_info.get('dcf_fv') or d.get('dcf_fv')
            target_mean = stock_info.get('target_mean')
            if model_fv and model_fv > 0 and target_mean and target_mean > 0:
                biases.append(model_fv / target_mean - 1)
    if not biases:
        return None
    arr = np.array(biases)
    return {
        'mean_bias': float(np.mean(arr)),
        'median_bias': float(np.median(arr)),
        'n_stocks': len(arr),
    }


# ---------------------------------------------------------------------------
# Console summary
# ---------------------------------------------------------------------------

def signal_quartile_accuracy(all_metrics, signal_key,
                             min_per_snapshot=MIN_SIGNAL_PER_SNAPSHOT):
    """Forward EXCESS return by quartile of a continuous signal.

    Method (deliberately conservative about the daily-snapshot data shape):

      1. Rank stocks into quartiles CROSS-SECTIONALLY within each snapshot
         (Q1 = lowest signal value, Q4 = highest), so a single day's market
         backdrop can't tilt the split and every snapshot contributes a
         balanced four-way partition. Excess return (vs SPY) is the outcome,
         so the market move is already stripped out.
      2. Pool each quartile's excess returns across all evaluable snapshots.
      3. Report the Q4−Q1 spread — the money question. A signal worth adding
         to the composite score shows a positive, monotone-ish spread.

    Independence caveat: consecutive daily snapshots hold nearly the same
    universe over overlapping windows, so the pooled observations are highly
    autocorrelated — n_obs OVERSTATES the true sample size. To counter that,
    we also compute a Fama–MacBeth-style average of PER-SNAPSHOT Spearman
    rank correlations (signal vs excess return) plus the fraction of
    snapshots where that correlation is positive: a sign-consistency check
    that treats each snapshot as one observation, not each stock.

    Returns None if no snapshot carries >= min_per_snapshot signal values,
    else a dict:
        quartiles:    {1..4: {mean_alpha, median_alpha, hit_rate, count}}
        spread:       Q4 mean_alpha − Q1 mean_alpha (or None)
        mean_rho:     mean per-snapshot Spearman rho (or None)
        rho_pos_frac: fraction of snapshots with rho > 0 (or None)
        n_snapshots:  distinct snapshots contributing
        n_obs:        total (non-independent) stock-observations
    """
    quartile_alphas = {1: [], 2: [], 3: [], 4: []}
    per_snap_rho = []
    n_snapshots = 0
    n_obs = 0

    for m in all_metrics:
        stocks = m.get('_source_stocks', {})
        obs = []  # (signal_value, excess_return)
        for d in m['details']:
            info = stocks.get(d['ticker'], {})
            sig = info.get(signal_key)
            ex = d['excess_return']
            # Guard against non-finite inputs: a NaN Close in a parquet file
            # propagates ticker → ret → excess_return, and a single NaN poisons
            # np.mean for the whole quartile. Require both signal and outcome
            # finite. (This NaN leak also affects the sector/MoS stats above;
            # tracked separately.)
            if (isinstance(sig, bool) or not isinstance(sig, (int, float))
                    or not np.isfinite(sig)
                    or not isinstance(ex, (int, float)) or not np.isfinite(ex)):
                continue
            obs.append((float(sig), float(ex)))

        if len(obs) < min_per_snapshot:
            continue

        n_snapshots += 1
        n_obs += len(obs)
        sigs = [o[0] for o in obs]
        exs = [o[1] for o in obs]
        n = len(obs)

        # Cross-sectional quartile via 0-based ordinal rank. argsort-of-
        # argsort breaks ties arbitrarily; acceptable for near-continuous
        # ratios where exact ties are rare.
        ordinal = np.argsort(np.argsort(sigs))
        for i, ex in enumerate(exs):
            q = min(int(ordinal[i] / n * 4), 3)
            quartile_alphas[q + 1].append(ex)

        # Per-snapshot Spearman (reuse the file's tie-averaged rank()).
        rank_s = rank(sigs)
        rank_e = rank(exs)
        d_sq = sum((rs - re) ** 2 for rs, re in zip(rank_s, rank_e))
        per_snap_rho.append(1 - (6 * d_sq) / (n * (n ** 2 - 1)) if n > 1 else 0.0)

    if n_snapshots == 0:
        return None

    quartiles = {}
    for q in (1, 2, 3, 4):
        arr = np.array(quartile_alphas[q])
        if len(arr) == 0:
            continue
        quartiles[q] = {
            'mean_alpha': float(np.mean(arr)),
            'median_alpha': float(np.median(arr)),
            'hit_rate': float(np.sum(arr > 0) / len(arr)),
            'count': len(arr),
        }

    q4 = quartiles.get(4, {}).get('mean_alpha')
    q1 = quartiles.get(1, {}).get('mean_alpha')
    spread = (q4 - q1) if (q4 is not None and q1 is not None) else None

    return {
        'quartiles': quartiles,
        'spread': spread,
        'mean_rho': float(np.mean(per_snap_rho)) if per_snap_rho else None,
        'rho_pos_frac': (float(np.mean([1.0 if r > 0 else 0.0
                                        for r in per_snap_rho]))
                         if per_snap_rho else None),
        'n_snapshots': n_snapshots,
        'n_obs': n_obs,
    }


def filter_metrics_to_cohort(all_metrics, cohort_key='mos',
                             top_quartile=True, exclude_capped=False):
    """Restrict each snapshot's details to a cross-sectional cohort.

    The value-trap question is conditional: an unconditional trap read mixes
    expensive decliners (which the model already avoids) into the sample.
    This filter keeps, per snapshot, only the top quartile by `cohort_key`
    (default `mos` — the statistically-cheap cohort), so a downstream
    signal_quartile_accuracy call measures the signal WITHIN the names the
    model would actually consider buying.

    exclude_capped additionally drops rows already vetoed by the existing
    distress caps (beneish_flag / altman distress) — the incrementality
    check: a trap signal that only re-flags those adds nothing over
    _rating_cap_for_row.

    Returns a new all_metrics list with the same dict shape (details filtered,
    _source_stocks passed through untouched); snapshots whose cohort would be
    empty are dropped.
    """
    out = []
    for m in all_metrics:
        stocks = m.get('_source_stocks', {})
        rows = []
        for d in m['details']:
            info = stocks.get(d['ticker'], {})
            cv = info.get(cohort_key)
            if isinstance(cv, bool) or not isinstance(cv, (int, float)) \
                    or not np.isfinite(cv):
                continue
            if exclude_capped and (
                    info.get('beneish_flag') is True
                    or info.get('altman_z_zone') == 'distress'):
                continue
            rows.append((float(cv), d))
        if len(rows) < 4:
            continue
        rows.sort(key=lambda x: x[0])
        cut = int(len(rows) * 0.75)
        kept = [d for _, d in (rows[cut:] if top_quartile else rows[:cut])]
        if not kept:
            continue
        fm = dict(m)
        fm['details'] = kept
        out.append(fm)
    return out


def strong_buy_hit_rate(all_metrics):
    """What percentage of BUY-rated stocks actually outperformed SPY?

    Returns float (0-1) or None if no BUY stocks.
    """
    buy_alphas = []
    for m in all_metrics:
        for d in m['details']:
            if d['rating'] == 'BUY':
                buy_alphas.append(d['excess_return'])
    if not buy_alphas:
        return None
    return sum(1 for a in buy_alphas if a > 0) / len(buy_alphas)


def consensus_comparison(all_metrics):
    """Compare model fair values vs analyst target prices.

    Returns dict with: mean_bias, median_bias, n_stocks.
    """
    biases = []
    for m in all_metrics:
        stocks = m.get('_source_stocks', {})
        for d in m['details']:
            ticker = d['ticker']
            stock_info = stocks.get(ticker, {})
            model_fv = stock_info.get('dcf_fv') or d.get('dcf_fv')
            target_mean = stock_info.get('target_mean')
            if model_fv and model_fv > 0 and target_mean and target_mean > 0:
                biases.append(model_fv / target_mean - 1)
    if not biases:
        return None
    arr = np.array(biases)
    return {
        'mean_bias': float(np.mean(arr)),
        'median_bias': float(np.median(arr)),
        'n_stocks': len(arr),
    }


def print_summary(all_metrics):
    """Print a concise console summary of backtest results."""
    if not all_metrics:
        return

    print(f"\n{'='*70}")
    print("BACKTEST RESULTS")
    print(f"{'='*70}")

    for m in all_metrics:
        b = m['buckets']
        print(f"\n{m['run_date']} + {m['horizon']}d  ({m['n_stocks']} stocks, SPY {m['spy_return']:+.1%})")
        print(f"  {'Rating':<12s} {'Mean':>8s} {'Alpha':>8s} {'Hit%':>6s} {'#':>4s}")
        print(f"  {'-'*40}")
        for rating in RATING_ORDER:
            if rating in b:
                r = b[rating]
                print(f"  {rating:<12s} {r['mean']:>+7.1%} {r['alpha']:>+7.1%}"
                      f" {r['hit_rate']:>5.0%} {r['count']:>4d}")

        if m.get('gates_corr') is not None:
            print(f"  Gates ρ: {m['gates_corr']:.3f}")
        fv = m.get('fv_metrics')
        if fv:
            print(f"  FV MAE: {fv['mae']:.1%}  Bias: {fv['signed_error']:+.1%}"
                  f"  ±20%: {fv['within_20']:.0%}")

    # --- Enhanced analytics (aggregated across all runs) ---
    print(f"\n{'='*70}")
    print("ENHANCED ANALYTICS")
    print(f"{'='*70}")

    # Strong BUY hit rate
    buy_hit = strong_buy_hit_rate(all_metrics)
    if buy_hit is not None:
        print(f"\n  BUY Hit Rate (beat SPY): {buy_hit:.0%}")

    # Sector accuracy
    sec_acc = sector_accuracy(all_metrics)
    if sec_acc:
        print(f"\n  {'Sector':<25s} {'Mean Ret':>9s} {'Alpha':>8s} {'Hit%':>6s} {'#':>4s}")
        print(f"  {'-'*54}")
        for sector in sorted(sec_acc, key=lambda s: sec_acc[s]['mean_alpha'], reverse=True):
            sa = sec_acc[sector]
            print(f"  {sector:<25s} {sa['mean_return']:>+8.1%} {sa['mean_alpha']:>+7.1%}"
                  f" {sa['hit_rate']:>5.0%} {sa['count']:>4d}")

    # MoS bucket accuracy
    mos_acc = mos_bucket_accuracy(all_metrics)
    if mos_acc:
        print(f"\n  {'MoS Bucket':<12s} {'Mean Ret':>9s} {'Hit%':>6s} {'#':>4s}")
        print(f"  {'-'*33}")
        for bucket in ['>30%', '10-30%', '0-10%', 'Negative']:
            if bucket in mos_acc:
                ma = mos_acc[bucket]
                print(f"  {bucket:<12s} {ma['mean_return']:>+8.1%}"
                      f" {ma['hit_rate']:>5.0%} {ma['count']:>4d}")

    # Consensus comparison
    consensus = consensus_comparison(all_metrics)
    if consensus:
        print(f"\n  Model vs Analyst Targets ({consensus['n_stocks']} stocks):")
        print(f"    Mean bias:   {consensus['mean_bias']:+.1%}")
        print(f"    Median bias: {consensus['median_bias']:+.1%}")


    # --- Signal quartile spreads (does the signal predict excess return?) ---
    print_signal_tables(all_metrics,
                        'SIGNAL QUARTILE SPREADS  (forward alpha by signal quartile)')


def print_signal_tables(all_metrics, title,
                        min_per_snapshot=MIN_SIGNAL_PER_SNAPSHOT):
    """Quartile table per SIGNAL_SPECS entry — shared by the unconditional
    print_summary block and the conditional cohort runs from the CLI."""
    print(f"\n  {'-'*60}")
    print(f"  {title}")
    print(f"  {'-'*60}")
    for key, label in SIGNAL_SPECS:
        sq = signal_quartile_accuracy(all_metrics, key,
                                      min_per_snapshot=min_per_snapshot)
        if sq is None:
            print(f"\n  {label}: not enough data yet "
                  f"(no snapshot has ≥{min_per_snapshot} signal "
                  f"values with a measured forward return)")
            continue
        print(f"\n  {label}  "
              f"({sq['n_snapshots']} snapshots, {sq['n_obs']} obs — "
              f"NOT independent)")
        print(f"    {'Quartile':<10s} {'Mean α':>9s} {'Median α':>10s}"
              f" {'Hit%':>6s} {'#':>6s}")
        for q in (1, 2, 3, 4):
            qd = sq['quartiles'].get(q)
            if not qd:
                continue
            tag = 'Q%d%s' % (q, ' (lo)' if q == 1 else ' (hi)' if q == 4 else '')
            print(f"    {tag:<10s} {qd['mean_alpha']:>+8.1%}"
                  f" {qd['median_alpha']:>+9.1%} {qd['hit_rate']:>5.0%}"
                  f" {qd['count']:>6d}")
        if sq['spread'] is not None:
            print(f"    Q4−Q1 spread: {sq['spread']:>+7.1%}"
                  f"   mean per-snapshot ρ: {sq['mean_rho']:+.3f}"
                  f"   ρ>0 in {sq['rho_pos_frac']:.0%} of snapshots")


# ===========================================================================
# CALIBRATE: Walk-forward parameter optimization
# ===========================================================================

# ---------------------------------------------------------------------------
# Window splitting
# ---------------------------------------------------------------------------

def _spaced_dates(dates, spacing_days):
    """Greedily thin a sorted date list so consecutive picks are >= spacing apart.

    Used to de-overlap walk-forward windows: two snapshots less than one
    horizon apart have forward-return windows covering almost the same market
    period, so they are not independent observations.
    """
    picked = []
    for d in sorted(dates):
        if not picked or (d - picked[-1]).days >= spacing_days:
            picked.append(d)
    return picked


def generate_windows(snapshot_dates, train_size=3, test_size=1, step=1):
    """Generate rolling train / test windows from sorted snapshot dates.

    Args:
        snapshot_dates: Sorted list of date objects with available snapshots.
        train_size: Number of snapshots in each training window.
        test_size: Number of snapshots in each test window.
        step: How many snapshots to advance between windows.

    Returns:
        List of ``{'train_dates': [...], 'test_dates': [...]}`` dicts.
        Empty list if there aren't enough snapshots.
    """
    total_needed = train_size + test_size
    if len(snapshot_dates) < total_needed:
        return []

    windows = []
    i = 0
    while i + total_needed <= len(snapshot_dates):
        train = snapshot_dates[i:i + train_size]
        test = snapshot_dates[i + train_size:i + total_needed]
        windows.append({'train_dates': train, 'test_dates': test})
        i += step
    return windows


# ---------------------------------------------------------------------------
# Objective functions
# ---------------------------------------------------------------------------

def compute_objective(backtest_metrics, objective='hit_rate'):
    """Compute a scalar objective from backtest metrics.

    Args:
        backtest_metrics: List of per-snapshot metric dicts from
            ``run_backtest()`` or compatible structure.
        objective: 'hit_rate' | 'alpha' | 'information_ratio' | 'composite'

    Returns:
        float: Objective value (higher is better).  Returns 0.0 if
        insufficient data.
    """
    fn = {
        'rank_ic': rank_ic_objective,
        'hit_rate': hit_rate_objective,
        'alpha': alpha_objective,
        'information_ratio': information_ratio_objective,
        'composite': composite_objective,
    }.get(objective, rank_ic_objective)
    return fn(backtest_metrics)


def rank_ic_objective(metrics):
    """Spearman rank IC between ``_composite_score`` and forward excess return.

    Computed across the FULL population (every row that has both a composite
    score and a forward excess return), not just BUY-rated names.  With only
    ~2 BUY per snapshot the BUY-only objectives are near-degenerate; rank IC
    uses ~2,000 rows/snapshot, so it actually discriminates between candidate
    parameter sets.  Higher = the model's score ranks future winners above
    losers.  Reuses the same Spearman formula as ``analyze_run``'s gates_corr.
    """
    scores, excess = [], []
    for m in metrics:
        for detail in m.get('details', []):
            s = detail.get('_composite_score')
            er = detail.get('excess_return')
            if (s is not None and er is not None
                    and math.isfinite(s) and math.isfinite(er)):
                scores.append(s)
                excess.append(er)
    n = len(scores)
    if n < 10:
        return 0.0
    rank_s = rank(scores)
    rank_e = rank(excess)
    d_sq = sum((a - b) ** 2 for a, b in zip(rank_s, rank_e))
    return 1 - (6 * d_sq) / (n * (n ** 2 - 1))


def hit_rate_objective(metrics):
    """Fraction of BUY-rated stocks that outperformed SPY.

    Aggregates across all snapshots in *metrics*.  Rows with no
    ``excess_return`` (i.e. snapshots not yet annotated with forward
    returns) are skipped — they contribute neither to numerator nor
    denominator.
    """
    beats = 0
    total = 0
    for m in metrics:
        for detail in m.get('details', []):
            if detail.get('rating') != 'BUY':
                continue
            er = detail.get('excess_return')
            if er is None:
                continue
            total += 1
            if er > 0:
                beats += 1
    return beats / total if total > 0 else 0.0


def alpha_objective(metrics):
    """Mean excess return of BUY-rated stocks across all snapshots."""
    excess = []
    for m in metrics:
        for detail in m.get('details', []):
            if detail.get('rating') == 'BUY':
                er = detail.get('excess_return')
                if er is not None:
                    excess.append(er)
    return float(np.mean(excess)) if excess else 0.0


def information_ratio_objective(metrics):
    """Alpha / tracking error for the BUY bucket."""
    excess = []
    for m in metrics:
        for detail in m.get('details', []):
            if detail.get('rating') == 'BUY':
                er = detail.get('excess_return')
                if er is not None:
                    excess.append(er)
    if len(excess) < 3:
        return 0.0
    mu = float(np.mean(excess))
    sigma = float(np.std(excess, ddof=1))
    return mu / sigma if sigma > 0 else 0.0


def composite_objective(metrics):
    """Blended objective: 0.4 * hit_rate + 0.3 * norm_alpha + 0.3 * fv_accuracy."""
    hr = hit_rate_objective(metrics)
    alpha = alpha_objective(metrics)
    # Normalise alpha to [0, 1] (cap at 10%)
    norm_alpha = max(0.0, min(1.0, alpha / 0.10))

    # Fair-value accuracy: fraction within ±20% of actual price
    within = 0
    fv_total = 0
    for m in metrics:
        for detail in m.get('details', []):
            fv = detail.get('dcf_fv')
            actual = detail.get('end_price')
            if fv and actual and actual > 0:
                fv_total += 1
                if abs(fv - actual) / actual <= 0.20:
                    within += 1
    fv_acc = within / fv_total if fv_total > 0 else 0.0

    return 0.4 * hr + 0.3 * norm_alpha + 0.3 * fv_acc


# ---------------------------------------------------------------------------
# Parameter search space
# ---------------------------------------------------------------------------

# Calibration re-scores snapshots with FROZEN fair values (no DCF re-run), so
# the objective can only respond to parameters consumed by the scoring layer:
# the score_weight_* category weights (growth derived as the residual,
# ownership held at its default). Sweeping anything else is a no-op — two
# candidates differing only in e.g. erp produce bit-identical objectives.
# The trimmed space is small enough to enumerate EXHAUSTIVELY (7x7x7 = 343
# raw combos, fewer after the growth>=0.05 constraint), so results carry no
# sampling-seed noise.
SEARCH_SPACE = {
    # Scoring weights (3 free; growth = 1 - sum - ownership_default)
    'score_weight_valuation': (0.15, 0.45, 0.05),
    'score_weight_quality':   (0.10, 0.40, 0.05),
    'score_weight_moat':      (0.10, 0.40, 0.05),
}

# Parameters that shape FAIR VALUES, not scores. Calibrating these requires a
# full valuation re-run per candidate (hours each), not the lightweight
# re-scoring above — kept here for reference so nobody re-adds them to
# SEARCH_SPACE and silently sweeps dead dimensions again.
# Rating thresholds ARE live under re-scoring (rating_from_composite reads
# them), but they only move rating-bucket objectives (hit_rate/alpha/...) —
# rank_ic correlates the raw composite score and cannot see them. They also
# multiply the grid ~12x, so they are opt-in via calibrate --include-thresholds
# rather than part of the weekly exhaustive sweep. (Ported from main's
# independently-developed calibrate.py.)
THRESHOLD_SPACE = {
    'rating_threshold_buy':  (52, 67, 5),
    'rating_threshold_lean': (34, 44, 5),
}

LIVE_RUN_ONLY_SPACE = {
    'erp': (0.04, 0.07, 0.005),
    'blend_trigger':    (1.2, 2.0, 0.1),
    'blend_dcf_weight': (0.40, 0.80, 0.05),
    'growth_weight_analyst_lt':  (0.15, 0.45, 0.05),
    'growth_weight_fundamental': (0.10, 0.35, 0.05),
    'analyst_haircut':          (0.60, 1.00, 0.05),
    'margin_trend_sensitivity': (0.0, 1.0, 0.25),
}


def _grid_axes(search_space):
    """Build per-parameter value lists for the search space.

    Returns (keys, ranges, total) where total = product of len(ranges[i]).
    Caller can iterate combinations on demand without materializing the
    full Cartesian product, which can easily exceed available RAM
    (the default SEARCH_SPACE has ~370M combos).
    """
    keys = sorted(search_space.keys())
    ranges = []
    total = 1
    for k in keys:
        lo, hi, step = search_space[k]
        vals = np.arange(lo, hi + step * 0.5, step)
        vals = [round(float(v), 6) for v in vals]
        ranges.append(vals)
        total *= len(vals)
    return keys, ranges, total


def _generate_grid(search_space):
    """Generate all combinations from the search space.

    WARNING: materializes the full Cartesian product.  For the default
    SEARCH_SPACE this is ~370M dicts (~370 GB peak memory).  Only safe
    for trimmed search spaces (used in tests).  Production sweeps go
    through grid_search() which samples indices via ``_grid_axes``.

    Returns:
        list[dict]: Every parameter combination as a dict.
    """
    keys, ranges, _ = _grid_axes(search_space)
    grid = []
    for combo in itertools.product(*ranges):
        d = dict(zip(keys, combo))
        grid.append(d)
    return grid


def _apply_derived_params(candidate):
    """Compute derived parameters and merge with defaults.

    - ``score_weight_growth`` = 1.0 - sum of other three scoring weights
      (only computed when at least one scoring weight is in the candidate)
    - ``blend_mult_weight`` = 1.0 - ``blend_dcf_weight``

    Returns:
        dict: Full ParamSet (defaults + candidate + derived), or None
        if constraints violated.
    """
    derived = dict(candidate)

    # Derive score_weight_growth only when scoring weights are being tuned.
    # Ownership weight is fixed at its default; growth absorbs the remainder.
    sw_keys = ('score_weight_valuation', 'score_weight_quality', 'score_weight_moat')
    if any(k in candidate for k in sw_keys):
        defs = default_params()
        sw_sum = (candidate.get('score_weight_valuation', defs['score_weight_valuation'])
                  + candidate.get('score_weight_quality', defs['score_weight_quality'])
                  + candidate.get('score_weight_moat', defs['score_weight_moat'])
                  + candidate.get('score_weight_ownership', defs['score_weight_ownership']))
        sw_growth = round(1.0 - sw_sum, 4)
        if sw_growth < 0.05:
            return None  # Constraint violation
        derived['score_weight_growth'] = sw_growth

    # Derive blend_mult_weight
    if 'blend_dcf_weight' in derived:
        derived['blend_mult_weight'] = round(1.0 - derived['blend_dcf_weight'], 4)

    try:
        params = merge_params(derived)
    except ValueError:
        return None

    errors = validate_params(params)
    if errors:
        return None
    return params


def _sample_grid(full_grid, n, seed=42):
    """Stratified random sampling from a materialized grid.

    Uses numpy to select *n* samples that are approximately
    evenly distributed across the grid.  Only used when the caller has
    already chosen to materialize the full grid; grid_search() prefers
    ``_sample_grid_from_space`` to avoid the materialization.
    """
    rng = np.random.default_rng(seed)
    if n >= len(full_grid):
        return full_grid
    indices = rng.choice(len(full_grid), size=n, replace=False)
    return [full_grid[i] for i in sorted(indices)]


def _sample_grid_from_space(search_space, n, seed=42):
    """Sample *n* parameter combos directly from the search space without
    materializing the full Cartesian product.

    For a default SEARCH_SPACE with ~370M combos, this uses ~n * 1KB of
    memory instead of ~370 GB.  Picks *n* random integer indices from
    [0, total) without replacement, then decodes each to a combo dict.
    """
    keys, ranges, total = _grid_axes(search_space)
    rng = np.random.default_rng(seed)
    if n >= total:
        # Tiny grid — just enumerate everything
        return [dict(zip(keys, combo)) for combo in itertools.product(*ranges)]
    indices = sorted(rng.choice(total, size=n, replace=False).tolist())
    sizes = [len(r) for r in ranges]
    combos = []
    for idx in indices:
        combo = []
        for r, s in zip(ranges, sizes):
            combo.append(r[idx % s])
            idx //= s
        combos.append(dict(zip(keys, combo)))
    return combos


def grid_search(evaluate_fn, search_space=None, max_evaluations=500):
    """Search over parameter space to maximise the objective.

    If the full grid exceeds *max_evaluations*, samples a subset
    *without materializing the full Cartesian product* — the default
    SEARCH_SPACE has ~370M combos, which would exhaust system memory.

    Args:
        evaluate_fn: Callable(params_dict) -> float (objective value).
        search_space: ``{name: (min, max, step)}``.  Defaults to
            module-level ``SEARCH_SPACE``.
        max_evaluations: Cap on evaluations.

    Returns:
        List of ``{'params': dict, 'objective': float}`` sorted by
        objective descending.
    """
    if search_space is None:
        search_space = SEARCH_SPACE

    raw_grid = _sample_grid_from_space(search_space, max_evaluations)

    results = []
    for candidate in raw_grid:
        params = _apply_derived_params(candidate)
        if params is None:
            continue
        obj = evaluate_fn(params)
        results.append({'params': params, 'objective': obj})

    results.sort(key=lambda x: x['objective'], reverse=True)
    return results


# ---------------------------------------------------------------------------
# Overfitting prevention
# ---------------------------------------------------------------------------

def regularized_objective(base_obj, params, lambda_reg=0.05):
    """Penalise large deviations from default parameter values.

    Args:
        base_obj: Raw objective value (higher is better).
        params: Candidate ParamSet dict.
        lambda_reg: Regularisation strength.

    Returns:
        float: Penalised objective.
    """
    defaults = default_params()
    deviation = 0.0
    for k, v in params.items():
        dv = defaults.get(k)
        if dv is not None and isinstance(v, (int, float)) and isinstance(dv, (int, float)):
            # Normalise by default value to make deviations comparable
            denom = abs(dv) if dv != 0 else 1.0
            deviation += ((v - dv) / denom) ** 2
    return base_obj - lambda_reg * deviation


def _calibrated_weights(params):
    """Reduce a full ParamSet to the block calibration actually swept.

    Only the five score_weight_* values respond to the frozen-fair-value
    re-scoring; reporting the rest as 'recommended' would present untouched
    defaults (or sampler noise, pre-trim) as calibrated output.
    """
    keys = list(SEARCH_SPACE) + ['score_weight_growth', 'score_weight_ownership']
    return {k: params[k] for k in sorted(keys) if k in params}


def compute_stability(window_results):
    """Measure how much each parameter varies across windows.

    Args:
        window_results: List of window dicts with 'best_params'.

    Returns:
        dict: ``{param_name: std_dev}`` for each parameter.
    """
    if not window_results:
        return {}

    all_params = [w['best_params'] for w in window_results if 'best_params' in w]
    if not all_params:
        return {}

    stability = {}
    for key in all_params[0]:
        vals = [p[key] for p in all_params if isinstance(p.get(key), (int, float))]
        if len(vals) >= 2:
            stability[key] = float(np.std(vals, ddof=1))
    return stability


# ---------------------------------------------------------------------------
# Walk-forward calibration loop
# ---------------------------------------------------------------------------

def walk_forward_calibrate(results_dir='output', horizons=None,
                           train_size=3, test_size=1, step=1,
                           objective='rank_ic', max_evaluations=400,
                           lambda_reg=0.05, yf_client=None,
                           prices_dir='output/prices', cache_dir='output/returns',
                           today=None,
                           search_space=None):
    """Run the full walk-forward calibration.

    Pipeline:
      1. Census snapshot maturity per horizon; refuse loudly if a horizon has
         zero matured snapshots (forward returns can't exist yet).
      2. Load every snapshot once and annotate forward returns once (cache-aware,
         shared by reference across windows) — returns are param-independent.
      3. For each rolling window whose TEST date is matured:
         a. sweep candidate ParamSets on the train window
         b. select the best, measure it out-of-sample on the test window

    Args:
        results_dir: Directory containing results_YYYY-MM-DD.json files.
        horizons: Forward-return horizons in days.  Defaults to [30] (the only
            matured horizon for current data; 90/180 mature later in 2026).
        train_size/test_size/step: Walk-forward window geometry.
        objective: Objective function name (default 'rank_ic').
        max_evaluations: Max parameter combinations per window.
        lambda_reg: Regularisation strength (0 = disabled).
        yf_client: Optional YFinanceClient (constructed lazily on cache miss).
        prices_dir: Local parquet price dir for forward-return fetches.
        cache_dir: Sidecar dir for cached forward returns.
        today: Override "today" (for testing maturity logic).

    Returns:
        dict with 'windows', 'overall', 'recommended_params', 'matured_counts'.
    """
    if horizons is None:
        horizons = [30]
    today = today or date.today()

    snapshot_dates = _discover_snapshot_dates(results_dir)

    # 1. Per-horizon maturity census (loud guard — never silently score 0.0).
    matured_counts = {
        h: sum(1 for d in snapshot_dates if is_matured(d.isoformat(), h, today))
        for h in horizons
    }
    usable_horizons = [h for h in horizons if matured_counts[h] > 0]
    if not usable_horizons:
        earliest = snapshot_dates[0].isoformat() if snapshot_dates else 'none'
        msg = (f"No matured snapshots for horizon(s) {horizons} as of {today}. "
               f"A (snapshot, horizon) pair matures when snapshot_date + horizon "
               f"<= today; earliest snapshot is {earliest}. "
               f"Nothing to calibrate — refusing to emit a 0.0 result.")
        print(f"[calibrate] {msg}")
        return {
            'date': today.isoformat(),
            'objective': objective,
            'horizon_days': horizons,
            'matured_counts': matured_counts,
            'n_windows': 0,
            'windows': [],
            'overall': {'error': msg},
            'recommended_params': _calibrated_weights(default_params()),
        }
    if set(usable_horizons) != set(horizons):
        dropped = [h for h in horizons if h not in usable_horizons]
        print(f"[calibrate] dropping immature horizon(s) {dropped} "
              f"(0 matured snapshots); using {usable_horizons}")
    horizons = usable_horizons

    # Window spacing: daily snapshots produce return windows that overlap by
    # (horizon - 1) days, so adjacent walk-forward "test" sets are ~99% the
    # same market period as their train sets — not independent evidence. Prefer
    # snapshot dates spaced >= the horizon apart (true de-overlap). Until
    # enough spaced matured snapshots exist, fall back to the overlapped set
    # with a LOUD warning and report the effective independent sample size
    # (span / horizon) so nobody mistakes 12 overlapping windows for 12
    # observations.
    spacing = max(horizons)
    spaced_dates = _spaced_dates(snapshot_dates, spacing)
    spaced_matured = [d for d in spaced_dates
                      if any(is_matured(d.isoformat(), h, today) for h in horizons)]
    overlapped_fallback = len(spaced_matured) < (train_size + test_size)
    if overlapped_fallback:
        dates_for_windows = snapshot_dates
        matured_all = [d for d in snapshot_dates
                       if any(is_matured(d.isoformat(), h, today) for h in horizons)]
        span_days = (max(matured_all) - min(matured_all)).days if matured_all else 0
        effective_n = max(1, span_days // spacing + 1)
        print(f"[calibrate] WARNING: only {len(spaced_matured)} matured snapshots "
              f"spaced >= {spacing}d apart (need {train_size + test_size}). Falling "
              f"back to OVERLAPPED daily windows — test sets share most of their "
              f"return period with training data. Effective independent sample "
              f"size ~{effective_n}, regardless of window count.")
    else:
        dates_for_windows = spaced_dates
        effective_n = len(spaced_matured)
        print(f"[calibrate] using {len(spaced_dates)} snapshot dates spaced "
              f">= {spacing}d apart ({len(spaced_matured)} matured) — "
              f"de-overlapped walk-forward.")

    windows = generate_windows(dates_for_windows, train_size, test_size, step)
    if not windows:
        return {
            'date': today.isoformat(),
            'objective': objective,
            'horizon_days': horizons,
            'matured_counts': matured_counts,
            'n_windows': 0,
            'windows': [],
            'overall': {'error': 'Insufficient snapshots for walk-forward'},
            'recommended_params': _calibrated_weights(default_params()),
        }

    # 2. Load every snapshot once, share by reference, annotate returns once.
    if yf_client is None:
        from data.yfinance_client import YFinanceClient
        yf_client = YFinanceClient()
    pdir = prices_dir if (prices_dir and os.path.isdir(prices_dir)) else None

    snaps_by_date = {}
    for d in snapshot_dates:
        loaded = _load_snapshots(results_dir, [d])
        if loaded:
            snaps_by_date[d] = loaded[0]
    annotated_rows = 0
    for snap in snaps_by_date.values():
        res = annotate_snapshot_returns(snap, horizons, yf_client,
                                        prices_dir=pdir, cache_dir=cache_dir,
                                        today=today)
        annotated_rows += sum(n for n in res.values() if n)

    # Coverage guard: dates can be matured while the PRICE data doesn't reach
    # the eval date (e.g. stale parquet). Without usable forward returns every
    # objective is 0.0 and "calibration" is meaningless — refuse loudly rather
    # than emit a noise result.
    if annotated_rows == 0:
        msg = (f"0 usable forward-return rows for horizon(s) {horizons}. "
               f"{matured_counts} snapshots matured by date, but no price series "
               f"reaches the eval date — local prices are likely stale "
               f"(refresh output/prices). Refusing to emit a 0.0 result.")
        print(f"[calibrate] {msg}")
        return {
            'date': today.isoformat(),
            'objective': objective,
            'horizon_days': horizons,
            'matured_counts': matured_counts,
            'annotated_rows': 0,
            'n_windows': 0,
            'windows': [],
            'overall': {'error': msg},
            'recommended_params': _calibrated_weights(default_params()),
        }

    # 3. Walk forward over windows with a matured test date.
    window_results = []
    skipped_immature = 0
    for win in windows:
        test_matured = any(
            is_matured(td.isoformat(), h, today)
            for td in win['test_dates'] for h in horizons
        )
        if not test_matured:
            skipped_immature += 1
            continue

        train_results = [snaps_by_date[d] for d in win['train_dates']
                         if d in snaps_by_date]
        test_results = [snaps_by_date[d] for d in win['test_dates']
                        if d in snaps_by_date]

        def evaluate(params, _train=train_results):
            metrics = _evaluate_params_on_snapshots(_train, params, horizons)
            raw = compute_objective(metrics, objective)
            if lambda_reg > 0:
                return regularized_objective(raw, params, lambda_reg)
            return raw

        search_results = grid_search(evaluate, search_space=search_space,
                                     max_evaluations=max_evaluations)
        if not search_results:
            continue

        best = search_results[0]
        best_params = best['params']
        train_obj = best['objective']

        test_metrics = _evaluate_params_on_snapshots(test_results, best_params,
                                                     horizons)
        test_obj = compute_objective(test_metrics, objective)

        window_results.append({
            'train_dates': [d.isoformat() for d in win['train_dates']],
            'test_dates': [d.isoformat() for d in win['test_dates']],
            # Serialize only the swept weight block — the full ParamSet would
            # imply the other ~30 params were calibrated when they were not.
            'best_params': _calibrated_weights(best_params),
            'train_objective': round(train_obj, 4),
            'test_objective': round(test_obj, 4),
        })

    # Aggregate results
    stability = compute_stability(window_results)
    train_objs = [w['train_objective'] for w in window_results]
    test_objs = [w['test_objective'] for w in window_results]

    # Recommended params: majority vote across windows (the set selected most
    # often), tie-broken by mean test objective. Never max-test — picking the
    # single window whose params scored best out-of-sample is selection ON the
    # test data and rewards noise.
    if window_results:
        from collections import Counter, defaultdict as _dd
        key_of = lambda p: tuple(sorted(p.items()))
        votes = Counter(key_of(w['best_params']) for w in window_results)
        test_by_key = _dd(list)
        for w in window_results:
            test_by_key[key_of(w['best_params'])].append(w['test_objective'])
        best_key = max(votes, key=lambda k: (votes[k],
                                             float(np.mean(test_by_key[k]))))
        recommended = dict(best_key)
        recommendation_basis = {
            'rule': 'majority vote across windows, tie-break mean test objective',
            'windows_won': votes[best_key],
            'windows_total': len(window_results),
            'mean_test_objective_of_winner': round(float(np.mean(test_by_key[best_key])), 4),
        }
    else:
        recommended = _calibrated_weights(default_params())
        recommendation_basis = {'rule': 'defaults (no windows)'}

    overall = {
        'mean_train_objective': round(float(np.mean(train_objs)), 4) if train_objs else None,
        'mean_test_objective': round(float(np.mean(test_objs)), 4) if test_objs else None,
        'overfit_gap': round(
            float(np.mean(train_objs)) - float(np.mean(test_objs)), 4
        ) if train_objs and test_objs else None,
        'param_stability': {k: round(v, 6) for k, v in stability.items()},
    }

    return {
        'date': today.isoformat(),
        'objective': objective,
        'horizon_days': horizons,
        'matured_counts': matured_counts,
        'annotated_rows': annotated_rows,
        'window_spacing_days': spacing,
        'overlapped_fallback': overlapped_fallback,
        'effective_independent_obs': effective_n,
        'n_windows': len(window_results),
        'n_windows_skipped_immature': skipped_immature,
        'n_evaluations_per_window': max_evaluations,
        'calibrated_keys': sorted(SEARCH_SPACE) + ['score_weight_growth',
                                                   'score_weight_ownership'],
        'windows': window_results,
        'overall': overall,
        'recommended_params': recommended,
        'recommendation_basis': recommendation_basis,
    }


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _discover_snapshot_dates(results_dir):
    """Find all results_YYYY-MM-DD.json files and return sorted dates."""
    dates = []
    if not os.path.isdir(results_dir):
        return dates
    for fname in os.listdir(results_dir):
        if fname.startswith('results_') and fname.endswith('.json'):
            date_str = fname.replace('results_', '').replace('.json', '')
            try:
                dates.append(date.fromisoformat(date_str))
            except ValueError:
                continue
    return sorted(dates)


def _load_snapshots(results_dir, dates):
    """Load results JSON files for the given dates.

    Returns:
        list[dict]: Each dict is the full JSON structure with 'date' and 'results'.
    """
    snapshots = []
    for d in dates:
        path = os.path.join(results_dir, f'results_{d.isoformat()}.json')
        if os.path.exists(path):
            with open(path) as f:
                snapshots.append(json.load(f))
    return snapshots


def _evaluate_params_on_snapshots(snapshots, params, horizons):
    """Re-score snapshot results with *params* and compute backtest metrics.

    This performs a lightweight re-scoring (updating composite scores and
    ratings based on new category weights) rather than a full DCF re-run.
    The fair values from the original snapshot are preserved; only the
    scoring/rating layer changes.

    Mutates each snapshot's ``results`` rows in place (no deepcopy).  The
    original rating is cached on each row as ``_rating_orig`` and restored
    before every rescore so that ``apply_composite_rating_override`` (which
    only downgrades) starts from a clean baseline on each candidate.  This
    is the hot path of walk_forward_calibrate: a deepcopy here costs ~50ms
    per snapshot and ~3GB peak RSS across a typical sweep.

    Returns:
        list[dict]: Backtest-compatible metric dicts, one per snapshot×horizon.
    """
    from scripts.scoring import (compute_continuous_scores,
                                 apply_composite_rating_override)

    metrics = []
    for snap in snapshots:
        results = snap.get('results', [])
        run_date = snap.get('date', '')

        # Cache original rating once per row, then reset before rescoring so
        # downgrades from a prior candidate don't bleed into this one.
        for r in results:
            if '_rating_orig' not in r:
                r['_rating_orig'] = r.get('rating')
            r['rating'] = r['_rating_orig']

        # Re-score with candidate params (only changes composite weights)
        compute_continuous_scores(results, params=params)
        apply_composite_rating_override(results, params=params)

        # Build a lightweight metric dict compatible with objective functions.
        # Forward returns are horizon-keyed under r['_fwd'][h], pre-computed
        # once by annotate_snapshot_returns (None when the horizon is immature
        # or the ticker had no price data).
        for h in horizons:
            details = []
            for r in results:
                fwd = (r.get('_fwd') or {}).get(h)
                details.append({
                    'ticker': r.get('ticker'),
                    'rating': r.get('rating'),
                    'dcf_fv': r.get('dcf_fv'),
                    'price': r.get('price'),
                    'mos': r.get('mos'),
                    '_composite_score': r.get('_composite_score'),
                    'excess_return': fwd['excess_return'] if fwd else None,
                    'end_price': fwd['end_price'] if fwd else None,
                })
            metrics.append({
                'run_date': run_date,
                'horizon': h,
                'details': details,
            })
    return metrics


# ---------------------------------------------------------------------------
# Weight calibration via Cohen's d
# ---------------------------------------------------------------------------

def _cohens_d(group_a, group_b):
    """Compute Cohen's d effect size between two score lists."""
    if len(group_a) < 2 or len(group_b) < 2:
        return 0.0
    mean_a = sum(group_a) / len(group_a)
    mean_b = sum(group_b) / len(group_b)
    var_a = sum((x - mean_a) ** 2 for x in group_a) / (len(group_a) - 1)
    var_b = sum((x - mean_b) ** 2 for x in group_b) / (len(group_b) - 1)
    pooled_sd = math.sqrt((var_a + var_b) / 2)
    return (mean_a - mean_b) / pooled_sd if pooled_sd > 0 else 0.0


def optimize_weights(results_json_path, output_path=None):
    """Grid search over category weight combinations to maximize Cohen's d.

    Loads a results JSON, generates all 4-weight combos that sum to 1.0,
    re-scores each combo, and computes Cohen's d between quality/poor groups.

    Args:
        results_json_path: Path to results_YYYY-MM-DD.json.
        output_path: Optional path to write calibration results JSON.

    Returns:
        dict with best weights, Cohen's d, and top-10 results.
    """
    from scripts.scoring import compute_continuous_scores, apply_composite_rating_override
    import copy

    with open(results_json_path) as f:
        data = json.load(f)
    all_results = data.get('results', data) if isinstance(data, dict) else data

    # Check we have quality/poor labels
    quality = [r for r in all_results if r.get('source_group') == 'quality']
    poor = [r for r in all_results if r.get('source_group') == 'poor']
    if len(quality) < 3 or len(poor) < 3:
        print(f"Insufficient labelled data: {len(quality)} quality, {len(poor)} poor")
        return None

    # Generate weight grid: all combos of 4 weights from step=0.05 that sum to 1.0
    steps = [round(v, 2) for v in np.arange(0.10, 0.40, 0.05)]
    grid = []
    for wv in steps:
        for wq in steps:
            for wm in steps:
                wg = round(1.0 - wv - wq - wm, 2)
                if 0.05 <= wg <= 0.40:
                    grid.append((wv, wq, wm, wg))

    print(f"Optimizing weights: {len(grid)} combos, "
          f"{len(quality)} quality, {len(poor)} poor stocks")

    # Compute baseline Cohen's d
    baseline_results = copy.deepcopy(all_results)
    compute_continuous_scores(baseline_results)
    apply_composite_rating_override(baseline_results)
    q_baseline = [r['_composite_score'] for r in baseline_results
                  if r.get('source_group') == 'quality' and r.get('_composite_score') is not None]
    p_baseline = [r['_composite_score'] for r in baseline_results
                  if r.get('source_group') == 'poor' and r.get('_composite_score') is not None]
    baseline_d = _cohens_d(q_baseline, p_baseline)
    print(f"Baseline Cohen's d: {baseline_d:.3f} "
          f"(weights: 0.30/0.25/0.25/0.20)")

    results_list = []
    for wv, wq, wm, wg in grid:
        trial = copy.deepcopy(all_results)
        params = {
            'score_weight_valuation': wv,
            'score_weight_quality': wq,
            'score_weight_moat': wm,
            'score_weight_growth': wg,
        }
        compute_continuous_scores(trial, params=params)
        apply_composite_rating_override(trial, params=params)

        q_scores = [r['_composite_score'] for r in trial
                    if r.get('source_group') == 'quality' and r.get('_composite_score') is not None]
        p_scores = [r['_composite_score'] for r in trial
                    if r.get('source_group') == 'poor' and r.get('_composite_score') is not None]
        d = _cohens_d(q_scores, p_scores)
        q_mean = sum(q_scores) / len(q_scores) if q_scores else 0
        p_mean = sum(p_scores) / len(p_scores) if p_scores else 0

        results_list.append({
            'weights': {'valuation': wv, 'quality': wq, 'moat': wm, 'growth': wg},
            'cohens_d': round(d, 4),
            'quality_mean': round(q_mean, 1),
            'poor_mean': round(p_mean, 1),
        })

    results_list.sort(key=lambda x: x['cohens_d'], reverse=True)
    best = results_list[0]

    print(f"\n--- Best Weights ---")
    w = best['weights']
    print(f"  Valuation: {w['valuation']:.0%}  Quality: {w['quality']:.0%}  "
          f"Moat: {w['moat']:.0%}  Growth: {w['growth']:.0%}")
    print(f"  Cohen's d: {best['cohens_d']:.3f} (baseline: {baseline_d:.3f})")
    print(f"  Quality mean: {best['quality_mean']}  Poor mean: {best['poor_mean']}")

    print(f"\n--- Top 10 ---")
    for i, r in enumerate(results_list[:10]):
        w = r['weights']
        print(f"  {i+1}. V={w['valuation']:.0%} Q={w['quality']:.0%} "
              f"M={w['moat']:.0%} G={w['growth']:.0%}  "
              f"d={r['cohens_d']:.3f}  q={r['quality_mean']} p={r['poor_mean']}")

    output = {
        'date': date.today().isoformat(),
        'baseline_cohens_d': round(baseline_d, 4),
        'best': best,
        'top_10': results_list[:10],
        'n_quality': len(quality),
        'n_poor': len(poor),
        'grid_size': len(grid),
    }

    if output_path:
        with open(output_path, 'w') as f:
            json.dump(output, f, indent=2)
        print(f"\nResults written to {output_path}")

    return output


# ===========================================================================
# CLI: Subcommand dispatch
# ===========================================================================

def _cli_measure(args):
    """Run backtest analysis on accumulated snapshots."""
    # Override the evaluated-signal set if requested. Labels default to the
    # raw field name for ad-hoc signals not in SIGNAL_SPECS.
    if getattr(args, 'signals', None):
        _known = dict(SIGNAL_SPECS)
        SIGNAL_SPECS[:] = [(k.strip(), _known.get(k.strip(), k.strip()))
                           for k in args.signals.split(',') if k.strip()]

    horizons = [int(h.strip()) for h in args.horizons.split(',')]

    from data.yfinance_client import YFinanceClient
    yf_client = YFinanceClient()

    prices_dir = args.prices_dir if os.path.isdir(args.prices_dir) else None
    if prices_dir:
        print(f"Using local price files from: {prices_dir}")

    all_metrics = run_backtest(args.results_dir, horizons, yf_client,
                               prices_dir=prices_dir)

    if all_metrics:
        print_summary(all_metrics)

        if getattr(args, 'cohort', None):
            coh = filter_metrics_to_cohort(all_metrics, cohort_key=args.cohort)
            print_signal_tables(
                coh,
                f'COHORT SPREADS — within top-quartile {args.cohort}',
                min_per_snapshot=10)
            if args.exclude_capped:
                coh2 = filter_metrics_to_cohort(all_metrics,
                                                cohort_key=args.cohort,
                                                exclude_capped=True)
                print_signal_tables(
                    coh2,
                    f'COHORT SPREADS — top-quartile {args.cohort}, '
                    'EXCL. beneish/altman-capped',
                    min_per_snapshot=10)

        os.makedirs('output', exist_ok=True)
        xlsx = os.path.join('output', f'backtest_{date.today().isoformat()}.xlsx')
        build_backtest_excel(all_metrics, xlsx)
        print(f"\nExcel: {xlsx}")
    else:
        print("\nNo backtest results to report yet.")
        print("As you accumulate snapshots over time, the backtest will activate.")


def _cli_calibrate(args):
    """Run walk-forward parameter calibration."""
    horizons = [int(h) for h in args.horizons.split(',')]

    space = None
    if getattr(args, 'include_thresholds', False):
        space = dict(SEARCH_SPACE)
        space.update(THRESHOLD_SPACE)
        if args.objective == 'rank_ic':
            print("[calibrate] NOTE: rating thresholds cannot affect rank_ic "
                  "(it reads the raw composite score); use a rating-bucket "
                  "objective (hit_rate/alpha/composite) to calibrate them.")

    result = walk_forward_calibrate(
        results_dir=args.results_dir,
        horizons=horizons,
        train_size=args.train_size,
        test_size=args.test_size,
        objective=args.objective,
        max_evaluations=args.max_evals,
        lambda_reg=args.lambda_reg,
        prices_dir=args.prices_dir,
        search_space=space,
    )

    overall = result.get('overall', {})
    # Refuse to write a meaningless 0.0 result when no horizon is matured.
    if result.get('n_windows', 0) == 0:
        print(f"\nNo calibration performed: {overall.get('error', 'no usable windows')}")
        print(f"Matured snapshots by horizon: {result.get('matured_counts')}")
        return

    # Default filename embeds the horizon so per-horizon runs don't clobber
    # each other (the weekly job calibrates 30d and 90d separately).
    hz_tag = '-'.join(str(h) for h in result.get('horizon_days', horizons))
    out_path = args.output or os.path.join(
        args.results_dir,
        f'calibration_h{hz_tag}_{date.today().isoformat()}.json')
    with open(out_path, 'w') as f:
        json.dump(result, f, indent=2, default=str)
    print(f'Calibration results written to {out_path}')

    print(f"\nObjective:            {result.get('objective')}")
    print(f"Horizon(s):           {result.get('horizon_days')}")
    print(f"Matured by horizon:   {result.get('matured_counts')}")
    print(f"Window mode:          "
          + ("OVERLAPPED fallback — effective independent obs "
             f"~{result.get('effective_independent_obs')}"
             if result.get('overlapped_fallback')
             else f"de-overlapped (spacing {result.get('window_spacing_days')}d)"))
    print(f"Windows:              {result['n_windows']}"
          f" (skipped immature: {result.get('n_windows_skipped_immature', 0)})")
    print(f"Mean train objective: {overall.get('mean_train_objective')}")
    print(f"Mean test objective:  {overall.get('mean_test_objective')}")
    print(f"Overfit gap:          {overall.get('overfit_gap')}")
    rb = result.get('recommendation_basis', {})
    print(f"Recommended weights:  {result.get('recommended_params')}")
    print(f"  basis: {rb.get('rule')} "
          f"(won {rb.get('windows_won')}/{rb.get('windows_total')} windows, "
          f"mean test {rb.get('mean_test_objective_of_winner')})")


def _cli_annotate(args):
    """Warm the forward-return sidecar cache for all snapshots x horizons."""
    from data.yfinance_client import YFinanceClient
    horizons = [int(h) for h in args.horizons.split(',')]
    yf_client = YFinanceClient()
    pdir = args.prices_dir if os.path.isdir(args.prices_dir) else None
    if pdir:
        print(f"Using local price files from: {pdir}")

    dates = _discover_snapshot_dates(args.results_dir)
    today = date.today()
    matured_counts = {
        h: sum(1 for d in dates if is_matured(d.isoformat(), h, today))
        for h in horizons
    }
    print(f"Snapshots: {len(dates)}   Matured by horizon: {matured_counts}")
    for h in horizons:
        if matured_counts[h] == 0:
            print(f"[annotate] horizon {h}d: 0 matured snapshots — skipping "
                  f"(nothing to compute yet).")

    total = 0
    for d in dates:
        loaded = _load_snapshots(args.results_dir, [d])
        if not loaded:
            continue
        res = annotate_snapshot_returns(loaded[0], horizons, yf_client,
                                        prices_dir=pdir,
                                        cache_dir=args.cache_dir, today=today)
        done = {h: n for h, n in res.items() if n is not None}
        if done:
            total += len(done)
            print(f"  {d.isoformat()}: annotated {done}")
    print(f"\nCache warmed under {args.cache_dir} ({total} matured (date,horizon) pairs).")


def _cli_optimize_weights(args):
    """Run Cohen's d weight optimization on a single snapshot."""
    out = args.output or os.path.join(
        os.path.dirname(args.snapshot) or '.',
        f'weight_calibration_{date.today().isoformat()}.json')
    optimize_weights(args.snapshot, output_path=out)


if __name__ == '__main__':
    import argparse

    # Legacy-CLI shim: before the subcommand split, this script was invoked as
    # `python scripts/backtest.py --results-dir ...`. Keep those invocations
    # working by defaulting to the `measure` subcommand.
    _subcommands = {'measure', 'calibrate', 'annotate', 'optimize-weights', '-h', '--help'}
    if len(sys.argv) == 1 or sys.argv[1] not in _subcommands:
        if len(sys.argv) > 1 and not sys.argv[1].startswith('-'):
            pass  # unknown positional — let argparse report it
        else:
            print("[backtest] no subcommand given — defaulting to 'measure' "
                  "(legacy CLI compatibility)")
            sys.argv.insert(1, 'measure')

    parser = argparse.ArgumentParser(
        description='Backtest + calibrate the stock model. Subcommands share '
                    'the same results_*.json snapshots and objective functions.')
    sub = parser.add_subparsers(dest='command', required=True,
                                metavar='{measure,calibrate,annotate,optimize-weights}')

    # ---- measure ----
    p_measure = sub.add_parser(
        'measure',
        help='Run backtest analysis on accumulated snapshots.',
        description='Loads results_*.json snapshots, computes forward returns '
                    'over the given horizons, and writes a styled Excel report.')
    p_measure.add_argument('--results-dir', default='output',
                           help='Directory containing results_*.json files')
    p_measure.add_argument('--horizons', default='30,90,180',
                           help='Comma-separated horizon days (default: 30,90,180)')
    p_measure.add_argument('--prices-dir', default='output/prices',
                           help='Directory of per-ticker Parquet price files '
                                '(default: output/prices)')
    p_measure.add_argument('--signals', default=None,
                           help='Comma-separated signal field names to evaluate '
                                'by quartile (default: SIGNAL_SPECS). Any field '
                                'present on the stock rows works.')
    p_measure.add_argument('--cohort', default=None, metavar='FIELD',
                           help='Also evaluate each signal WITHIN the per-snapshot '
                                'top quartile of this field (e.g. --cohort mos).')
    p_measure.add_argument('--exclude-capped', action='store_true',
                           help='With --cohort: drop rows already vetoed by the '
                                'beneish/altman-distress rating caps.')
    p_measure.set_defaults(func=_cli_measure)

    # ---- calibrate ----
    p_cal = sub.add_parser(
        'calibrate',
        help='Walk-forward parameter calibration over rolling train/test windows.',
        description='Splits snapshots into rolling train/test windows; for each '
                    'window, sweeps the parameter grid using the chosen objective '
                    'and records out-of-sample performance.')
    p_cal.add_argument('--results-dir', default='output',
                       help='Directory with results_*.json files')
    p_cal.add_argument('--horizons', default='30',
                       help='Comma-separated forward return horizons in days '
                            '(default: 30 — the only matured horizon for current data)')
    p_cal.add_argument('--train-size', type=int, default=3)
    p_cal.add_argument('--test-size', type=int, default=1)
    p_cal.add_argument('--objective', default='rank_ic',
                       choices=['rank_ic', 'hit_rate', 'alpha',
                                'information_ratio', 'composite'],
                       help='Objective to maximise (default: rank_ic — '
                            'full-population Spearman IC, robust to thin BUY counts)')
    p_cal.add_argument('--max-evals', type=int, default=400,
                       help='Cap on parameter combos per window. The trimmed '
                            'score-weight grid has ~343 raw combos, so the '
                            'default enumerates it exhaustively (no sampling).')
    p_cal.add_argument('--lambda-reg', type=float, default=0.05)
    p_cal.add_argument('--prices-dir', default='output/prices',
                       help='Directory of per-ticker Parquet price files '
                            '(default: output/prices)')
    p_cal.add_argument('--include-thresholds', action='store_true',
                       help='Also sweep rating_threshold_buy/lean (grid ~12x '
                            'larger, sampled above --max-evals). Only rating-'
                            'bucket objectives can see thresholds.')
    p_cal.add_argument('--output', default=None,
                       help='Output JSON path (default: output/calibration_DATE.json)')
    p_cal.set_defaults(func=_cli_calibrate)

    # ---- annotate ----
    p_ann = sub.add_parser(
        'annotate',
        help='Warm the forward-return sidecar cache (no calibration).',
        description='Computes forward returns for every matured (snapshot, '
                    'horizon) pair and caches them under output/returns/, so '
                    'calibrate/measure runs reuse them without refetching.')
    p_ann.add_argument('--results-dir', default='output',
                       help='Directory with results_*.json files')
    p_ann.add_argument('--horizons', default='30',
                       help='Comma-separated horizon days (default: 30)')
    p_ann.add_argument('--prices-dir', default='output/prices',
                       help='Directory of per-ticker Parquet price files')
    p_ann.add_argument('--cache-dir', default='output/returns',
                       help='Sidecar cache directory (default: output/returns)')
    p_ann.set_defaults(func=_cli_annotate)

    # ---- optimize-weights ----
    p_opt = sub.add_parser(
        'optimize-weights',
        help="Cohen's d weight optimization on a single snapshot.",
        description='Searches the 4-weight valuation/quality/moat/growth grid '
                    "for the combo that best separates quality vs poor source "
                    'groups in a single snapshot.')
    p_opt.add_argument('snapshot',
                       help='Path to a results_YYYY-MM-DD.json snapshot file')
    p_opt.add_argument('--output', default=None,
                       help='Output JSON path '
                            '(default: weight_calibration_DATE.json next to the snapshot)')
    p_opt.set_defaults(func=_cli_optimize_weights)

    args = parser.parse_args()
    args.func(args)
